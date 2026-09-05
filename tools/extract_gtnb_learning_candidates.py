#!/usr/bin/env python3
"""Extract compact, unreviewed GTNB measured-learning authoring candidates.

Only injection-moulding records are selected using the source-documented I/INY machine
prefix rule already used by the governed benchmark. Numeric extraction deliberately
matches the established benchmark reader: direct cell values and cached formula results
stored in the exact publisher workbook are eligible, while formula-backed cells are
counted explicitly so source-derived quality metrics are never presented as raw sensors.
"""
from __future__ import annotations

import hashlib
import json
import math
import re
import tempfile
from collections import defaultdict
from itertools import zip_longest
from pathlib import Path

from openpyxl import load_workbook

from prove_mendeley_open_sources import SOURCES, public_files, resolve_file, download_first

OUT=Path('measured-source-proof/gtnb-unreviewed-learning-candidates.json')
EXPECTED_ROWS=4502
WINDOW_SIZE=400
MIN_NUMERIC_PER_REQUIRED_CHANNEL=20
HEADER_MAP={
    'Maquina':'machine', 'Nombre_producto':'product', 'Peso_producto_gramos':'Product_Weight',
    'Peso_prom_bruto':'Avg_Gross_Weight', 'Consumo_PP_kilos':'PP_Consumption',
    'Consumo_pigmento_kilos':'Pigment_Consumption', 'Kilos_colada':'Flash_kg',
    'Kilos_defectuosos':'Defective_kg', '%Colada':'%Flash', '%Defectuosos':'%Defective',
    '%Reproceso':'%Reprocess', 'Presion_inyeccion_bares':'Injection_Pressure',
    'Presion_retencion_bares':'Retention_Pressure', 'Temp_mat_fundido':'Melt_Temp',
    'Temp_molde_centigrados':'Mold_Temp', 'Tiempo_ciclo':'Cycle_Time',
    'Tiempo_enfriamiento_inyeccion_seg':'Cooling_Time_Injection',
    'Tiempo_expulsion_inyeccion_seg':'Ejection_Time_Injection',
    'Tiempo_retencion_inyeccion_seg':'Retention_Time_Injection',
    'Velocidad_de_Inyección_mm/s':'Injection_Speed',
}
CHANNEL_META={
    'Product_Weight':('recorded-product-unit-weight','g'), 'Avg_Gross_Weight':('recorded-average-gross-weight','g'),
    'PP_Consumption':('recorded-polypropylene-consumption','kg'), 'Pigment_Consumption':('recorded-pigment-consumption','kg'),
    'Flash_kg':('recorded-flash-mass','kg'), 'Defective_kg':('recorded-defective-product-mass','kg'),
    '%Flash':('recorded-flash-percentage','%'), '%Defective':('recorded-defective-percentage','%'),
    '%Reprocess':('recorded-reprocess-percentage','%'), 'Injection_Pressure':('recorded-injection-pressure','bar'),
    'Retention_Pressure':('recorded-holding-pressure','bar'), 'Melt_Temp':('recorded-melt-temperature','degC'),
    'Mold_Temp':('recorded-mould-temperature','degC'), 'Cycle_Time':('recorded-cycle-time','s'),
    'Cooling_Time_Injection':('recorded-injection-cooling-time','s'),
    'Ejection_Time_Injection':('recorded-injection-ejection-time','s'),
    'Retention_Time_Injection':('recorded-holding-time','s'), 'Injection_Speed':('recorded-injection-speed','mm/s'),
}
GROUP_CHANNELS=['Product_Weight','Cycle_Time','Injection_Pressure','Melt_Temp']
QUALITY_CHANNELS=['Injection_Pressure','Retention_Pressure','Mold_Temp','Flash_kg','Defective_kg','%Flash','%Defective']
PROCESS_CHANNELS=['Cycle_Time','Cooling_Time_Injection','Ejection_Time_Injection','Retention_Time_Injection','Injection_Speed','Product_Weight']


def canonical_sha(value):
    raw=json.dumps(value,sort_keys=True,separators=(',',':'),ensure_ascii=False).encode('utf-8')
    return 'sha256:'+hashlib.sha256(raw).hexdigest()
def finite(v): return isinstance(v,(int,float)) and not isinstance(v,bool) and math.isfinite(float(v))
def header_text(v): return str(v or '').strip()
def source_spec(): return next(s for s in SOURCES if s['datasetId']=='mendeley-gtnb4j7bfx-v1')

def download_verified():
    source=source_spec(); file_id,name,expected_sha=source['files'][0]
    _,meta=public_files(source['shortId'],source['version']); _,_,urls=resolve_file(meta,file_id,name,source['shortId'],source['version'])
    td=tempfile.TemporaryDirectory(); path=Path(td.name)/name; download_first(urls,path)
    digest=hashlib.sha256(path.read_bytes()).hexdigest()
    if digest!=expected_sha: td.cleanup(); raise RuntimeError(f'GTNB SHA mismatch: {digest}')
    return path,td,'sha256:'+digest

def numeric_counts(rows, channels):
    return {ch:sum(1 for r in rows if finite(r.get(ch))) for ch in channels}

def best_contiguous_window(rows, channels, size=WINDOW_SIZE):
    """Maximize weakest-channel numeric coverage, then total coverage; earliest wins ties."""
    if not rows: return [], {}, 0
    if len(rows)<=size:
        counts=numeric_counts(rows,channels); return list(rows),counts,0
    counts={ch:0 for ch in channels}
    for r in rows[:size]:
        for ch in channels: counts[ch]+=int(finite(r.get(ch)))
    best_start=0; best_counts=dict(counts); best_score=(min(counts.values()),sum(counts.values()))
    for start in range(1,len(rows)-size+1):
        outgoing=rows[start-1]; incoming=rows[start+size-1]
        for ch in channels:
            counts[ch]-=int(finite(outgoing.get(ch))); counts[ch]+=int(finite(incoming.get(ch)))
        score=(min(counts.values()),sum(counts.values()))
        if score>best_score:
            best_start=start; best_counts=dict(counts); best_score=score
    return rows[best_start:best_start+size],best_counts,best_start

def require_numeric_floor(label, counts):
    weak={ch:n for ch,n in counts.items() if n<MIN_NUMERIC_PER_REQUIRED_CHANNEL}
    if weak: raise RuntimeError(f'{label}: numeric evidence floor {MIN_NUMERIC_PER_REQUIRED_CHANNEL} not met: {weak}')
def formula_subset(formula_counts, channels): return {ch:int(formula_counts.get(ch,0)) for ch in channels}

def signal(channel, rows, formula_counts):
    semantic,unit=CHANNEL_META[channel]; points=[r for r in rows if finite(r[channel])]
    xs=[float(r['_sourceRow']) for r in points]; ys=[float(r[channel]) for r in points]
    if len(xs)<MIN_NUMERIC_PER_REQUIRED_CHANNEL:
        raise RuntimeError(f'{channel}: expected >= {MIN_NUMERIC_PER_REQUIRED_CHANNEL} numeric values, got {len(xs)}')
    rep={'xSemantic':'observation-index','xUnit':'index','xDirection':'increasing','reductionMethod':'coverage-qualified-source-order-window-no-interpolation','originalPointCount':len(rows),'x':xs,'y':ys}
    return {'id':channel.lower().replace('%','pct-'),'label':channel.replace('_',' '),'sourceChannel':channel,'semantic':semantic,'unit':unit,'sourceValueMode':'delivered-direct-or-cached-formula-result','sourceInjectionFormulaCellCount':int(formula_counts.get(channel,0)),'representation':rep,'representationFingerprint':canonical_sha(rep)}
def candidate(cid, rows, channels, suggested, selection, source_fp, formula_counts):
    sigs=[signal(ch,rows,formula_counts) for ch in channels]
    scope=dict(selection); scope['sourceValueMode']='publisher-workbook-stored-values-consistent-with-existing-benchmark-reader'; scope['sourceInjectionFormulaCellCountsByChannel']=formula_subset(formula_counts,channels)
    return {'candidateId':cid,'datasetId':'mendeley-gtnb4j7bfx-v1','sourceArtifact':'modelo.xlsx','sourceFingerprint':source_fp,'sourceScope':scope,'signals':sigs,'candidateFingerprint':canonical_sha(sigs),'suggestedCatalogueCases':suggested,'evidenceBoundary':'Historical public injection-production records only. Formula-backed workbook results remain source-derived records, not raw sensors. Record order is not asserted to be shot-resolved; process fields are not relabelled as actual versus commanded values; quality associations do not establish root cause.'}

def main():
    path,td,fp=download_verified()
    try:
        formula_wb=load_workbook(path,read_only=True,data_only=False)
        value_wb=load_workbook(path,read_only=True,data_only=True)
        if 'nicky' not in formula_wb.sheetnames or 'nicky' not in value_wb.sheetnames: raise RuntimeError('GTNB nicky worksheet missing')
        formula_iter=formula_wb['nicky'].iter_rows(); value_iter=value_wb['nicky'].iter_rows(values_only=True)
        try:
            formula_header=next(formula_iter); value_header=next(value_iter)
        except StopIteration: raise RuntimeError('GTNB worksheet is empty')
        actual=[header_text(v) for v in value_header]; formula_actual=[header_text(c.value) for c in formula_header]
        if actual!=formula_actual: raise RuntimeError('GTNB formula/value workbook header disagreement')
        nonempty=[h for h in actual if h]
        if len(nonempty)!=len(set(nonempty)): raise RuntimeError('GTNB duplicate normalized headers')
        expected=list(HEADER_MAP); missing=[h for h in expected if h not in actual]
        if missing: raise RuntimeError(f'GTNB header drift after whitespace normalization: {missing}')
        col={h:actual.index(h) for h in expected}
        rows=[]; groups=defaultdict(list); formula_counts={canonical:0 for canonical in CHANNEL_META}
        for source_row,pair in enumerate(zip_longest(formula_iter,value_iter,fillvalue=None),start=2):
            formula_cells,values=pair
            if formula_cells is None or values is None: raise RuntimeError('GTNB formula/value workbook row-count disagreement')
            machine=header_text(values[col['Maquina']] if col['Maquina']<len(values) else None).upper()
            if not (re.fullmatch(r'I(?:-|_)?\d+',machine) or machine.startswith('INY')): continue
            product=header_text(values[col['Nombre_producto']] if col['Nombre_producto']<len(values) else None)
            row={'_sourceRow':source_row,'_machine':machine,'_product':product}
            for source_header,canonical in HEADER_MAP.items():
                if canonical in {'machine','product'}: continue
                idx=col[source_header]; v=values[idx] if idx<len(values) else None
                row[canonical]=float(v) if finite(v) else None
                if canonical in formula_counts and idx<len(formula_cells) and formula_cells[idx].data_type=='f': formula_counts[canonical]+=1
            rows.append(row); groups[(machine,product)].append(row)
        if len(rows)!=EXPECTED_ROWS: raise RuntimeError(f'GTNB injection-row count drift: {len(rows)}')

        eligible_groups=[]; best_failed=None
        for key,group_source_rows in groups.items():
            displayed,counts,start=best_contiguous_window(group_source_rows,GROUP_CHANNELS)
            score=(min(counts.values()),sum(counts.values()),len(group_source_rows))
            if best_failed is None or score>best_failed[0]: best_failed=(score,counts,len(group_source_rows))
            if min(counts.values())>=MIN_NUMERIC_PER_REQUIRED_CHANNEL: eligible_groups.append((key,group_source_rows,displayed,counts,start))
        if not eligible_groups:
            summary={'bestWindowCounts':best_failed[1] if best_failed else {},'bestGroupRecordCount':best_failed[2] if best_failed else 0}
            raise RuntimeError(f'GTNB no machine/product group meets numeric evidence floor {MIN_NUMERIC_PER_REQUIRED_CHANNEL}: {summary}')
        largest_key,largest_source_rows,group_rows,group_counts,group_start=max(eligible_groups,key=lambda item:(len(item[1]),min(item[3].values()),sum(item[3].values()),item[0][0],item[0][1]))
        require_numeric_floor('GTNB selected product group',group_counts)
        group_alias='sha256:'+hashlib.sha256(('\u241f'.join(largest_key)).encode('utf-8')).hexdigest()

        quality_rows,quality_counts,quality_start=best_contiguous_window(rows,QUALITY_CHANNELS); require_numeric_floor('GTNB quality window',quality_counts)
        process_rows,process_counts,process_start=best_contiguous_window(rows,PROCESS_CHANNELS); require_numeric_floor('GTNB process window',process_counts)
        candidates=[
            candidate('GTNB-LARGEST-PRODUCT-GROUP-01',group_rows,GROUP_CHANNELS,['MLM-004','MLM-007','MLM-019','MLM-067'],{'selection':'largest injection machine/product group whose best bounded source-order window meets the numeric evidence floor for every required learner channel','minimumNumericValuesPerRequiredChannel':MIN_NUMERIC_PER_REQUIRED_CHANNEL,'groupAlias':group_alias,'groupSourceRecordCount':len(largest_source_rows),'selectedGroupOrdinalStart':group_start,'displayedRecords':len(group_rows),'numericCountsByChannel':group_counts},fp,formula_counts),
            candidate('GTNB-QUALITY-ASSOCIATION-01',quality_rows,QUALITY_CHANNELS,['MLM-040'],{'selection':'bounded contiguous injection-record window maximizing weakest required-channel numeric coverage without sorting measured values','minimumNumericValuesPerRequiredChannel':MIN_NUMERIC_PER_REQUIRED_CHANNEL,'sourceInjectionRecords':len(rows),'selectedInjectionOrdinalStart':quality_start,'displayedRecords':len(quality_rows),'numericCountsByChannel':quality_counts},fp,formula_counts),
            candidate('GTNB-PROCESS-WINDOW-01',process_rows,PROCESS_CHANNELS,['MLM-019','MLM-038','MLM-067'],{'selection':'bounded contiguous injection-record window maximizing weakest required-channel numeric coverage without sorting measured values','minimumNumericValuesPerRequiredChannel':MIN_NUMERIC_PER_REQUIRED_CHANNEL,'sourceInjectionRecords':len(rows),'selectedInjectionOrdinalStart':process_start,'displayedRecords':len(process_rows),'numericCountsByChannel':process_counts},fp,formula_counts),
        ]
        result={'schemaVersion':1,'status':'unreviewed-source-derived-candidates','promotionEligible':False,'candidateCount':len(candidates),'numericEvidenceFloorPerRequiredChannel':MIN_NUMERIC_PER_REQUIRED_CHANNEL,'sourceInjectionFormulaCellCountsByCanonicalChannel':formula_counts,'candidates':candidates,'boundary':'Authoring evidence only. Numeric values follow the existing benchmark policy and may include publisher-stored cached formula results; formula-backed quality metrics remain source-derived records. Product and machine identifiers are not emitted. Candidate selection is based on bounded numeric evidence coverage, not outcome magnitude. Independent engineering review and case-specific binding remain mandatory.'}
        OUT.parent.mkdir(parents=True,exist_ok=True); OUT.write_text(json.dumps(result,indent=2,ensure_ascii=False)+'\n',encoding='utf-8')
        print(json.dumps({'status':result['status'],'candidateCount':len(candidates),'candidateIds':[c['candidateId'] for c in candidates],'selectedGroupNumericCounts':group_counts,'qualityWindowNumericCounts':quality_counts,'processWindowNumericCounts':process_counts,'formulaCountsForQualityChannels':formula_subset(formula_counts,QUALITY_CHANNELS)},separators=(',',':')))
    finally: td.cleanup()
    return 0
if __name__=='__main__': raise SystemExit(main())
