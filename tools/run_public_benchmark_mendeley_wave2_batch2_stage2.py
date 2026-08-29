#!/usr/bin/env python3
from __future__ import annotations

import argparse
import datetime as dt
import hashlib
import io
import json
import re
import urllib.request
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
MANIFEST = ROOT / "data/public-benchmark-results/mendeley-wave2-batch2-stage1.json"
UA = "MouldMaster-Educational-Evidence-Profiler/1.0"
API_ROOT = "https://api.data.mendeley.com"


def get(url: str) -> tuple[bytes, str]:
    req = urllib.request.Request(url, headers={"User-Agent": UA, "Accept": "*/*"})
    with urllib.request.urlopen(req, timeout=120) as r:
        return r.read(), r.geturl()


def public_listing(dataset_id: str, version: int):
    url = f"https://data.mendeley.com/public-api/datasets/{dataset_id}/files?folder_id=root&version={version}"
    raw, _ = get(url)
    payload = json.loads(raw.decode("utf-8"))
    if isinstance(payload, list):
        return payload
    if isinstance(payload, dict):
        for key in ("results", "files", "items", "data"):
            if isinstance(payload.get(key), list):
                return payload[key]
    return []


def file_id(item):
    return str(item.get("id") or item.get("file_id") or item.get("uuid") or "")


def file_url(dataset_id: str, version: int, item):
    details = item.get("content_details") or item.get("contentDetails") or {}
    for key in ("download_url", "downloadUrl"):
        if details.get(key):
            return str(details[key])
        if item.get(key):
            return str(item[key])
    fid = file_id(item)
    if not fid:
        raise RuntimeError("publisher item has no file id")
    return f"{API_ROOT}/datasets/{dataset_id}/files/{fid}/file_downloaded?version={version}"


def safe_text(value):
    if not isinstance(value, str):
        return None
    text = " ".join(value.strip().split())
    if not text or len(text) > 120 or not re.search(r"[A-Za-z]", text):
        return None
    if re.fullmatch(r"[-+0-9.,Ee% /]+", text):
        return None
    return text


def value_token(value):
    if isinstance(value, dt.datetime):
        return value.isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    return repr(value)


def sheet_content_sha256(ws):
    h = hashlib.sha256()
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            h.update(cell.coordinate.encode())
            h.update(b"\0")
            h.update(str(cell.data_type).encode())
            h.update(b"\0")
            h.update(value_token(cell.value).encode("utf-8", "replace"))
            h.update(b"\n")
    return h.hexdigest()


def rich_text(title_obj):
    try:
        paras = title_obj.tx.rich.p
        pieces = []
        for p in paras:
            for r in getattr(p, "r", []) or []:
                if getattr(r, "t", None):
                    pieces.append(r.t)
            if getattr(p, "endParaRPr", None) is not None and getattr(p, "t", None):
                pieces.append(p.t)
        text = " ".join(str(x).strip() for x in pieces if str(x).strip())
        return text or None
    except Exception:
        return None


def title_text(obj):
    if obj is None:
        return None
    direct = rich_text(obj)
    if direct:
        return direct
    try:
        if obj.tx and obj.tx.strRef and obj.tx.strRef.f:
            return obj.tx.strRef.f
    except Exception:
        pass
    return None


def series_ref(ser, attr):
    try:
        obj = getattr(ser, attr, None)
        if obj is None:
            return None
        num = getattr(obj, "numRef", None)
        if num is not None and getattr(num, "f", None):
            return num.f
        st = getattr(obj, "strRef", None)
        if st is not None and getattr(st, "f", None):
            return st.f
    except Exception:
        pass
    return None


def chart_profile(ws):
    out = []
    for chart in getattr(ws, "_charts", []) or []:
        series = []
        for ser in getattr(chart, "ser", []) or []:
            title = None
            try:
                tx = getattr(ser, "tx", None)
                if tx is not None:
                    if getattr(tx, "v", None):
                        title = tx.v
                    elif getattr(tx, "strRef", None) is not None:
                        title = getattr(tx.strRef, "f", None)
            except Exception:
                pass
            series.append({
                "titleOrReference": title,
                "valuesReference": series_ref(ser, "val") or series_ref(ser, "yVal"),
                "categoriesReference": series_ref(ser, "cat") or series_ref(ser, "xVal"),
            })
        out.append({
            "type": type(chart).__name__,
            "title": title_text(getattr(chart, "title", None)),
            "xAxisTitle": title_text(getattr(getattr(chart, "x_axis", None), "title", None)),
            "yAxisTitle": title_text(getattr(getattr(chart, "y_axis", None), "title", None)),
            "series": series,
        })
    return out


def profile_workbook(data: bytes):
    wb = load_workbook(io.BytesIO(data), read_only=False, data_only=False)
    sheets = []
    for ws in wb.worksheets:
        nonempty = numeric = formulas = strings = booleans = dates = 0
        text_labels = []
        seen = set()
        first_text_cells = []
        formula_cells = []
        row_profiles = []
        for row_idx in range(1, ws.max_row + 1):
            row_numeric = row_formula = row_string = 0
            row_labels = []
            for cell in ws[row_idx]:
                v = cell.value
                if v is None:
                    continue
                nonempty += 1
                if cell.data_type == "f":
                    formulas += 1
                    row_formula += 1
                    if len(formula_cells) < 150:
                        formula_cells.append({"coordinate": cell.coordinate, "formula": str(v)[:160]})
                    continue
                if isinstance(v, bool):
                    booleans += 1
                elif isinstance(v, (int, float)):
                    numeric += 1
                    row_numeric += 1
                elif getattr(cell, "is_date", False):
                    dates += 1
                elif isinstance(v, str):
                    strings += 1
                    row_string += 1
                    t = safe_text(v)
                    if t:
                        if t not in seen and len(text_labels) < 160:
                            seen.add(t)
                            text_labels.append(t)
                        if len(first_text_cells) < 120:
                            first_text_cells.append({"coordinate": cell.coordinate, "text": t})
                        if len(row_labels) < 12:
                            row_labels.append({"coordinate": cell.coordinate, "text": t})
            if row_numeric or row_formula or row_string:
                row_profiles.append({
                    "row": row_idx,
                    "numericConstantCells": row_numeric,
                    "formulaCells": row_formula,
                    "stringCells": row_string,
                    "textLabels": row_labels,
                })
        sheets.append({
            "sheet": ws.title,
            "sheetContentSha256": sheet_content_sha256(ws),
            "maxRow": ws.max_row,
            "maxColumn": ws.max_column,
            "nonEmptyCells": nonempty,
            "numericConstantCells": numeric,
            "formulaCells": formulas,
            "stringCells": strings,
            "booleanCells": booleans,
            "dateCells": dates,
            "textLabels": text_labels,
            "firstTextCells": first_text_cells,
            "rowTypeProfiles": row_profiles,
            "formulaCoordinatesAndExpressions": formula_cells,
            "charts": chart_profile(ws),
            "rawNumericValuesEmitted": False,
        })
    return sheets


def profile_source(source, dataset_id: str, version: int):
    listing = public_listing(dataset_id, version)
    by_id = {file_id(x): x for x in listing}
    files = []
    for expected in source["apiFiles"]:
        fid = expected["id"]
        item = by_id.get(fid)
        if item is None:
            raise RuntimeError(f"publisher file id disappeared: {fid}")
        data, final_url = get(file_url(dataset_id, version, item))
        digest = hashlib.sha256(data).hexdigest()
        if digest.lower() != expected["sha256"].lower():
            raise RuntimeError(f"publisher SHA mismatch for {expected['name']}: {digest}")
        files.append({
            "fileId": fid,
            "fileName": expected["name"],
            "sizeBytes": len(data),
            "sha256": digest,
            "publisherSha256Matched": True,
            "resolvedUrl": final_url,
            "workbook": profile_workbook(data),
            "rawPublisherFileCommitted": False,
            "rawRowsOrNumericValuesEmitted": False,
        })
    return files


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--output", type=Path, required=True)
    ap.add_argument("--retrieved-date", required=True)
    args = ap.parse_args()
    manifest = json.loads(MANIFEST.read_text(encoding="utf-8"))
    selected = []
    specs = {
        "mendeley-c3pt29jt7c-v1": ("c3pt29jt7c", 1),
        "mendeley-yxz2w7ctnh-v1": ("yxz2w7ctnh", 1),
    }
    for source in manifest["sources"]:
        if source["datasetId"] not in specs:
            continue
        did, version = specs[source["datasetId"]]
        selected.append({
            "datasetId": source["datasetId"],
            "doi": source["doi"],
            "license": source["license"],
            "status": "retrieved-profile-needs-semantic-review",
            "files": profile_source(source, did, version),
            "countsAsFullyProfiledMeasuredDataset": False,
            "acceptedMeasuredTimeSeriesSamples": 0,
        })
    result = {
        "schema": 2,
        "status": "retrieved-workbook-layouts-needing-semantic-review",
        "retrievedDate": args.retrieved_date,
        "sources": selected,
        "summary": {
            "sourcesRetrieved": len(selected),
            "filesRetrieved": sum(len(x["files"]) for x in selected),
            "fullyProfiledAccepted": 0,
            "acceptedMeasuredTimeSeriesSamples": 0,
            "rawPublisherFilesCommitted": False,
            "rawRowsOrNumericValuesEmitted": False,
        },
        "evidenceBoundary": "Exact CC BY 4.0 publisher workbooks are temporarily retrieved and SHA-verified. Only workbook structure, safe text labels, cell-type/row counts, formulas, chart references and cryptographic fingerprints are emitted. Raw numeric measurements and source files are not retained. Worksheet content hashes are used only to detect duplicate sheets. Acceptance requires a source-specific semantic mapping."
    }
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(result, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(json.dumps(result["summary"], indent=2))


if __name__ == "__main__":
    main()
