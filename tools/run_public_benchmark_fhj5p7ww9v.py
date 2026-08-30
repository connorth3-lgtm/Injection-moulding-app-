#!/usr/bin/env python3
from __future__ import annotations

import argparse
import hashlib
import json
import tempfile
import urllib.request
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
CONTRACT = ROOT / "data/public-benchmark-contracts/fhj5p7ww9v-v1.json"
DATASET_ID = "fhj5p7ww9v"
VERSION = 1
PUBLIC_FILES_ENDPOINT = f"https://data.mendeley.com/public-api/datasets/{DATASET_ID}/files?folder_id=root&version={VERSION}"
API_ROOT = "https://api.data.mendeley.com"
UA = "MouldMaster-Educational-Evidence-Profiler/1.0"

MEASURED_SHEETS = {"weight", "flexural strength", "flexural modulus"}
MATERIAL_ROWS = {
    4: "Solid PP",
    7: "All-PP composite",
    10: "PP foam",
    13: "All-PP composite foam",
}
PROCESS_ROWS = {2: "injection temperature", 3: "injection speed"}
SPREAD_ROWS = {5, 6, 8, 9, 11, 12, 14, 15}


def get(url, accept="*/*"):
    req = urllib.request.Request(url, headers={"Accept": accept, "User-Agent": UA})
    with urllib.request.urlopen(req, timeout=120) as r:
        return r.read(), r.geturl()


def flatten_files(payload):
    if isinstance(payload, list):
        return payload
    if isinstance(payload, dict):
        for key in ("results", "files", "items", "data"):
            if isinstance(payload.get(key), list):
                return payload[key]
    raise RuntimeError("Mendeley public file list did not contain files")


def file_id(item):
    return str(item.get("id") or item.get("file_id") or item.get("uuid") or "").strip()


def file_name(item):
    return str(item.get("filename") or item.get("name") or "").strip()


def file_url(item):
    details = item.get("content_details") or item.get("contentDetails") or {}
    for candidate in (details.get("download_url"), details.get("downloadUrl"), item.get("download_url"), item.get("downloadUrl")):
        if candidate:
            return str(candidate)
    fid = file_id(item)
    if fid:
        return f"{API_ROOT}/datasets/{DATASET_ID}/files/{fid}/file_downloaded?version={VERSION}"
    return None


def aggregate_sheet(ws):
    text_cells = []
    numeric_by_row = {}
    numeric_by_column = {}
    formula_count = 0
    for row in ws.iter_rows():
        row_numeric = 0
        for cell in row:
            value = cell.value
            if value is None:
                continue
            if isinstance(value, str):
                if value.startswith("="):
                    formula_count += 1
                else:
                    cleaned = " ".join(value.split())
                    if cleaned:
                        text_cells.append({"cell": cell.coordinate, "text": cleaned[:200]})
            elif isinstance(value, (int, float)) and not isinstance(value, bool):
                row_numeric += 1
                numeric_by_column[cell.column_letter] = numeric_by_column.get(cell.column_letter, 0) + 1
        if row_numeric:
            numeric_by_row[row[0].row] = row_numeric
    text_by_cell = {x["cell"]: x["text"] for x in text_cells}

    exact_layout = (
        ws.title.strip().lower() in MEASURED_SHEETS
        and text_by_cell.get("A1") == "No."
        and text_by_cell.get("A2") == "injection temperature"
        and text_by_cell.get("A3") == "injection speed"
        and all(text_by_cell.get(f"A{r}") == label for r, label in MATERIAL_ROWS.items())
        and all(numeric_by_row.get(r) == 8 for r in [1, 2, 3, *MATERIAL_ROWS.keys(), *SPREAD_ROWS])
    )

    direct_measured = sum(numeric_by_row.get(r, 0) for r in MATERIAL_ROWS) if exact_layout else 0
    process_factor = sum(numeric_by_row.get(r, 0) for r in PROCESS_ROWS) if exact_layout else 0
    spread_summary = sum(numeric_by_row.get(r, 0) for r in SPREAD_ROWS) if exact_layout else 0
    condition_ids = numeric_by_row.get(1, 0) if exact_layout else 0
    return {
        "textCells": text_cells,
        "textCellCount": len(text_cells),
        "numericCellsByRow": [{"row": r, "numericCells": c} for r, c in sorted(numeric_by_row.items())],
        "numericCellsByColumn": numeric_by_column,
        "numericCellCount": sum(numeric_by_row.values()),
        "formulaCellCount": formula_count,
        "semanticMap": {
            "recognized": exact_layout,
            "outcomeFamily": ws.title.strip().lower() if exact_layout else None,
            "experimentalConditionColumns": condition_ids,
            "processFactorRows": [{"row": r, "label": label, "numericCells": numeric_by_row.get(r, 0)} for r, label in PROCESS_ROWS.items()] if exact_layout else [],
            "directMeasuredMaterialRows": [{"row": r, "label": label, "numericCells": numeric_by_row.get(r, 0)} for r, label in MATERIAL_ROWS.items()] if exact_layout else [],
            "spreadSummaryRows": sorted(SPREAD_ROWS) if exact_layout else [],
            "directMeasuredOutcomeCells": direct_measured,
            "processFactorCellsExcludedFromMeasuredOutcomes": process_factor,
            "spreadSummaryCellsExcludedFromDirectMeasurements": spread_summary,
            "formulaCellsExcludedAsDerived": formula_count,
        },
        "numericValuesEmitted": False,
    }


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--output", type=Path, required=True)
    ap.add_argument("--retrieved-date", required=True)
    args = ap.parse_args()
    contract = json.loads(CONTRACT.read_text(encoding="utf-8"))

    raw, _ = get(PUBLIC_FILES_ENDPOINT, "application/json")
    files = flatten_files(json.loads(raw.decode("utf-8")))
    expected = contract["source"]["expectedPublisherFile"]
    matches = [x for x in files if file_name(x) == expected]
    if len(matches) != 1:
        raise RuntimeError(f"expected exactly one {expected!r}; found {[file_name(x) for x in files]}")
    item = matches[0]
    url = file_url(item)
    if not url:
        raise RuntimeError("publisher file has no download route")
    data, final_url = get(url)
    digest = hashlib.sha256(data).hexdigest()

    with tempfile.TemporaryDirectory() as td:
        path = Path(td) / "source.xlsx"
        path.write_bytes(data)
        wb = load_workbook(path, read_only=True, data_only=False)
        profiles = []
        total_direct = 0
        total_rows = 0
        total_columns = 0
        recognized_sheets = 0
        for ws in wb.worksheets:
            layout = aggregate_sheet(ws)
            if layout["semanticMap"]["recognized"]:
                recognized_sheets += 1
                total_direct += layout["semanticMap"]["directMeasuredOutcomeCells"]
            # pandas dimensions give a stable substantive table shape after empty edges are removed.
            df = pd.read_excel(path, sheet_name=ws.title).dropna(axis=0, how="all").dropna(axis=1, how="all")
            total_rows += int(len(df))
            total_columns += int(len(df.columns))
            profiles.append({"sheet": ws.title, "rows": int(len(df)), "columns": int(len(df.columns)), "layout": layout})

    details = item.get("content_details") or item.get("contentDetails") or {}
    publisher_sha = details.get("sha256_hash") or details.get("sha256Hash")
    recognized = recognized_sheets == 3 and total_direct == 96
    status = "completed-restricted-noncommercial-measured-benchmark" if recognized else "retrieved-profile-needs-semantic-review"
    result = {
        "schema": 3,
        "status": status,
        "retrievedDate": args.retrieved_date,
        "source": {
            "datasetId": contract["datasetId"],
            "datasetDoi": contract["source"]["datasetDoi"],
            "version": VERSION,
            "license": contract["source"]["license"],
            "publisherFileId": file_id(item),
            "publisherFileName": expected,
            "publisherReportedSizeBytes": details.get("size") if details.get("size") is not None else item.get("size"),
            "publisherSha256": publisher_sha,
            "retrievedSizeBytes": len(data),
            "sha256": digest,
            "resolvedUrl": final_url,
        },
        "profile": {
            "sheets": profiles,
            "sheetCount": len(profiles),
            "recognizedMeasuredOutcomeSheets": recognized_sheets,
            "experimentalConditionsPerSheet": 8 if recognized else None,
            "materialGroups": list(MATERIAL_ROWS.values()) if recognized else [],
            "measuredOutcomeFamilies": sorted(MEASURED_SHEETS) if recognized else [],
            "totalRowsAcrossSheets": total_rows,
            "totalColumnsAcrossSheets": total_columns,
            "recordLevelMeasuredOutcomeValues": total_direct,
            "measurementLevel": "source-reported outcome values by material and experimental condition; not raw replicate-level observations" if recognized else None,
            "acceptedMeasuredTimeSeriesSamples": 0,
            "rawRowsOrCellValuesEmitted": False,
            "numericMeasurementValuesEmitted": False,
        },
        "acceptance": {
            "countsAsFullyProfiledMeasuredDataset": recognized,
            "semanticLayoutRecognized": recognized,
            "useScope": "noncommercial-education-research-only",
            "commercialReuseAllowed": False,
            "rawRedistributionAllowedUnderProjectPolicy": False,
            "acceptedMeasuredTimeSeriesSamples": 0,
        },
        "retrieval": {
            "rawPublisherFileCommitted": False,
            "rawRowsOrCellValuesUploadedAsArtifact": False,
        },
        "evidenceBoundary": "The exact CC BY-NC 3.0 workbook is fingerprinted. Direct measured outcome cells are only the four material rows on the three outcome sheets across eight experimental-condition columns (4 x 3 x 8 = 96). Injection-temperature/speed rows, deviation summaries and formula-based comparison/percentage cells are explicitly excluded. These 96 cells are source-reported outcome values, not raw replicate measurements or high-frequency time-series samples." if recognized else "The exact publisher workbook was retrieved, but the expected aggregate semantic layout did not reconcile; it remains non-counting."
    }
    if publisher_sha and len(str(publisher_sha)) == 64:
        result["source"]["publisherSha256Matched"] = str(publisher_sha).lower() == digest.lower()
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(result, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(json.dumps({
        "status": status,
        "sha256": digest,
        "recognizedMeasuredOutcomeSheets": recognized_sheets,
        "recordLevelMeasuredOutcomeValues": total_direct,
        "acceptedMeasuredTimeSeriesSamples": 0,
    }, indent=2))


if __name__ == "__main__":
    main()
