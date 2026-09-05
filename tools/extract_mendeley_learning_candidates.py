#!/usr/bin/env python3
"""Create compact, unreviewed measured-learning candidates from proven open Mendeley workbooks.

This is a source-recovery/authoring aid, not a promotion path. Exact publisher files are
re-downloaded and SHA-256 verified through prove_mendeley_open_sources before any numeric
values are read. Outputs are bounded/transformed learner candidate representations only;
raw workbooks and full raw rows are never retained or uploaded.
"""
from __future__ import annotations

import copy
import hashlib
import json
import math
import tempfile
from pathlib import Path

from openpyxl import load_workbook

from prove_mendeley_open_sources import SOURCES, public_files, resolve_file, download_first

OUT = Path("measured-source-proof/mendeley-unreviewed-learning-candidates.json")


def sha256_json(value) -> str:
    payload=json.dumps(value,sort_keys=True,separators=(",",":"),ensure_ascii=False).encode("utf-8")
    return "sha256:"+hashlib.sha256(payload).hexdigest()


def isnum(v):
    return isinstance(v,(int,float)) and not isinstance(v,bool) and math.isfinite(float(v))


def norm(v) -> str:
    return "".join(str(v or "").strip().lower().split())


def download_verified(source: dict, expected_name: str) -> tuple[Path, tempfile.TemporaryDirectory, str]:
    file_spec=next(x for x in source["files"] if x[1]==expected_name)
    file_id,name,expected_sha=file_spec
    _,meta=public_files(source["shortId"],source["version"])
    _chosen,_resolved,urls=resolve_file(meta,file_id,name,source["shortId"],source["version"])
    td=tempfile.TemporaryDirectory()
    path=Path(td.name)/name
    download_first(urls,path)
    digest=hashlib.sha256(path.read_bytes()).hexdigest()
    if digest!=expected_sha:
        td.cleanup(); raise RuntimeError(f"{source['datasetId']}/{name}: SHA mismatch {digest}")
    return path,td,"sha256:"+digest


def numeric_column(ws, col: str, start: int, end: int) -> list[float]:
    values=[]
    for row in range(start,end+1):
        cell=ws[f"{col}{row}"]
        if cell.data_type=="f": continue
        if isnum(cell.value): values.append(float(cell.value))
    return values


def reduce_points(x: list[float], y: list[float], limit: int=400) -> tuple[list[float],list[float]]:
    if len(x)!=len(y): raise ValueError("x/y mismatch")
    if len(x)<=limit: return x,y
    idx=sorted({round(i*(len(x)-1)/(limit-1)) for i in range(limit)})
    return [x[i] for i in idx],[y[i] for i in idx]


def maximal_decreasing_branch(x: list[float], y: list[float], *, maximum_terminal_reversal_points: int=2):
    """Select the source-ordered decreasing branch without sorting or interpolating."""
    if len(x)!=len(y) or len(x)<3:
        raise RuntimeError("decreasing branch requires aligned source pairs")
    cut=len(x)
    for i,(a,b) in enumerate(zip(x,x[1:])):
        if float(b)>float(a):
            cut=i+1
            break
    selected_x=x[:cut]; selected_y=y[:cut]
    excluded=len(x)-cut
    if not all(float(a)>=float(b) for a,b in zip(selected_x,selected_x[1:])):
        raise RuntimeError("source series is not monotonic before terminal reversal")
    if excluded>maximum_terminal_reversal_points:
        raise RuntimeError(f"source series terminal reversal grew to {excluded} points")
    if cut < len(x)-maximum_terminal_reversal_points:
        raise RuntimeError("source series reversal is not confined to the terminal tail")
    return selected_x,selected_y,{"sourceNumericPairCount":len(x),"selectedDecreasingPairCount":cut,"excludedTerminalReversalPairCount":excluded}


def experiment_replicate_summary(ws, cols: list[str], start: int, end: int) -> tuple[list[float],list[float],list[float]]:
    x=[]; medians=[]; spreads=[]
    for row in range(start,end+1):
        vals=[float(ws[f"{c}{row}"].value) for c in cols if isnum(ws[f"{c}{row}"].value) and ws[f"{c}{row}"].data_type!="f"]
        if not vals: continue
        ordered=sorted(vals); n=len(ordered)
        median=ordered[n//2] if n%2 else (ordered[n//2-1]+ordered[n//2])/2
        x.append(float(row-start+1)); medians.append(median); spreads.append(max(vals)-min(vals))
    return x,medians,spreads


def signal(id_,source_channel,semantic,unit,x_semantic,x_unit,x,y,reduction,original,x_direction=None):
    rep={"xSemantic":x_semantic,"xUnit":x_unit,"reductionMethod":reduction,"originalPointCount":original,"x":x,"y":y}
    if x_direction is not None:
        rep["xDirection"]=x_direction
    return {"id":id_,"sourceChannel":source_channel,"semantic":semantic,"unit":unit,"representation":rep,"representationFingerprint":sha256_json(rep)}


def four_h98(source):
    path,td,fp=download_verified(source,"Raw Data.xlsx")
    try:
        wb=load_workbook(path,read_only=False,data_only=False); ws=wb["Sheet1"]
        expected={"B1":"Input","E1":"Output","A2":"Expt No.","B2":"GNP %","C2":"Temperature","D2":"Pressure (MPa)","J2":"Average Tensile modulus","P2":"Average Hardness","V2":"Average Toughness (J)"}
        for cell,text in expected.items():
            if str(ws[cell].value or "").strip()!=text: raise RuntimeError(f"4h98 label drift {cell}: {ws[cell].value!r}")
        series=[]
        for ident,cols,semantic,unit in [
            ("tensile",list("EFGHI"),"tensile-modulus-replicate-summary","GPa"),
            ("hardness",list("KLMNO"),"hardness-replicate-summary","HV"),
            ("toughness",list("QRSTU"),"toughness-replicate-summary","J")]:
            x,median,spread=experiment_replicate_summary(ws,cols,4,38)
            if len(x)!=35: raise RuntimeError(f"4h98 {ident}: expected 35 experiment summaries, got {len(x)}")
            series.append(signal(f"{ident}-median",f"Sheet1!{cols[0]}:{cols[-1]}",semantic+"-median",unit,"experiment-index","index",x,median,"per-experiment-median-of-five-direct-replicates",175,"increasing"))
            series.append(signal(f"{ident}-spread",f"Sheet1!{cols[0]}:{cols[-1]}",semantic+"-range",unit,"experiment-index","index",x,spread,"per-experiment-max-minus-min-of-five-direct-replicates",175,"increasing"))
        return {"candidateId":"MEND-4H98-REPLICATE-SUMMARY-01","datasetId":source["datasetId"],"sourceArtifact":"Raw Data.xlsx","sourceFingerprint":fp,"sourceScope":{"sheet":"Sheet1","rows":"4:38","excludedDerivedColumns":["J","P","V"]},"signals":series,"candidateFingerprint":sha256_json(series),"suggestedCatalogueCases":["MLM-030","MLM-049","MLM-056"],"evidenceBoundary":"Per-experiment summaries of direct replicate measurements. Derived publisher average columns are excluded; these data do not establish production root cause."}
    finally: td.cleanup()


def six_k8(source):
    path,td,fp=download_verified(source,"Data.xlsx")
    try:
        wb=load_workbook(path,read_only=False,data_only=False); candidates=[]
        ws=wb["Figure7"]
        expected7={"A2":"time(s)","B2":"p(mpa)","C2":"vsp(mm³/g)"}
        actual7={cell:norm(ws[cell].value) for cell in expected7}
        if any(actual7[cell]!=expected for cell,expected in expected7.items()):
            raise RuntimeError(f"6k8 Figure7 header drift: {actual7}")
        xs=[]; pressure=[]; volume=[]
        for row in range(3,ws.max_row+1):
            a,b,c=ws[f"A{row}"].value,ws[f"B{row}"].value,ws[f"C{row}"].value
            if isnum(a) and isnum(b) and isnum(c): xs.append(float(a)); pressure.append(float(b)); volume.append(float(c))
        if len(xs)!=2818: raise RuntimeError(f"6k8 Figure7 row count drift: {len(xs)}")
        rx,rp=reduce_points(xs,pressure); rx2,rv=reduce_points(xs,volume)
        if rx!=rx2: raise RuntimeError("6k8 Figure7 reduction axis mismatch")
        signals=[
            signal("pressure","Figure7!B","pressure-special-isothermal","MPa","time","s",rx,rp,"deterministic-endpoint-preserving-index-reduction",len(xs),"increasing"),
            signal("specific-volume","Figure7!C","specific-volume-special-isothermal","mm3/g","time","s",rx,rv,"deterministic-endpoint-preserving-index-reduction",len(xs),"increasing")]
        candidates.append({"candidateId":"MEND-6K8-FIGURE7-01","datasetId":source["datasetId"],"sourceArtifact":"Data.xlsx","sourceFingerprint":fp,"sourceScope":{"sheet":"Figure7","columns":["A","B","C"]},"signals":signals,"candidateFingerprint":sha256_json(signals),"suggestedCatalogueCases":["MLM-050"],"evidenceBoundary":"Polypropylene pvT material-characterization trace; not an injection-machine cycle trace."})

        ws=wb["Figure2"]; series=[]; branch_selection={}
        for ident,xcol,ycol,semantic in [("200bar","A","B","specific-volume-200bar-isobaric-cooling"),("400bar","E","F","specific-volume-400bar-isobaric-cooling"),("800bar","I","J","specific-volume-800bar-isobaric-cooling")]:
            xs=[]; ys=[]
            for row in range(1,ws.max_row+1):
                a,b=ws[f"{xcol}{row}"].value,ws[f"{ycol}{row}"].value
                if isnum(a) and isnum(b): xs.append(float(a)); ys.append(float(b))
            if not xs: raise RuntimeError(f"6k8 Figure2 no numeric pair for {ident}")
            branch_x,branch_y,selection=maximal_decreasing_branch(xs,ys)
            if selection["excludedTerminalReversalPairCount"]!=1:
                raise RuntimeError(f"6k8 Figure2 {ident}: expected exactly one terminal reversal pair, got {selection['excludedTerminalReversalPairCount']}")
            branch_selection[ident]=selection
            rx,ry=reduce_points(branch_x,branch_y)
            series.append(signal(ident,f"Figure2!{ycol}",semantic,"mm3/g","temperature","degC",rx,ry,"source-order-maximal-decreasing-branch-then-deterministic-index-reduction",len(xs),"decreasing"))
        candidates.append({"candidateId":"MEND-6K8-FIGURE2-01","datasetId":source["datasetId"],"sourceArtifact":"Data.xlsx","sourceFingerprint":fp,"sourceScope":{"sheet":"Figure2","pressureSeriesBar":[200,400,800],"branchSelection":branch_selection},"signals":series,"candidateFingerprint":sha256_json(series),"suggestedCatalogueCases":["MLM-051"],"evidenceBoundary":"Specific-volume response is limited to the source-ordered decreasing-temperature branch for each source-labelled isobaric series. One terminal temperature-reversal pair per series is explicitly excluded from this cooling-branch candidate; no causal interpretation is assigned to that reversal."})
        return candidates
    finally: td.cleanup()


def _copy_signal_with_artifact(candidate: dict, source_channel: str) -> dict:
    selected=copy.deepcopy(next(s for s in candidate["signals"] if s["sourceChannel"]==source_channel))
    selected["sourceArtifact"]=candidate["sourceArtifact"]
    return selected


def yxz(source):
    results=[]
    for name,sheet,marker,cols,semantics,units,rows,cases in [
        ("data_tensile_3d_print_d_ryan.xlsx","tensile_PLA","M1",list("OPQR"),["pla-injection-tensile-modulus","pla-injection-maximum-force","pla-injection-maximum-stress","pla-injection-elongation-at-maximum"],["GPa","N","MPa","%"],(4,22),["MLM-033"]),
        ("data_3pbending_3d_print_d_ryan.xlsx","bending_PLA","O1",list("PQRS"),["pla-injection-bending-thickness","pla-injection-bending-width","pla-injection-bending-maximum-force","pla-injection-bending-deflection-at-max"],["mm","mm","N","mm"],(4,34),["MLM-034"])]:
        path,td,fp=download_verified(source,name)
        try:
            wb=load_workbook(path,read_only=False,data_only=False); ws=wb[sheet]
            marker_text=str(ws[marker].value or "")
            if "moulded" not in marker_text.lower() or ("injection" not in marker_text.lower() and "inke" not in marker_text.lower()):
                raise RuntimeError(f"yxz injection marker drift {sheet}/{marker}: {marker_text!r}")
            sigs=[]
            for col,semantic,unit in zip(cols,semantics,units):
                vals=numeric_column(ws,col,rows[0],rows[1]); x=[float(i+1) for i in range(len(vals))]
                if not vals: raise RuntimeError(f"yxz no direct values in {sheet}!{col}")
                sigs.append(signal(col,f"{sheet}!{col}",semantic,unit,"observation-index","index",x,vals,"direct-injection-block-values-no-interpolation",len(vals),"increasing"))
            results.append({"candidateId":f"MEND-YXZ-{sheet.upper()}-01","datasetId":source["datasetId"],"sourceArtifact":name,"sourceFingerprint":fp,"sourceScope":{"sheet":sheet,"marker":marker_text,"columns":cols,"rows":list(rows)},"signals":sigs,"candidateFingerprint":sha256_json(sigs),"suggestedCatalogueCases":cases,"evidenceBoundary":"Only the explicitly labelled injection-moulded block is included; FDM, energy, impact and formula-derived content remain excluded."})
        finally: td.cleanup()

    tensile,bending=results
    combined_signals=[
        _copy_signal_with_artifact(tensile,"tensile_PLA!P"),
        _copy_signal_with_artifact(bending,"bending_PLA!R"),
    ]
    combined={
        "candidateId":"MEND-YXZ-TENSILE-BENDING-FORCE-01",
        "datasetId":source["datasetId"],
        "sourceArtifacts":[
            {"name":tensile["sourceArtifact"],"sha256":tensile["sourceFingerprint"]},
            {"name":bending["sourceArtifact"],"sha256":bending["sourceFingerprint"]},
        ],
        "sourceScope":{
            "comparison":"injection-moulded PLA tensile versus three-point-bending maximum-force measurements",
            "components":[
                {"artifact":tensile["sourceArtifact"],"sheet":"tensile_PLA","sourceChannel":"tensile_PLA!P"},
                {"artifact":bending["sourceArtifact"],"sheet":"bending_PLA","sourceChannel":"bending_PLA!R"},
            ],
            "coordinateBoundary":"Each signal retains its own source-order observation index; the records are not paired specimen-by-specimen.",
        },
        "signals":combined_signals,
        "candidateFingerprint":sha256_json(combined_signals),
        "suggestedCatalogueCases":["MLM-055"],
        "evidenceBoundary":"Both measurements are direct injection-moulded PLA maximum-force outcomes in N, but tensile and bending tests use different loading modes and specimen contexts. Equal units do not make them the same mechanical property, and the comparison does not establish a production root cause.",
    }
    results.append(combined)
    return results


def main() -> int:
    by_id={s["datasetId"]:s for s in SOURCES}
    candidates=[four_h98(by_id["mendeley-4h98rz9f92-v3"])]
    candidates.extend(six_k8(by_id["mendeley-6k8fpbrd9s-v1"]))
    candidates.extend(yxz(by_id["mendeley-yxz2w7ctnh-v1"]))
    result={"schemaVersion":1,"status":"unreviewed-source-derived-candidates","promotionEligible":False,"candidateCount":len(candidates),"candidates":candidates,"boundary":"Authoring evidence only. These compact source-derived representations require independent engineering review and a case-specific binding before learner promotion. Raw publisher workbooks and full raw rows are not retained."}
    OUT.parent.mkdir(parents=True,exist_ok=True); OUT.write_text(json.dumps(result,indent=2,ensure_ascii=False)+"\n",encoding="utf-8")
    print(json.dumps({"status":result["status"],"candidateCount":len(candidates),"candidateIds":[c["candidateId"] for c in candidates]},separators=(",",":")))
    return 0

if __name__=="__main__": raise SystemExit(main())
