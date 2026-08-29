#!/usr/bin/env python3
from __future__ import annotations

import argparse
import hashlib
import json
import tempfile
import urllib.request
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

DATASET_ID = "4h98rz9f92"
VERSION = 3
DOI = "10.17632/4h98rz9f92.3"
EXPECTED_FILE_ID = "368356fe-618c-4eab-82e6-53dc86762943"
EXPECTED_FILE = "Raw Data.xlsx"
EXPECTED_SHA256 = "39210169aac62a1455603d37cdffaca93cf0c46189ea4258c5f3c0a4a37255c9"
PUBLIC_FILES_ENDPOINT = f"https://data.mendeley.com/public-api/datasets/{DATASET_ID}/files?folder_id=root&version={VERSION}"
API_ROOT = "https://api.data.mendeley.com"
UA = "MouldMaster-Educational-Evidence-Profiler/1.0"


def get(url: str, accept: str = "*/*"):
    req = urllib.request.Request(url, headers={"Accept": accept, "User-Agent": UA})
    with urllib.request.urlopen(req, timeout=120) as r:
        return r.read(), r.geturl()


def flatten_files(payload):
    if isinstance(payload, list): return payload
    if isinstance(payload, dict):
        for key in ("results", "files", "items", "data"):
            if isinstance(payload.get(key), list): return payload[key]
    return []


def file_id(item): return str(item.get("id") or item.get("file_id") or item.get("uuid") or "").strip()
def file_name(item): return str(item.get("filename") or item.get("name") or "").strip()


def file_url(item):
    details = item.get("content_details") or item.get("contentDetails") or {}
    for candidate in (details.get("download_url"), details.get("downloadUrl"), item.get("download_url"), item.get("downloadUrl")):
        if candidate: return str(candidate)
    return f"{API_ROOT}/datasets/{DATASET_ID}/files/{file_id(item)}/file_downloaded?version={VERSION}"


def publisher_sha(item):
    details = item.get("content_details") or item.get("contentDetails") or {}
    return details.get("sha256_hash") or details.get("sha256Hash") or item.get("sha256") or item.get("sha256_hash")


def profile_sheet(ws):
    numeric_by_column = {}
    text_cells = {}
    formula_cells = 0
    nonempty_rows = set()
    nonempty_cols = set()
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if v is None: continue
            nonempty_rows.add(cell.row); nonempty_cols.add(cell.column)
            if isinstance(v, str):
                if v.startswith("="): formula_cells += 1
                else: text_cells[cell.coordinate] = " ".join(v.split())[:240]
            elif isinstance(v, (int, float)) and not isinstance(v, bool):
                col = get_column_letter(cell.column)
                numeric_by_column[col] = numeric_by_column.get(col, 0) + 1
    return {
        "sheet": ws.title,
        "maxRow": ws.max_row,
        "maxColumn": ws.max_column,
        "nonEmptyRowCount": len(nonempty_rows),
        "nonEmptyColumnCount": len(nonempty_cols),
        "numericCells": sum(numeric_by_column.values()),
        "numericCellsByColumn": numeric_by_column,
        "formulaCells": formula_cells,
        "textLabels": text_cells,
        "numericValuesEmitted": False
    }


def main():
    ap = argparse.ArgumentParser(); ap.add_argument("--output", required=True); ap.add_argument("--retrieved-date", required=True); args = ap.parse_args()
    raw, _ = get(PUBLIC_FILES_ENDPOINT, "application/json")
    files = flatten_files(json.loads(raw.decode("utf-8")))
    matches = [x for x in files if file_id(x) == EXPECTED_FILE_ID and file_name(x) == EXPECTED_FILE]
    if len(matches) != 1: raise RuntimeError("exact HDPE/GNP raw workbook identity drifted")
    item = matches[0]
    psha = str(publisher_sha(item) or "").lower()
    if psha != EXPECTED_SHA256: raise RuntimeError(f"publisher manifest SHA drifted: {psha}")
    data, final_url = get(file_url(item))
    digest = hashlib.sha256(data).hexdigest()
    if digest != EXPECTED_SHA256: raise RuntimeError(f"retrieved HDPE/GNP workbook SHA drifted: {digest}")
    with tempfile.TemporaryDirectory() as td:
        p = Path(td) / EXPECTED_FILE; p.write_bytes(data)
        wb = load_workbook(p, read_only=True, data_only=False)
        sheets = [profile_sheet(ws) for ws in wb.worksheets]
    total_numeric = sum(s["numericCells"] for s in sheets)
    total_formulas = sum(s["formulaCells"] for s in sheets)
    result = {
        "schema": 1,
        "status": "retrieved-profile-needs-semantic-review",
        "retrievedDate": args.retrieved_date,
        "source": {"datasetId":"mendeley-4h98rz9f92-v3","datasetDoi":DOI,"version":VERSION,"license":"CC BY 4.0","publisherFileId":EXPECTED_FILE_ID,"publisherFileName":EXPECTED_FILE,"sha256":digest,"publisherSha256Matched":True,"retrievedSizeBytes":len(data),"resolvedUrl":final_url},
        "profile": {"sheetCount":len(sheets),"sheets":sheets,"totalNumericCells":total_numeric,"totalFormulaCells":total_formulas,"rawRowsOrCellValuesEmitted":False,"numericMeasurementValuesEmitted":False},
        "acceptance": {"countsAsFullyProfiledMeasuredDataset":False,"acceptedMeasuredTimeSeriesSamples":0,"stage3SemanticMappingRequired":True},
        "retrieval": {"rawPublisherFileCommitted":False,"rawRowsOrCellValuesUploadedAsArtifact":False},
        "evidenceBoundary": "Stage 2 retrieves only the exact SHA-pinned Raw Data.xlsx workbook and emits aggregate workbook shape, text labels and numeric-cell counts. No numeric measurement values or raw rows are emitted. Direct experimental inputs and measured tensile modulus, toughness and hardness must be mapped from the delivered workbook before acceptance; any class labels or derived/model fields remain excluded."
    }
    out=Path(args.output); out.parent.mkdir(parents=True,exist_ok=True); out.write_text(json.dumps(result,indent=2,ensure_ascii=False)+"\n",encoding="utf-8")
    print(json.dumps({"status":result["status"],"sha256":digest,"sheetCount":len(sheets),"totalNumericCells":total_numeric,"totalFormulaCells":total_formulas,"sheets":sheets},indent=2))


if __name__ == "__main__": main()
