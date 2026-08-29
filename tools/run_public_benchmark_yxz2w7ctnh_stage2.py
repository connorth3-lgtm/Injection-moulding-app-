#!/usr/bin/env python3
from __future__ import annotations
import argparse, hashlib, io, json, re, urllib.request
from pathlib import Path
from openpyxl import load_workbook

ROOT=Path(__file__).resolve().parents[1]
CONTRACT=ROOT/'data/public-benchmark-contracts/yxz2w7ctnh-v1.json'
DATASET_ID='yxz2w7ctnh'; VERSION=1; API_ROOT='https://api.data.mendeley.com'; UA='MouldMaster-Educational-Evidence-Profiler/1.0'

def get(url):
    req=urllib.request.Request(url,headers={'User-Agent':UA,'Accept':'*/*'})
    with urllib.request.urlopen(req,timeout=120) as r: return r.read(),r.geturl()

def normal(v): return re.sub(r'\s+',' ',str(v).strip())
def marker_counts(labels):
    text=' '.join(labels).lower()
    return {
      'injection':sum(text.count(x) for x in ['injection','injection mould','injection mold','injected']),
      'fdm':sum(text.count(x) for x in ['fdm','3d print','3d-print','printed','printing']),
      'abs':text.count('abs'), 'pla':text.count('pla'),
      'tensile':text.count('tensile'), 'bending':text.count('bend'), 'impact':text.count('impact')
    }

def main():
    ap=argparse.ArgumentParser(); ap.add_argument('--output',required=True); ap.add_argument('--retrieved-date',required=True); a=ap.parse_args()
    c=json.loads(CONTRACT.read_text()); profiles=[]; hashes_ok=True
    for f in c['stage1Evidence']['mechanicalFiles']:
        url=f"{API_ROOT}/datasets/{DATASET_ID}/files/{f['id']}/file_downloaded?version={VERSION}"
        data,final=get(url); digest=hashlib.sha256(data).hexdigest(); matched=digest.lower()==f['sha256'].lower(); hashes_ok &= matched
        wb=load_workbook(io.BytesIO(data),read_only=True,data_only=False)
        sheets=[]
        for ws in wb.worksheets:
            numeric=0; formulas=0; text_cells=0; labels=[]; nonempty=0
            for row in ws.iter_rows():
                for cell in row:
                    v=cell.value
                    if v is None: continue
                    nonempty+=1
                    if cell.data_type=='f' or (isinstance(v,str) and v.startswith('=')):
                        formulas+=1; continue
                    if isinstance(v,(int,float)) and not isinstance(v,bool): numeric+=1; continue
                    if isinstance(v,str):
                        text_cells+=1; s=normal(v)
                        if s and s not in labels and len(labels)<160: labels.append(s)
            sheets.append({'sheet':ws.title,'maxRow':ws.max_row,'maxColumn':ws.max_column,'nonEmptyCells':nonempty,'numericCells':numeric,'formulaCells':formulas,'textCells':text_cells,'textLabels':labels,'routeMaterialTestMarkers':marker_counts(labels),'rawNumericValuesEmitted':False})
        profiles.append({'fileId':f['id'],'fileName':f['filename'],'expectedSha256':f['sha256'],'sha256':digest,'publisherSha256Matched':matched,'sizeBytes':len(data),'resolvedUrl':final,'sheetCount':len(sheets),'sheets':sheets,'rawPublisherFileCommitted':False})
    result={'schema':1,'status':'mechanical-workbook-schema-profiled' if hashes_ok else 'publisher-hash-mismatch','retrievedDate':a.retrieved_date,'source':{'datasetId':c['datasetId'],'datasetDoi':c['source']['datasetDoi'],'license':c['source']['license'],'version':VERSION},'profile':{'files':profiles,'mechanicalFilesProfiled':len(profiles),'allPublisherHashesMatched':hashes_ok,'energyWorkbookRetrieved':False,'rawNumericValuesEmitted':False,'rawRowsOrArraysEmitted':False},'acceptance':{'countsAsFullyProfiledMeasuredDataset':False,'acceptedMaterialTestTraceValues':0,'acceptedInjectionProcessTimeSeriesSamples':0,'stage3Required':True},'retrieval':{'rawPublisherFilesCommitted':False,'rawRowsOrNumericValuesUploadedAsArtifact':False},'evidenceBoundary':c['evidenceBoundary']}
    Path(a.output).parent.mkdir(parents=True,exist_ok=True); Path(a.output).write_text(json.dumps(result,indent=2)+'\n')
    print(json.dumps({'status':result['status'],'hashesMatched':hashes_ok,'files':[{'file':p['fileName'],'sheets':[{'sheet':s['sheet'],'numeric':s['numericCells'],'formula':s['formulaCells'],'markers':s['routeMaterialTestMarkers']} for s in p['sheets']]} for p in profiles]},indent=2))
if __name__=='__main__': main()
