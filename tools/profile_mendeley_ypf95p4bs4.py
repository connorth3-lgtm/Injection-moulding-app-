#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import hashlib
import html
import io
import json
import re
import urllib.parse
import urllib.request
from pathlib import Path

DATASET_ID = "ypf95p4bs4"
VERSION = "1"
DOI = "10.17632/ypf95p4bs4.1"
TITLE = "Data - Model to increase availability using H-SMED, SMED and TPM in the injection molding process in a plastic company"
LICENSE = "CC BY 4.0"
PAGE = f"https://data.mendeley.com/datasets/{DATASET_ID}/{VERSION}"
API = f"https://data.mendeley.com/public-api/datasets/{DATASET_ID}/files?folder_id=root&version={VERSION}"
UA = "MouldMaster-TPM-SMED-profiler/1.0"

PLANT_MARKERS = {
    "injection", "inyect", "machine", "maquina", "máquina", "downtime", "failure", "falla", "fault",
    "maintenance", "mantenimiento", "availability", "disponibilidad", "changeover", "setup", "setup time",
    "production", "produccion", "producción", "stop", "parada", "reparacion", "reparación", "mtbf", "mttr",
    "tpm", "smed", "h-smed", "cycle", "ciclo", "operator", "operador"
}
SIM_MARKERS = {"arena", "simulation", "simulacion", "simulación", "model", "modelo", "scenario", "escenario"}


def sha256(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def get(url: str, accept: str = "*/*"):
    req = urllib.request.Request(url, headers={"User-Agent": UA, "Accept": accept})
    with urllib.request.urlopen(req, timeout=240) as response:
        return response.read(), response.headers, response.geturl()


def flatten(value):
    if isinstance(value, list):
        return value
    if isinstance(value, dict):
        for key in ("results", "files", "items", "data"):
            if isinstance(value.get(key), list):
                return value[key]
    return []


def discover():
    found = []
    api_error = None
    try:
        raw, _, _ = get(API, "application/json")
        for item in flatten(json.loads(raw.decode("utf-8"))):
            details = item.get("content_details") or {}
            url = item.get("download_url") or item.get("downloadUrl") or details.get("download_url")
            name = item.get("name") or item.get("filename") or item.get("file_name")
            if url:
                found.append({"name": name, "url": url, "publisherId": item.get("id")})
    except Exception as exc:
        api_error = str(exc)
    if found:
        return found, {"mode": "mendeley-public-api", "apiError": api_error}

    raw, _, _ = get(PAGE, "text/html")
    text = html.unescape(raw.decode("utf-8", "replace")).replace("\\u002F", "/").replace("\\/", "/")
    pattern = re.compile(rf"https://data\.mendeley\.com/public-files/datasets/{DATASET_ID}/files/([0-9a-fA-F-]{{36}})/file_downloaded")
    seen = set()
    for match in pattern.finditer(text):
        url = match.group(0)
        if url not in seen:
            seen.add(url)
            found.append({"name": None, "url": url, "publisherId": match.group(1)})
    if not found:
        raise RuntimeError(f"No version-pinned Mendeley file links discovered; API error={api_error}")
    return found, {"mode": "mendeley-page-public-files", "apiError": api_error}


def filename(final_url, headers, fallback):
    cd = headers.get("Content-Disposition")
    if cd:
        match = re.search(r"filename\*?=(?:UTF-8''|\")?([^\";]+)", cd, re.I)
        if match:
            return urllib.parse.unquote(match.group(1).strip().strip('"'))
    candidate = Path(urllib.parse.urlparse(final_url).path).name
    return candidate if Path(candidate).suffix else fallback


def text_tokens(value) -> set[str]:
    text = re.sub(r"[^a-z0-9áéíóúñ_-]+", " ", str(value).lower())
    return {x for x in text.split() if x}


def classify_text(text: str) -> dict:
    lower = text.lower()
    plant_hits = sorted({m for m in PLANT_MARKERS if m in lower})
    sim_hits = sorted({m for m in SIM_MARKERS if m in lower})
    if plant_hits and not sim_hits:
        category = "candidate-real-plant-records"
    elif sim_hits and not plant_hits:
        category = "simulation-or-model-output"
    elif plant_hits and sim_hits:
        category = "mixed-plant-and-simulation-review-required"
    else:
        category = "unresolved"
    return {"category": category, "plantMarkerHits": plant_hits, "simulationMarkerHits": sim_hits}


def profile_csv(raw: bytes) -> dict:
    text = raw.decode("utf-8-sig", "replace")
    try:
        dialect = csv.Sniffer().sniff(text[:65536], delimiters=",;\t|")
    except Exception:
        dialect = csv.excel
    reader = csv.reader(io.StringIO(text), dialect)
    try:
        header = [str(x).strip() for x in next(reader)]
    except StopIteration:
        return {"rows": 0, "columns": 0, "header": [], "preview": [], "semanticClassification": classify_text("")}
    rows = []
    count = 0
    for row in reader:
        if not row or not any(str(x).strip() for x in row):
            continue
        count += 1
        if len(rows) < 8:
            rows.append([str(x)[:160] for x in row[:40]])
    semantics = classify_text(" ".join(header) + " " + " ".join(" ".join(r) for r in rows))
    return {"rows": count, "columns": len(header), "header": header, "preview": rows, "delimiter": repr(dialect.delimiter), "semanticClassification": semantics}


def profile_xlsx(raw: bytes) -> dict:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    sheets = []
    for ws in wb.worksheets:
        nonempty_rows = 0
        max_width = 0
        preview = []
        header_candidate = []
        for row in ws.iter_rows(values_only=True):
            values = ["" if v is None else str(v) for v in row]
            if not any(v.strip() for v in values):
                continue
            nonempty_rows += 1
            max_width = max(max_width, len(values))
            if not header_candidate:
                header_candidate = [v.strip()[:160] for v in values[:60]]
            if len(preview) < 10:
                preview.append([v[:160] for v in values[:60]])
        body_rows = max(0, nonempty_rows - 1)
        semantic_text = ws.title + " " + " ".join(" ".join(r) for r in preview)
        sheets.append({
            "name": ws.title,
            "nonEmptyRows": nonempty_rows,
            "bodyRowsApprox": body_rows,
            "maxColumnsObserved": max_width,
            "firstNonEmptyRow": header_candidate,
            "preview": preview,
            "semanticClassification": classify_text(semantic_text),
        })
    return {"format": "xlsx", "sheets": sheets}


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", default="mendeley-ypf95p4bs4-v1.json")
    args = parser.parse_args()

    records, acquisition = discover()
    files = []
    for index, record in enumerate(records, 1):
        raw, headers, final_url = get(record["url"])
        name = record.get("name") or filename(final_url, headers, f"publisher-file-{index}")
        suffix = Path(name).suffix.lower()
        item = {
            "name": name,
            "publisherId": record.get("publisherId"),
            "sizeBytes": len(raw),
            "sha256": sha256(raw),
            "suffix": suffix,
        }
        if suffix in {".xlsx", ".xlsm"}:
            item.update(profile_xlsx(raw))
        elif suffix in {".csv", ".tsv", ".txt"}:
            item.update({"format": suffix.lstrip("."), "table": profile_csv(raw)})
        else:
            item["format"] = suffix.lstrip(".") or "unknown"
            item["semanticClassification"] = classify_text(name)
        files.append(item)

    sheet_classes = []
    for f in files:
        for s in f.get("sheets") or []:
            sheet_classes.append({"file": f["name"], "sheet": s["name"], **s["semanticClassification"], "rows": s["bodyRowsApprox"], "columns": s["maxColumnsObserved"]})
        if f.get("table"):
            t = f["table"]
            sheet_classes.append({"file": f["name"], "sheet": None, **t["semanticClassification"], "rows": t["rows"], "columns": t["columns"]})
        elif not f.get("sheets"):
            sheet_classes.append({"file": f["name"], "sheet": None, **f.get("semanticClassification", classify_text(f["name"])), "rows": None, "columns": None})

    real_candidates = [x for x in sheet_classes if x["category"] == "candidate-real-plant-records"]
    mixed = [x for x in sheet_classes if x["category"] == "mixed-plant-and-simulation-review-required"]
    simulation = [x for x in sheet_classes if x["category"] == "simulation-or-model-output"]

    payload = {
        "schema": 1,
        "status": "exact-file-profile-review-required",
        "completedDate": "2026-08-28",
        "source": {
            "datasetId": DATASET_ID,
            "version": VERSION,
            "doi": DOI,
            "title": TITLE,
            "publisher": "Mendeley Data",
            "page": PAGE,
            "license": LICENSE,
            "publishedContext": "Injection-moulding plastic-company availability study using H-SMED, SMED and TPM; source package also contains Arena simulation results.",
        },
        "acquisition": acquisition,
        "files": files,
        "fileCount": len(files),
        "tableOrSheetClassifications": sheet_classes,
        "candidateRealPlantTablesOrSheets": len(real_candidates),
        "mixedPlantSimulationTablesOrSheets": len(mixed),
        "simulationTablesOrSheets": len(simulation),
        "acceptedMeasuredRecords": 0,
        "acceptedMeasuredTimeSeriesSamples": 0,
        "rawSourceRowsCommitted": False,
        "rawSourceFilesCommitted": False,
        "boundary": "This first exact-file pass is discovery-only. Arena simulation/model outputs never count as measured evidence. Promotion requires an exact spreadsheet/table that can be identified as source plant observations, with record unit, dates/grouping, machine/maintenance/changeover/production semantics and separation from derived KPI or simulation results."
    }
    Path(args.output).write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(json.dumps({
        "acquisition": acquisition,
        "fileCount": len(files),
        "files": [(f["name"], f["format"], f["sizeBytes"], f["sha256"]) for f in files],
        "candidateRealPlantTablesOrSheets": real_candidates,
        "mixedPlantSimulationTablesOrSheets": mixed,
        "simulationTablesOrSheets": simulation,
    }, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
