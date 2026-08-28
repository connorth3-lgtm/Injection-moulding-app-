#!/usr/bin/env python3
from __future__ import annotations
import argparse, io, json, sys
from datetime import datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook
from tools.profile_mendeley_ypf95p4bs4 import discover, get, filename, sha256, DOI, LICENSE

WORKBOOK='TF-IN92-241-G198-Excel.xlsx'
EXPECTED_SHA='b8bdf20f4e513d647ef4fc887f8846c254d0151657137a888f20752559ce1917'


def norm(v):
    return '' if v is None else str(v).strip()


def dt(v):
    if isinstance(v, datetime): return v
    text=norm(v)
    if not text: return None
    for fmt in ('%Y-%m-%d %H:%M:%S','%Y-%m-%d'):
        try: return datetime.strptime(text,fmt)
        except ValueError: pass
    try: return datetime.fromisoformat(text)
    except ValueError: return None


def num(v):
    try: return float(v)
    except Exception: return None


def find_header(rows, required):
    for i,row in enumerate(rows):
        vals=[norm(x) for x in row]
        if all(any(r.lower()==x.lower() for x in vals) for r in required):
            return i, vals
    raise RuntimeError(f'header not found: {required}')


def profile_sheet(ws, required, record_type, machine_filter=None):
    rows=[list(r) for r in ws.iter_rows(values_only=True)]
    hi, header=find_header(rows, required)
    index={h.lower():i for i,h in enumerate(header) if h}
    accepted=[]
    for row in rows[hi+1:]:
        values=list(row)+[None]*max(0,len(header)-len(row))
        when=dt(values[index['fecha']])
        if not when: continue
        if machine_filter:
            machine=norm(values[index['máquina'] if 'máquina' in index else index['maquina']])
            if machine.lower()!=machine_filter.lower(): continue
        if record_type=='setup':
            operation=norm(values[index['operación'] if 'operación' in index else index['operacion']])
            hours=num(values[index['tiempo(h)']])
            if not operation or hours is None: continue
        else:
            motive=norm(values[index['motivo']])
            detail=norm(values[index['detalle']])
            minutes=num(values[index['tiempo (m)']])
            hours=num(values[index['tiempo (h)']])
            if not motive or not detail or minutes is None or hours is None: continue
        accepted.append(when)
    return {
        'sheet': ws.title,
        'recordType': record_type,
        'acceptedRows': len(accepted),
        'firstDate': min(accepted).date().isoformat() if accepted else None,
        'lastDate': max(accepted).date().isoformat() if accepted else None,
        'machineFilter': machine_filter,
        'headerRowNumber': hi+1,
    }


def main():
    ap=argparse.ArgumentParser(); ap.add_argument('--output',default='tpm-smed-primary-v1.json'); args=ap.parse_args()
    records,_=discover(); selected=None; raw=None
    for i,rec in enumerate(records,1):
        b,h,u=get(rec['url']); name=rec.get('name') or filename(u,h,f'file-{i}')
        if name==WORKBOOK:
            selected={'name':name,'sizeBytes':len(b),'sha256':sha256(b),'publisherId':rec.get('publisherId')}; raw=b; break
    if raw is None: raise RuntimeError('primary workbook not found')
    if selected['sha256']!=EXPECTED_SHA: raise RuntimeError('workbook SHA-256 drift')
    wb=load_workbook(io.BytesIO(raw),read_only=True,data_only=True)
    profiles=[
        profile_sheet(wb['Paradas de Maquinaria 2023'],['Fecha','Máquina','Motivo','Detalle','Tiempo (m)','Tiempo (h)'],'injector-downtime-event','Inyectora'),
        profile_sheet(wb['Limpieza 2023 (Mtto)'],['Fecha','Máquina','Motivo','Detalle','Tiempo (m)','Tiempo (h)'],'injector-maintenance-event','Inyectora'),
        profile_sheet(wb['Setup 2023'],['Fecha','Operación','Tipo','Tiempo(h)'],'setup'),
    ]
    total=sum(x['acceptedRows'] for x in profiles)
    payload={
      'schema':1,'status':'completed-public-measured-record-benchmark' if total>0 else 'review-required','completedDate':'2026-08-28',
      'source':{'doi':DOI,'license':LICENSE,'publisher':'Mendeley Data','workbook':selected},
      'acceptedPrimaryTables':profiles,'acceptedMeasuredRecords':total,'acceptedMeasuredTimeSeriesSamples':0,
      'excludedSheets':['VINCULACION','Hoja3','Resumen Tiempos Improductivos','Resumen paradas de maquinaria','Disponibilidad Inyectora','Impacto Económico','Arbol de problemas','TPM','TPM 1','TPM 2','SMED','HSMED','HSMED2','Indicadores','PRESUPUESTO','Flujo de Caja','VALIDACION','VALIDACION 2','Validación no económica','MTBF - MTTR','Tiempos de Setup (Westinghouse)'],
      'rawSourceRowsCommitted':False,'rawSourceFilesCommitted':False,
      'boundary':'Only dated raw injector downtime, injector maintenance/cleaning, and mould-change/setup rows count. Summary pivots, KPI/availability/MTBF-MTTR calculations, standards, budgets, validation/scenario/model sheets and undated observational timing studies are excluded.'
    }
    Path(args.output).write_text(json.dumps(payload,indent=2,ensure_ascii=False)+'\n',encoding='utf-8')
    print(json.dumps({'status':payload['status'],'workbook':selected,'tables':profiles,'acceptedMeasuredRecords':total},indent=2,ensure_ascii=False))
if __name__=='__main__': main()
