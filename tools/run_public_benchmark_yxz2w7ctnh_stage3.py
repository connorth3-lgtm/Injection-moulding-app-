#!/usr/bin/env python3
from __future__ import annotations
import argparse, collections, hashlib, io, json, re, urllib.request
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

ROOT=Path(__file__).resolve().parents[1]; CONTRACT=ROOT/'data/public-benchmark-contracts/yxz2w7ctnh-v1.json'
DATASET_ID='yxz2w7ctnh'; PUBLIC_FILE_ROOT=f'https://data.mendeley.com/public-files/datasets/{DATASET_ID}/files'; UA='MouldMaster-Educational-Evidence-Profiler/1.0'

def get(url):
    req=urllib.request.Request(url,headers={'User-Agent':UA,'Accept':'*/*'})
    with urllib.request.urlopen(req,timeout=120) as r: return r.read(),r.geturl()

def clean(v): return re.sub(r'\s+',' ',str(v).strip())
def route_kind(text):
    low=text.lower()
    if any(x in low for x in ['injection','injection mould','injection mold','injected','inkection']): return 'injection'
    if any(x in low for x in ['3d print','3dprint','3d-print','fdm','printed','printing']): return 'fdm'
    return None

def main():
    ap=argparse.ArgumentParser(); ap.add_argument('--output',required=True); ap.add_argument('--retrieved-date',required=True); a=ap.parse_args()
    c=json.loads(CONTRACT.read_text()); by_name={f['filename']:f for f in c['stage1Evidence']['mechanicalFiles']}; canonical=c['stage2Evidence']['canonicalTestSheets']; out=[]; hashes_ok=True
    for family,spec in canonical.items():
        f=by_name[spec['file']]; url=f"{PUBLIC_FILE_ROOT}/{f['id']}/file_downloaded"; data,final=get(url); digest=hashlib.sha256(data).hexdigest(); matched=digest.lower()==f['sha256'].lower(); hashes_ok &= matched
        wb=load_workbook(io.BytesIO(data),read_only=True,data_only=False)
        for sheet_name in spec['sheets']:
            ws=wb[sheet_name]; texts=[]; anchors=[]; num_by_col=collections.Counter(); formula_by_col=collections.Counter(); num_by_row=collections.Counter()
            for row in ws.iter_rows():
                for cell in row:
                    v=cell.value
                    if v is None: continue
                    if cell.data_type=='f' or (isinstance(v,str) and v.startswith('=')):
                        formula_by_col[get_column_letter(cell.column)]+=1; continue
                    if isinstance(v,(int,float)) and not isinstance(v,bool):
                        num_by_col[get_column_letter(cell.column)]+=1; num_by_row[str(cell.row)]+=1; continue
                    if isinstance(v,str):
                        s=clean(v)
                        if len(texts)<240: texts.append({'cell':cell.coordinate,'text':s})
                        rk=route_kind(s)
                        if rk: anchors.append({'cell':cell.coordinate,'route':rk,'text':s})
            out.append({'testFamily':family,'fileName':f['filename'],'sheet':sheet_name,'publisherSha256Matched':matched,'resolvedUrl':final,'maxRow':ws.max_row,'maxColumn':ws.max_column,'routeAnchors':anchors,'textCells':texts,'numericCellsByColumn':dict(sorted(num_by_col.items())),'formulaCellsByColumn':dict(sorted(formula_by_col.items())),'numericCellsByRow':dict(sorted(num_by_row.items(),key=lambda x:int(x[0]))),'rawNumericValuesEmitted':False})
    result={'schema':1,'status':'route-coordinate-profiled' if hashes_ok else 'publisher-hash-mismatch','retrievedDate':a.retrieved_date,'source':{'datasetId':c['datasetId'],'datasetDoi':c['source']['datasetDoi'],'license':c['source']['license']},'profile':{'canonicalSheets':out,'canonicalSheetCount':len(out),'allPublisherHashesMatched':hashes_ok,'duplicateEmbeddedSheetsCounted':False,'energySheetsCounted':False,'rawNumericValuesEmitted':False},'acceptance':{'countsAsFullyProfiledMeasuredDataset':False,'acceptedMaterialTestTraceValues':0,'acceptedInjectionProcessTimeSeriesSamples':0,'stage4Required':True},'retrieval':{'rawPublisherFilesCommitted':False,'rawRowsOrNumericValuesUploadedAsArtifact':False},'evidenceBoundary':c['evidenceBoundary']}
    Path(a.output).parent.mkdir(parents=True,exist_ok=True); Path(a.output).write_text(json.dumps(result,indent=2)+'\n')
    print(json.dumps({'status':result['status'],'sheets':[{'family':s['testFamily'],'sheet':s['sheet'],'anchors':s['routeAnchors'],'numericCellsByColumn':s['numericCellsByColumn']} for s in out]},indent=2))
if __name__=='__main__': main()
