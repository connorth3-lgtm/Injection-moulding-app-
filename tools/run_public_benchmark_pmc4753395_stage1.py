#!/usr/bin/env python3
from __future__ import annotations

import argparse
import hashlib
import io
import json
import re
import subprocess
import tempfile
import urllib.parse
import urllib.request
import zipfile
from pathlib import Path, PurePosixPath

import openpyxl
import pandas as pd

ROOT = Path(__file__).resolve().parents[1]
CONTRACT = ROOT / "data/public-benchmark-contracts/pmc4753395-hdpe-cenosphere-v1.json"
UA = "MouldMaster-Educational-Evidence-Profiler/1.0"


def get_bytes(url: str, accept: str = "*/*"):
    req = urllib.request.Request(url, headers={"Accept": accept, "User-Agent": UA})
    with urllib.request.urlopen(req, timeout=120) as r:
        return r.read(), r.geturl(), r.headers.get("Content-Type")


def get_json(url: str):
    raw, final, content_type = get_bytes(url, "application/json,*/*")
    return json.loads(raw.decode("utf-8")), final, content_type, hashlib.sha256(raw).hexdigest()


def safe_member(name: str) -> bool:
    p = PurePosixPath(name)
    return not p.is_absolute() and ".." not in p.parts


def compact_text(v):
    if not isinstance(v, str):
        return None
    s = re.sub(r"\s+", " ", v).strip()
    return s[:160] if s else None


def s3_to_https(url: str) -> str:
    if not url.startswith("s3://pmc-oa-opendata/"):
        raise RuntimeError("unexpected PMC cloud media URL")
    return "https://pmc-oa-opendata.s3.amazonaws.com/" + url[len("s3://pmc-oa-opendata/"):]


def cloud_supplement(metadata_url: str, pmcid: str, supplement_label: str):
    meta, final, content_type, metadata_sha = get_json(metadata_url)
    if meta.get("pmcid") != pmcid:
        raise RuntimeError(f"PMC cloud metadata PMCID mismatch: {meta.get('pmcid')!r}")
    license_code = str(meta.get("license_code") or "").upper().replace("-", " ")
    if license_code != "CC BY":
        raise RuntimeError(f"PMC cloud metadata licence is not CC BY: {meta.get('license_code')!r}")
    media_urls = meta.get("media_urls") or []
    target = supplement_label.lower()
    candidates = [str(u) for u in media_urls if Path(urllib.parse.urlparse(str(u)).path).name.lower() == target]
    if len(candidates) != 1:
        raise RuntimeError(f"expected exactly one cloud media object {supplement_label!r}; found {len(candidates)}")
    s3_url = candidates[0]
    https_url = s3_to_https(s3_url)
    payload, payload_final, payload_type = get_bytes(https_url, "application/zip,*/*")
    return payload, {
        "metadataUrl": metadata_url,
        "metadataResolvedUrl": final,
        "metadataContentType": content_type,
        "metadataSha256": metadata_sha,
        "metadataPmcid": meta.get("pmcid"),
        "metadataVersion": meta.get("version"),
        "metadataDoi": meta.get("doi"),
        "metadataLicenseCode": meta.get("license_code"),
        "mediaObjectCount": len(media_urls),
        "supplementS3Url": s3_url,
        "supplementHttpsUrl": https_url,
        "supplementResolvedUrl": payload_final,
        "supplementContentType": payload_type,
        "supplementSizeBytes": len(payload),
        "supplementSha256": hashlib.sha256(payload).hexdigest(),
        "rawCloudMetadataOrMediaUploaded": False
    }


def profile_xlsx(data: bytes, member_name: str):
    with tempfile.NamedTemporaryFile(suffix=".xlsx") as f:
        f.write(data); f.flush()
        raw_sheets = pd.read_excel(f.name, sheet_name=None, header=None)
        wb_formula = openpyxl.load_workbook(f.name, data_only=False, read_only=True)
        out = []
        for sheet_name, df in raw_sheets.items():
            df = df.dropna(axis=0, how="all").dropna(axis=1, how="all")
            numeric_cells, text_labels, seen = 0, [], set()
            for value in df.to_numpy().ravel().tolist():
                if isinstance(value, bool):
                    continue
                if isinstance(value, (int, float)) and not pd.isna(value):
                    numeric_cells += 1
                else:
                    label = compact_text(value)
                    if label and label not in seen:
                        seen.add(label); text_labels.append(label)
            formula_cells = 0
            if sheet_name in wb_formula.sheetnames:
                for row in wb_formula[sheet_name].iter_rows():
                    formula_cells += sum(1 for cell in row if cell.data_type == "f")
            out.append({
                "sheet": str(sheet_name), "rows": int(len(df)), "columns": int(len(df.columns)),
                "numericCells": int(numeric_cells), "formulaCells": int(formula_cells),
                "textLabels": text_labels[:120], "rawNumericValuesEmitted": False
            })
        return {"member": member_name, "sheetCount": len(out), "sheets": out, "rawRowsOrNumericValuesEmitted": False}


def parse_7z_listing(text: str):
    entries, current = [], {}
    for line in text.splitlines() + [""]:
        if not line.strip():
            if current.get("Path") and current.get("Folder", "-") != "+":
                entries.append(current)
            current = {}
            continue
        if " = " in line:
            k, v = line.split(" = ", 1)
            current[k.strip()] = v.strip()
    return entries


def profile_nested_rar(rar_bytes: bytes, rar_name: str):
    with tempfile.TemporaryDirectory(prefix="mouldmaster-pmc-rar-") as td:
        rar_path = Path(td) / "source.rar"
        rar_path.write_bytes(rar_bytes)
        listing = subprocess.run(["7z", "l", "-slt", "-ba", str(rar_path)], check=True, capture_output=True, text=True)
        entries = parse_7z_listing(listing.stdout)
        nested_members, workbooks = [], []
        for entry in entries:
            name = entry.get("Path", "")
            if not safe_member(name):
                raise RuntimeError(f"unsafe nested RAR member: {name}")
            ext = Path(name).suffix.lower()
            item = {
                "name": name,
                "extension": ext,
                "sizeBytes": int(entry.get("Size") or 0),
                "packedSizeBytes": int(entry.get("Packed Size") or 0),
                "isMeasuredCandidateByName": Path(name).name.lower() == "tensile-data.xlsx",
                "isTheoreticalModelByName": "porfiri" in Path(name).name.lower(),
                "rawPayloadEmitted": False
            }
            if ext == ".xlsx":
                extracted = subprocess.run(["7z", "x", "-so", str(rar_path), name], check=True, capture_output=True).stdout
                item["sha256"] = hashlib.sha256(extracted).hexdigest()
                item["sizeBytes"] = len(extracted)
                workbooks.append(profile_xlsx(extracted, f"{rar_name}::{name}"))
            nested_members.append(item)
        return {
            "archive": rar_name,
            "memberCount": len(nested_members),
            "members": nested_members,
            "workbooks": workbooks,
            "rawNestedArchiveCommittedOrUploaded": False,
            "rawNestedMembersEmitted": False
        }


def main():
    ap = argparse.ArgumentParser(); ap.add_argument("--output", type=Path, required=True); ap.add_argument("--retrieved-date", required=True); args = ap.parse_args()
    contract = json.loads(CONTRACT.read_text(encoding="utf-8"))
    url = contract["source"]["supplementUrl"]; supplement_label = contract["source"]["supplementLabel"]
    direct_data, direct_final, direct_type = get_bytes(url, "application/zip,*/*")
    retrieval_route, cloud_diagnostics = "direct-supplement-url", None
    if zipfile.is_zipfile(io.BytesIO(direct_data)):
        data, final_url, content_type = direct_data, direct_final, direct_type
    else:
        data, cloud_diagnostics = cloud_supplement(contract["source"]["cloudMetadataUrl"], contract["source"]["pmcid"], supplement_label)
        final_url = cloud_diagnostics["supplementResolvedUrl"]; content_type = cloud_diagnostics["supplementContentType"]; retrieval_route = "pmc-current-cloud-media"

    archive_sha = hashlib.sha256(data).hexdigest(); valid_zip = zipfile.is_zipfile(io.BytesIO(data))
    members, workbook_profiles, nested_archives = [], [], []
    if valid_zip:
        with zipfile.ZipFile(io.BytesIO(data)) as zf:
            for info in zf.infolist():
                if info.is_dir(): continue
                if not safe_member(info.filename): raise RuntimeError(f"unsafe archive member: {info.filename}")
                payload = zf.read(info); ext = Path(info.filename).suffix.lower()
                members.append({
                    "name": info.filename, "extension": ext, "sizeBytes": len(payload),
                    "compressedSizeBytes": int(info.compress_size), "crc32": f"{info.CRC:08x}",
                    "sha256": hashlib.sha256(payload).hexdigest(),
                    "isMeasuredCandidateByName": Path(info.filename).name.lower() == "tensile-data.xlsx",
                    "isTheoreticalModelByName": "porfiri" in Path(info.filename).name.lower(), "rawPayloadEmitted": False
                })
                if ext == ".xlsx": workbook_profiles.append(profile_xlsx(payload, info.filename))
                elif ext == ".rar":
                    nested = profile_nested_rar(payload, info.filename); nested_archives.append(nested); workbook_profiles.extend(nested["workbooks"])

    all_names = [Path(x["name"]).name.lower() for x in members]
    for nested in nested_archives:
        all_names.extend(Path(x["name"]).name.lower() for x in nested["members"])
    measured_present = "tensile-data.xlsx" in all_names
    theoretical_present = any("porfiri" in n and n.endswith(".xlsx") for n in all_names)
    status = "retrieved-profile-needs-semantic-review" if valid_zip and measured_present else "retrieval-or-structure-blocked"
    result = {
        "schema": 1, "status": status, "retrievedDate": args.retrieved_date,
        "source": {
            "datasetId": contract["datasetId"], "datasetDoi": contract["source"]["datasetDoi"], "pmcid": contract["source"]["pmcid"],
            "license": contract["source"]["license"], "supplementUrl": url, "retrievalRoute": retrieval_route,
            "directAttempt": {"resolvedUrl": direct_final, "contentType": direct_type, "sizeBytes": len(direct_data), "sha256": hashlib.sha256(direct_data).hexdigest(), "zipStructureValid": zipfile.is_zipfile(io.BytesIO(direct_data)), "responseBodyEmitted": False},
            "cloudFallback": cloud_diagnostics, "resolvedSupplementLocation": final_url, "contentType": content_type,
            "retrievedSizeBytes": len(data), "sha256": archive_sha, "zipStructureValid": valid_zip
        },
        "archiveProfile": {
            "memberCount": len(members), "members": members, "nestedArchives": nested_archives,
            "measuredWorkbookPresent": measured_present, "theoreticalWorkbookPresent": theoretical_present,
            "workbooks": workbook_profiles, "rawArchiveCommitted": False, "rawMembersUploadedAsArtifact": False, "rawRowsOrNumericValuesEmitted": False
        },
        "acceptance": {"stage1ProfileComplete": status == "retrieved-profile-needs-semantic-review", "countsAsFullyProfiledMeasuredDataset": False, "acceptedMeasuredTimeSeriesSamples": 0, "stage2SemanticReviewRequired": status == "retrieved-profile-needs-semantic-review"},
        "evidenceBoundary": contract["evidenceBoundary"]
    }
    args.output.parent.mkdir(parents=True, exist_ok=True); args.output.write_text(json.dumps(result, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(json.dumps({"status": status, "retrievalRoute": retrieval_route, "archiveSha256": archive_sha, "sizeBytes": len(data), "memberCount": len(members), "nestedArchiveCount": len(nested_archives), "measuredWorkbookPresent": measured_present, "theoreticalWorkbookPresent": theoretical_present, "workbookNames": [x["member"] for x in workbook_profiles]}, indent=2))


if __name__ == "__main__": main()
