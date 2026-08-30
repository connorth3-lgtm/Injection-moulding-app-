#!/usr/bin/env python3
from __future__ import annotations

import argparse
import hashlib
import json
import tempfile
import urllib.request
from pathlib import Path
from openpyxl import load_workbook

DATASET_ID = "6k8fpbrd9s"
VERSION = 1
DOI = "10.17632/6k8fpbrd9s.1"
EXPECTED_FILE_ID = "8598d42d-f794-47e2-ad84-dd952c900d27"
EXPECTED_FILE = "Data.xlsx"
EXPECTED_SHA256 = "14c056dd86e11cc47e1e97834174631f9dc0806442917a7149ddf5856dc9b11c"
PUBLIC_FILES_ENDPOINT = f"https://data.mendeley.com/public-api/datasets/{DATASET_ID}/files?folder_id=root&version={VERSION}"
API_ROOT = "https://api.data.mendeley.com"
UA = "MouldMaster-Educational-Evidence-Profiler/1.0"


def get(url: str, accept: str = "*/*"):
    req = urllib.request.Request(url, headers={"Accept": accept, "User-Agent": UA})
    with urllib.request.urlopen(req, timeout=120) as r:
        return r.read(), r.geturl()


def flatten_files(payload):
    if isinstance(payload, list):
        return payload
    if isinstance(payload, dict):
        for key in ("results", "files", "items", "data"):
            if isinstance(payload.get(key), list):
                return payload[key]
    return []


def file_id(item):
    return str(item.get("id") or item.get("file_id") or item.get("uuid") or "").strip()


def file_name(item):
    return str(item.get("filename") or item.get("name") or "").strip()


def file_url(item):
    details = item.get("content_details") or item.get("contentDetails") or {}
    for candidate in (details.get("download_url"), details.get("downloadUrl"), item.get("download_url"), item.get("downloadUrl")):
        if candidate:
            return str(candidate)
    fid = file_id(item)
    return f"{API_ROOT}/datasets/{DATASET_ID}/files/{fid}/file_downloaded?version={VERSION}"


def aggregate_sheet(ws):
    text_cells = []
    numeric_by_row = {}
    numeric_by_column = {}
    formula_count = 0
    nonempty = 0
    for row_index, row in enumerate(ws.iter_rows(), start=1):
        row_numeric = 0
        for cell in row:
            value = cell.value
            if value is None:
                continue
            nonempty += 1
            if isinstance(value, str):
                if value.startswith("="):
                    formula_count += 1
                else:
                    cleaned = " ".join(value.split())
                    if cleaned:
                        coordinate = getattr(cell, "coordinate", None)
                        if coordinate:
                            text_cells.append({"cell": coordinate, "text": cleaned[:240]})
            elif isinstance(value, (int, float)) and not isinstance(value, bool):
                row_numeric += 1
                col = getattr(cell, "column_letter", None)
                if col:
                    numeric_by_column[col] = numeric_by_column.get(col, 0) + 1
        if row_numeric:
            numeric_by_row[row_index] = row_numeric
    labels = [x["text"].lower() for x in text_cells]
    markers = {
        "pressure": any("pressure" in x or x.strip() in {"p", "p/mpa", "p [mpa]"} for x in labels),
        "temperature": any("temperature" in x or x.strip() in {"t", "t/°c", "t [°c]", "t/c"} for x in labels),
        "specificVolume": any("specific volume" in x or "specific volume" in x.replace("_", " ") or "vsp" in x for x in labels),
        "cooling": any("cool" in x for x in labels),
        "heating": any("heat" in x for x in labels),
        "compression": any("compress" in x for x in labels),
        "decompression": any("decompress" in x for x in labels),
        "rate": any("rate" in x for x in labels),
    }
    return {
        "title": ws.title,
        "maxRow": ws.max_row,
        "maxColumn": ws.max_column,
        "nonEmptyCellCount": nonempty,
        "textCells": text_cells,
        "textCellCount": len(text_cells),
        "numericCellsByRow": [{"row": r, "numericCells": c} for r, c in sorted(numeric_by_row.items())],
        "numericCellsByColumn": numeric_by_column,
        "numericCellCount": sum(numeric_by_row.values()),
        "formulaCellCount": formula_count,
        "semanticMarkers": markers,
        "numericValuesEmitted": False,
    }


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--output", required=True)
    ap.add_argument("--retrieved-date", required=True)
    args = ap.parse_args()

    raw, _ = get(PUBLIC_FILES_ENDPOINT, "application/json")
    files = flatten_files(json.loads(raw.decode("utf-8")))
    matches = [x for x in files if file_id(x) == EXPECTED_FILE_ID and file_name(x) == EXPECTED_FILE]
    if len(matches) != 1:
        raise RuntimeError(f"exact pvT publisher file not found: {[(file_id(x), file_name(x)) for x in files]}")
    item = matches[0]
    data, final_url = get(file_url(item))
    digest = hashlib.sha256(data).hexdigest()
    if digest != EXPECTED_SHA256:
        raise RuntimeError(f"publisher file SHA-256 drifted: {digest}")

    with tempfile.TemporaryDirectory() as td:
        path = Path(td) / "Data.xlsx"
        path.write_bytes(data)
        wb = load_workbook(path, read_only=True, data_only=False)
        sheets = [aggregate_sheet(ws) for ws in wb.worksheets]

    result = {
        "schema": 1,
        "status": "retrieved-profile-needs-semantic-review",
        "retrievedDate": args.retrieved_date,
        "source": {
            "datasetId": "mendeley-6k8fpbrd9s-v1",
            "datasetDoi": DOI,
            "version": VERSION,
            "license": "CC BY 4.0",
            "publisherFileId": EXPECTED_FILE_ID,
            "publisherFileName": EXPECTED_FILE,
            "sha256": digest,
            "publisherSha256Matched": True,
            "retrievedSizeBytes": len(data),
            "resolvedUrl": final_url
        },
        "profile": {
            "sheetCount": len(sheets),
            "sheets": sheets,
            "totalNumericCells": sum(x["numericCellCount"] for x in sheets),
            "totalFormulaCells": sum(x["formulaCellCount"] for x in sheets),
            "rawRowsOrCellValuesEmitted": False,
            "numericMeasurementValuesEmitted": False
        },
        "acceptance": {
            "countsAsFullyProfiledMeasuredDataset": False,
            "acceptedMeasuredTimeSeriesSamples": 0,
            "semanticReviewRequired": True,
            "injectionMouldingCycleDataset": False
        },
        "retrieval": {
            "rawPublisherFileCommitted": False,
            "rawRowsOrCellValuesUploadedAsArtifact": False
        },
        "evidenceBoundary": "Exact CC BY 4.0 pvT workbook retrieved and SHA-256 matched. The aggregate artifact exposes text labels, sheet structure, formula counts and numeric-cell counts only; no numeric measurements are emitted. Direct pressure/temperature/specific-volume measurements remain non-counting until the delivered sheet layout is semantically mapped and fitted/derived fields are excluded."
    }
    out = Path(args.output)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(result, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(json.dumps({"status": result["status"], "sha256": digest, "sheetCount": len(sheets), "totalNumericCells": result["profile"]["totalNumericCells"], "sheetTitles": [x["title"] for x in sheets]}, indent=2))


if __name__ == "__main__":
    main()
