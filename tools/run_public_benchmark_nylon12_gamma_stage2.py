#!/usr/bin/env python3
from __future__ import annotations
import argparse, hashlib, io, json, re, urllib.request, zipfile
from pathlib import Path, PurePosixPath
import xml.etree.ElementTree as ET
import openpyxl

ROOT=Path(__file__).resolve().parents[1]
CONTRACT=ROOT/'data/public-benchmark-contracts/nylon12-gamma-8c8fjwcw86-v1.json'
UA='MouldMaster-Educational-Evidence-Profiler/1.0'
PUBLIC_BASE='https://data.mendeley.com/public-files/datasets/8c8fjwcw86/files'


def get(url):
    req=urllib.request.Request(url,headers={'Accept':'*/*','User-Agent':UA})
    with urllib.request.urlopen(req,timeout=180) as r: return r.read(),r.geturl(),r.headers.get('Content-Type')

def safe(name):
    p=PurePosixPath(name); return not p.is_absolute() and '..' not in p.parts

def text_tokens(xml_bytes):
    try: root=ET.fromstring(xml_bytes)
    except ET.ParseError: return []
    out=[]
    for elem in root.iter():
        if elem.text and elem.text.strip(): out.append(' '.join(elem.text.split()))
    return out

def route_markers(strings):
    joined=' '.join(strings).lower()
    return {'injection':sum(joined.count(x) for x in ['injection mold','injection mould','injection-mold','injection_mold']),'sls':sum(joined.count(x) for x in ['selective laser sinter','sls'])}

def xlsx_profile(data,name):
    with io.BytesIO(data) as bio:
        wb=openpyxl.load_workbook(bio,data_only=False,read_only=True)
        sheets=[]
        for ws in wb.worksheets:
            numeric=0; formulas=0; labels=[]; seen=set()
            for row in ws.iter_rows():
                for cell in row:
                    v=cell.value
                    if isinstance(v,bool) or v is None: continue
                    if cell.data_type=='f': formulas+=1
                    elif isinstance(v,(int,float)): numeric+=1
                    elif isinstance(v,str):
                        s=' '.join(v.split())[:120]
                        if s and s not in seen: seen.add(s); labels.append(s)
            sheets.append({'sheet':ws.title,'maxRow':ws.max_row,'maxColumn':ws.max_column,'numericCells':numeric,'formulaCells':formulas,'routeMarkers':route_markers(labels),'textLabels':labels[:60],'rawNumericValuesEmitted':False})
    return {'member':name,'sha256':hashlib.sha256(data).hexdigest(),'sheets':sheets,'rawWorkbookUploaded':False}

def chart_profile(xml_bytes,name):
    strings=text_tokens(xml_bytes); root=ET.fromstring(xml_bytes)
    num_points=sum(1 for e in root.iter() if e.tag.endswith('}pt') and any(a.tag.endswith('}v') for a in e))
    labels=[]
    for s in strings:
        if len(s)<=120 and not re.fullmatch(r'[-+0-9.eE%]+',s): labels.append(s)
    return {'member':name,'cachedPointElements':num_points,'routeMarkers':route_markers(labels),'textLabels':list(dict.fromkeys(labels))[:60],'rawNumericValuesEmitted':False}

def package_profile(data,filename):
    if not zipfile.is_zipfile(io.BytesIO(data)): raise RuntimeError(f'{filename} is not valid OOXML ZIP')
    members=[]; embeddings=[]; charts=[]; route_strings=[]
    with zipfile.ZipFile(io.BytesIO(data)) as z:
        for info in z.infolist():
            if info.is_dir(): continue
            if not safe(info.filename): raise RuntimeError('unsafe OOXML member')
            lower=info.filename.lower(); payload=z.read(info)
            members.append({'name':info.filename,'sizeBytes':len(payload),'sha256':hashlib.sha256(payload).hexdigest()})
            if lower.endswith('.xml') and ('/slides/' in lower or '/charts/' in lower or lower.endswith('document.xml')):
                route_strings.extend(text_tokens(payload))
            if '/embeddings/' in lower and lower.endswith(('.xlsx','.xlsm')):
                embeddings.append(xlsx_profile(payload,info.filename))
            if '/charts/' in lower and lower.endswith('.xml'):
                try: charts.append(chart_profile(payload,info.filename))
                except ET.ParseError: pass
    return {'filename':filename,'memberCount':len(members),'embeddedWorkbookCount':len(embeddings),'chartXmlCount':len(charts),'packageRouteMarkers':route_markers(route_strings),'embeddedWorkbooks':embeddings,'charts':charts,'rawDocumentTextEmitted':False,'rawImagesEmitted':False,'rawNumericValuesEmitted':False}

def main():
    a=argparse.ArgumentParser(); a.add_argument('--output',type=Path,required=True); a.add_argument('--retrieved-date',required=True); args=a.parse_args(); c=json.loads(CONTRACT.read_text())
    profiles=[]
    for f in c['stage1Evidence']['files']:
        url=f"{PUBLIC_BASE}/{f['id']}/file_downloaded"
        data,final,ctype=get(url); digest=hashlib.sha256(data).hexdigest()
        if digest!=f['sha256']: raise RuntimeError(f"publisher file hash drift: {f['filename']}")
        p=package_profile(data,f['filename']); p.update({'publisherFileId':f['id'],'retrievedSizeBytes':len(data),'sha256':digest,'resolvedUrl':final,'contentType':ctype}); profiles.append(p)
    embedded=sum(p['embeddedWorkbookCount'] for p in profiles); charts=sum(p['chartXmlCount'] for p in profiles); cached=sum(ch['cachedPointElements'] for p in profiles for ch in p['charts'])
    inj=sum(p['packageRouteMarkers']['injection'] for p in profiles)+sum(ch['routeMarkers']['injection'] for p in profiles for ch in p['charts'])+sum(s['routeMarkers']['injection'] for p in profiles for w in p['embeddedWorkbooks'] for s in w['sheets'])
    sls=sum(p['packageRouteMarkers']['sls'] for p in profiles)+sum(ch['routeMarkers']['sls'] for p in profiles for ch in p['charts'])+sum(s['routeMarkers']['sls'] for p in profiles for w in p['embeddedWorkbooks'] for s in w['sheets'])
    status='ooxml-profile-needs-semantic-review' if (embedded or charts or cached) else 'ooxml-no-machine-readable-measurement-payload'
    result={'schema':1,'status':status,'retrievedDate':args.retrieved_date,'source':{'datasetId':c['datasetId'],'datasetDoi':c['source']['datasetDoi'],'license':c['source']['license']},'profile':{'packages':profiles,'embeddedWorkbookCount':embedded,'chartXmlCount':charts,'chartCachedPointElements':cached,'routeMarkerTotals':{'injection':inj,'sls':sls},'rawNumericValuesEmitted':False,'rawImagesEmitted':False,'rawThirdPartyFilesCommittedOrUploaded':False},'acceptance':{'countsAsFullyProfiledMeasuredDataset':False,'acceptedInjectionProcessTimeSeriesSamples':0,'semanticReviewRequired':status=='ooxml-profile-needs-semantic-review'},'evidenceBoundary':c['evidenceBoundary']}
    args.output.parent.mkdir(parents=True,exist_ok=True); args.output.write_text(json.dumps(result,indent=2)+'\n')
    print(json.dumps({'status':status,'embeddedWorkbookCount':embedded,'chartXmlCount':charts,'chartCachedPointElements':cached,'routeMarkerTotals':result['profile']['routeMarkerTotals']},indent=2))
if __name__=='__main__': main()
