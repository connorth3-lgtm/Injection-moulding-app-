#!/usr/bin/env python3
from __future__ import annotations
import argparse,hashlib,json,tempfile,urllib.request
from pathlib import Path
from openpyxl import load_workbook
ROOT=Path(__file__).resolve().parents[1];M=json.loads((ROOT/'data/public-benchmark-results/ypf95p4bs4-stage1.json').read_text());DATASET='ypf95p4bs4';UA='MouldMaster-Educational-Evidence-Profiler/1.0'
MAIN=next(x for x in M['manifest']['files'] if x['filename']=='TF-IN92-241-G198-Excel.xlsx')
def get():
    u=f'https://data.mendeley.com/public-files/datasets/{DATASET}/files/{MAIN["id"]}/file_downloaded';req=urllib.request.Request(u,headers={'User-Agent':UA,'Accept':'*/*','Referer':f'https://data.mendeley.com/datasets/{DATASET}/1'})
    with urllib.request.urlopen(req,timeout=240) as r:return r.read(),r.geturl()
def norm(v):return ' '.join(str(v or '').strip().lower().split())
def find_header(ws, required):
    req={norm(x) for x in required}
    for row in ws.iter_rows():
        vals=[norm(c.value) for c in row]
        if req.issubset(set(vals)):
            return row[0].row,{norm(c.value):c.column for c in row if c.value is not None}
    raise RuntimeError(f'header not found in {ws.title}: {required}')
def is_literal_num(c):return isinstance(c.value,(int,float)) and not isinstance(c.value,bool) and c.data_type!='f'
def count_limpieza(ws):
    hr,h=find_header(ws,['Fecha','Máquina','Motivo','Tiempo (m)']);cm=h['máquina'];ct=h['tiempo (m)'];count=0;records=0
    for r in range(hr+1,ws.max_row+1):
        if norm(ws.cell(r,cm).value)=='inyectora' and is_literal_num(ws.cell(r,ct)):
            count+=1;records+=1
    return {'sheet':ws.title,'directMeasuredField':'Tiempo (m)','selection':'Máquina = Inyectora','directObservedValues':count,'sourceRecordsWithAcceptedValue':records}
def count_setup_history(ws):
    hr,h=find_header(ws,['Fecha','Operación','Tiempo(h)']);co=h['operación'];ct=h['tiempo(h)'];count=0;records=0
    for r in range(hr+1,ws.max_row+1):
        op=norm(ws.cell(r,co).value)
        if 'cambio de molde' in op and is_literal_num(ws.cell(r,ct)):
            count+=1;records+=1
    return {'sheet':ws.title,'directMeasuredField':'Tiempo(h)','selection':'Operación contains Cambio de molde','directObservedValues':count,'sourceRecordsWithAcceptedValue':records}
def count_paradas(ws):
    hr,h=find_header(ws,['Fecha','Máquina','Motivo','Tiempo (m)']);cm=h['máquina'];ct=h['tiempo (m)'];count=0;records=0
    for r in range(hr+1,ws.max_row+1):
        if norm(ws.cell(r,cm).value)=='inyectora' and is_literal_num(ws.cell(r,ct)):
            count+=1;records+=1
    return {'sheet':ws.title,'directMeasuredField':'Tiempo (m)','selection':'Máquina = Inyectora','directObservedValues':count,'sourceRecordsWithAcceptedValue':records}
def count_westinghouse(ws):
    # The workbook labels a merged block as Observaciones (min); count only literal numeric cells beneath that merged observation block for activity rows.
    target=None
    for mr in ws.merged_cells.ranges:
        if norm(ws.cell(mr.min_row,mr.min_col).value)=='observaciones (min)':target=mr;break
    if target is None:raise RuntimeError('Westinghouse observation merged range not found')
    header_row=target.max_row+1;activity_col=None
    for c in range(1,ws.max_column+1):
        if norm(ws.cell(header_row,c).value)=='actividad':activity_col=c;break
    if activity_col is None:raise RuntimeError('Westinghouse activity column not found')
    count=records=0
    for r in range(header_row+1,ws.max_row+1):
        activity=ws.cell(r,activity_col).value
        if not isinstance(activity,str) or not activity.strip():continue
        row_count=sum(1 for c in range(target.min_col,target.max_col+1) if is_literal_num(ws.cell(r,c)))
        if row_count:count+=row_count;records+=1
    return {'sheet':ws.title,'directMeasuredField':'Observaciones (min)','selection':f'merged observation columns {target.min_col}-{target.max_col}; activity rows only','directObservedValues':count,'sourceActivityRowsWithAcceptedValues':records,'observationColumns':target.max_col-target.min_col+1}
def main():
    ap=argparse.ArgumentParser();ap.add_argument('--output',type=Path,required=True);ap.add_argument('--retrieved-date',required=True);a=ap.parse_args();data,final=get();digest=hashlib.sha256(data).hexdigest()
    if digest!=MAIN['sha256']:raise RuntimeError('publisher SHA mismatch for operational workbook')
    with tempfile.TemporaryDirectory() as td:
        p=Path(td)/'source.xlsx';p.write_bytes(data);wb=load_workbook(p,data_only=False,read_only=False)
        blocks=[count_limpieza(wb['Limpieza 2023 (Mtto)']),count_setup_history(wb['Setup 2023']),count_paradas(wb['Paradas de Maquinaria 2023']),count_westinghouse(wb['Tiempos de Setup (Westinghouse)'])];wb.close()
    total=sum(x['directObservedValues'] for x in blocks)
    result={'schema':1,'status':'accepted-profiled-record-level-injection-operations','retrievedDate':a.retrieved_date,'source':{'datasetId':'mendeley-ypf95p4bs4-v1','datasetDoi':'10.17632/ypf95p4bs4.1','license':'CC BY 4.0','publisherFileName':MAIN['filename'],'sha256':digest,'publisherSha256Matched':True,'resolvedUrl':final},'profile':{'acceptedDirectObservedBlocks':blocks,'directRecordLevelInjectionOperationalMeasurements':total,'excludedSheetsAndArtifacts':{'arenaDoeFiles':3,'validationWorkbook':'simulation/model validation and scenario outputs','salesSheet':'commercial records, not injection-process measurements','summaryPivotFormulaSheets':'derived summaries/formulas','proposalSmedTpmSsdSheets':'proposed/intervention data, not baseline direct measurements','financeAndScheduleSheets':'economic/project calculations'},'rawNumericValuesEmitted':False},'acceptance':{'countsAsFullyProfiledMeasuredDataset':total>0,'acceptedRecordLevelMeasuredValues':total,'acceptedMeasuredTimeSeriesSamples':0,'evidenceClass':'record-level injection-operation duration measurements'},'retrieval':{'validationWorkbookDownloaded':False,'doeFilesDownloaded':False,'rawPublisherFileCommitted':False,'rawNumericValuesUploadedAsArtifact':False},'evidenceBoundary':'Acceptance is restricted to direct literal duration observations in four source-log/time-study blocks: injector cleaning/maintenance, historical mould-change setup, injector downtime incidents, and Westinghouse setup observations. Dates, identifiers, sales, formulas, pivots, availability/MTBF/MTTR calculations, proposal/intervention values, finance, Arena models and validation scenario outputs are excluded.'}
    a.output.parent.mkdir(parents=True,exist_ok=True);a.output.write_text(json.dumps(result,indent=2,ensure_ascii=False)+'\n');print(json.dumps({'status':result['status'],'blocks':blocks,'total':total},indent=2))
if __name__=='__main__':main()
