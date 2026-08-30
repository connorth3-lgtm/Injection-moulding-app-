#!/usr/bin/env python3
from __future__ import annotations

import argparse
import hashlib
import json
import tempfile
import urllib.request
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

DATASET_ID = "6k8fpbrd9s"
VERSION = 1
DOI = "10.17632/6k8fpbrd9s.1"
EXPECTED_FILE_ID = "8598d42d-f794-47e2-ad84-dd952c900d27"
EXPECTED_FILE = "Data.xlsx"
EXPECTED_SHA256 = "14c056dd86e11cc47e1e97834174631f9dc0806442917a7149ddf5856dc9b11c"
PUBLIC_FILES_ENDPOINT = f"https://data.mendeley.com/public-api/datasets/{DATASET_ID}/files?folder_id=root&version={VERSION}"
API_ROOT = "https://api.data.mendeley.com"
UA = "MouldMaster-Educational-Evidence-Profiler/1.0"


def get(url: str, accept: str = "*/*"):
    req = urllib.request.Request(url, headers={"Accept": accept, "User-Agent": UA})
    with urllib.request.urlopen(req, timeout=120) as r:
        return r.read(), r.geturl()


def flatten_files(payload):
    if isinstance(payload, list):
        return payload
    if isinstance(payload, dict):
        for key in ("results", "files", "items", "data"):
            if isinstance(payload.get(key), list):
                return payload[key]
    return []


def file_id(item):
    return str(item.get("id") or item.get("file_id") or item.get("uuid") or "").strip()


def file_name(item):
    return str(item.get("filename") or item.get("name") or "").strip()


def file_url(item):
    details = item.get("content_details") or item.get("contentDetails") or {}
    for candidate in (details.get("download_url"), details.get("downloadUrl"), item.get("download_url"), item.get("downloadUrl")):
        if candidate:
            return str(candidate)
    return f"{API_ROOT}/datasets/{DATASET_ID}/files/{file_id(item)}/file_downloaded?version={VERSION}"


ODD_72 = [get_column_letter(i) for i in range(1, 73, 2)]
EVEN_72 = [get_column_letter(i) for i in range(2, 73, 2)]
ROLE_COLUMNS = {
    "Figure2": {
        "temperature": ["A","C","E","G","I","K","N","P","R","T","V","X"],
        "specificVolume": ["B","D","F","H","J","L","O","Q","S","U","W","Y"],
    },
    "Figure3": {
        "time": ["A","C","E","G","I","K"],
        "temperature": ["B","D","F","H","J","L"],
    },
    "Figure4": {
        "temperature": ["A","C","E","H","J","L","O","Q","S"],
        "specificVolume": ["B","D","F","I","K","M","P","R","T"],
    },
    "Figure5": {"time": ["A","C"], "pressure": ["B","D"]},
    "Figure6": {
        "temperature": ["A","C","E","H","J","L","P","R","T","W","Y","AA"],
        "specificVolume": ["B","D","F","I","K","M","Q","S","U","X","Z","AB"],
    },
    "Figure7": {"time": ["A"], "pressure": ["B"], "specificVolume": ["C"]},
    "Figure8": {
        "temperature": ["A","H","O","V"],
        "specificVolume": ["B","C","D","E","F","I","J","K","L","M","P","Q","R","S","T","W","X","Y","Z","AA"],
    },
    "Figure9": {"pistonSpeed": ["A"], "specificVolume": [get_column_letter(i) for i in range(2,13)]},
    "Figure10abc": {"pressure": ODD_72, "specificVolume": EVEN_72},
    "Figure10def": {"pressure": ODD_72, "specificVolume": EVEN_72},
}
EXPECTED_SHEETS = list(ROLE_COLUMNS)
EXPECTED_ROLE_TOTALS = {
    "temperature": 2422,
    "pressure": 11782,
    "specificVolume": 14386,
    "time": 3207,
    "pistonSpeed": 20,
}
EXPECTED_TOTAL_NUMERIC = 31817
EXPECTED_DIRECT_PHYSICAL = 28590
EXPECTED_COORDINATES = 3227
EXPECTED_MATERIAL_TRACE_MEASUREMENTS = 6026


def count_numeric_by_column(ws):
    counts = {}
    formulas = 0
    text = {}
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if v is None:
                continue
            if isinstance(v, str):
                if v.startswith("="):
                    formulas += 1
                else:
                    coord = getattr(cell, "coordinate", None)
                    if coord:
                        text[coord] = " ".join(v.split())[:240]
            elif isinstance(v, (int, float)) and not isinstance(v, bool):
                col = getattr(cell, "column_letter", None)
                if col:
                    counts[col] = counts.get(col, 0) + 1
    return counts, formulas, text


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--output", required=True)
    ap.add_argument("--retrieved-date", required=True)
    args = ap.parse_args()

    raw, _ = get(PUBLIC_FILES_ENDPOINT, "application/json")
    files = flatten_files(json.loads(raw.decode("utf-8")))
    matches = [x for x in files if file_id(x) == EXPECTED_FILE_ID and file_name(x) == EXPECTED_FILE]
    if len(matches) != 1:
        raise RuntimeError("exact pvT workbook identity drifted")
    data, final_url = get(file_url(matches[0]))
    digest = hashlib.sha256(data).hexdigest()
    if digest != EXPECTED_SHA256:
        raise RuntimeError(f"publisher SHA-256 drifted: {digest}")

    with tempfile.TemporaryDirectory() as td:
        p = Path(td) / EXPECTED_FILE
        p.write_bytes(data)
        wb = load_workbook(p, read_only=True, data_only=False)
        if wb.sheetnames != EXPECTED_SHEETS:
            raise RuntimeError(f"pvT sheet set/order drifted: {wb.sheetnames}")
        sheet_profiles = []
        role_totals = {k: 0 for k in EXPECTED_ROLE_TOTALS}
        total_numeric = 0
        total_formulas = 0
        for ws in wb.worksheets:
            column_counts, formulas, text = count_numeric_by_column(ws)
            roles = {}
            classified = 0
            for role, cols in ROLE_COLUMNS[ws.title].items():
                n = sum(column_counts.get(c, 0) for c in cols)
                roles[role] = {"columns": cols, "numericCells": n}
                role_totals[role] += n
                classified += n
            sheet_numeric = sum(column_counts.values())
            if classified != sheet_numeric:
                raise RuntimeError(f"{ws.title}: role map classified {classified} of {sheet_numeric} numeric cells")
            total_numeric += sheet_numeric
            total_formulas += formulas
            sheet_profiles.append({
                "sheet": ws.title,
                "maxRow": ws.max_row,
                "maxColumn": ws.max_column,
                "roles": roles,
                "numericCells": sheet_numeric,
                "formulaCells": formulas,
                "keyLabels": text,
                "numericValuesEmitted": False,
            })

    if role_totals != EXPECTED_ROLE_TOTALS:
        raise RuntimeError(f"pvT role totals drifted: {role_totals}")
    if total_numeric != EXPECTED_TOTAL_NUMERIC or total_formulas != 0:
        raise RuntimeError(f"pvT aggregate totals drifted: numeric={total_numeric}, formulas={total_formulas}")
    direct_physical = role_totals["temperature"] + role_totals["pressure"] + role_totals["specificVolume"]
    coordinates = role_totals["time"] + role_totals["pistonSpeed"]
    if direct_physical != EXPECTED_DIRECT_PHYSICAL or coordinates != EXPECTED_COORDINATES:
        raise RuntimeError("pvT direct/coordinate totals drifted")
    trace_measurements = (
        next(x for x in sheet_profiles if x["sheet"] == "Figure3")["roles"]["temperature"]["numericCells"]
        + next(x for x in sheet_profiles if x["sheet"] == "Figure5")["roles"]["pressure"]["numericCells"]
        + next(x for x in sheet_profiles if x["sheet"] == "Figure7")["roles"]["pressure"]["numericCells"]
        + next(x for x in sheet_profiles if x["sheet"] == "Figure7")["roles"]["specificVolume"]["numericCells"]
    )
    if trace_measurements != EXPECTED_MATERIAL_TRACE_MEASUREMENTS:
        raise RuntimeError(f"pvT material trace total drifted: {trace_measurements}")

    result = {
        "schema": 1,
        "status": "completed-public-measured-material-characterization-benchmark",
        "retrievedDate": args.retrieved_date,
        "source": {
            "datasetId": "mendeley-6k8fpbrd9s-v1",
            "datasetDoi": DOI,
            "version": VERSION,
            "license": "CC BY 4.0",
            "publisherFileId": EXPECTED_FILE_ID,
            "publisherFileName": EXPECTED_FILE,
            "sha256": digest,
            "publisherSha256Matched": True,
            "retrievedSizeBytes": len(data),
            "resolvedUrl": final_url,
        },
        "profile": {
            "material": "polypropylene",
            "measurementDomain": "pressure-specific volume-temperature (pvT) material characterization",
            "injectionMouldingCycleDataset": False,
            "sheetCount": len(sheet_profiles),
            "sheets": sheet_profiles,
            "deliveredNumericCells": total_numeric,
            "deliveredDirectPhysicalValueCells": direct_physical,
            "directPhysicalRoleTotals": {
                "temperature": role_totals["temperature"],
                "pressure": role_totals["pressure"],
                "specificVolume": role_totals["specificVolume"],
            },
            "coordinateCellsExcludedFromDirectPhysicalCount": coordinates,
            "coordinateRoleTotals": {"time": role_totals["time"], "pistonSpeed": role_totals["pistonSpeed"]},
            "materialCharacterizationTraceMeasurementCells": trace_measurements,
            "formulaCells": total_formulas,
            "crossFigureReuseMayExist": True,
            "uniqueExperimentalMeasurementCount": None,
            "rawRowsOrCellValuesEmitted": False,
            "numericMeasurementValuesEmitted": False,
        },
        "acceptance": {
            "countsAsFullyProfiledMeasuredDataset": True,
            "acceptedMeasuredTimeSeriesSamples": 0,
            "materialCharacterizationTraceMeasurementsNotAddedToInjectionCycleSampleMetric": True,
            "deliveredCellCountsAreNotClaimedAsDeduplicatedExperiments": True,
        },
        "retrieval": {
            "rawPublisherFileCommitted": False,
            "rawRowsOrCellValuesUploadedAsArtifact": False,
        },
        "evidenceBoundary": "The exact CC BY 4.0 workbook is SHA-256 matched and every delivered numeric cell is classified by role across ten figure-data sheets. 28,590 delivered cells are physical temperature, pressure or specific-volume values; 3,227 time/piston-speed coordinates are excluded from that direct-physical count. Cross-figure reuse may occur, so no deduplicated experiment count is claimed. The dataset is accepted as measured polypropylene material-characterization evidence, not as an injection-moulding cycle dataset, and contributes zero to the injection-cycle high-frequency sample metric."
    }
    out = Path(args.output)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(result, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(json.dumps({
        "status": result["status"],
        "sha256": digest,
        "deliveredNumericCells": total_numeric,
        "deliveredDirectPhysicalValueCells": direct_physical,
        "coordinateCellsExcluded": coordinates,
        "materialCharacterizationTraceMeasurementCells": trace_measurements,
        "acceptedMeasuredTimeSeriesSamples": 0,
    }, indent=2))


if __name__ == "__main__":
    main()
