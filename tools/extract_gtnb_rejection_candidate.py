#!/usr/bin/env python3
"""Recover governed GTNB rejection numerator/denominator evidence for MLM-039 authoring.

Numeric values follow the established GTNB benchmark policy: publisher-stored direct cell
values and cached formula results are read from the exact SHA-pinned workbook. Formula-
backed cells are counted explicitly. A reproducible rejection-rate feature recipe is
emitted, but no learner-facing conclusion is promoted by this candidate layer.
"""
from __future__ import annotations
import hashlib, json, math, re, tempfile
from itertools import zip_longest
from pathlib import Path
from openpyxl import load_workbook
from prove_mendeley_open_sources import SOURCES, public_files, resolve_file, download_first

OUT=Path('measured-source-proof/gtnb-rejection-unreviewed-learning-candidate.json')
def sha(v): return 'sha256:'+hashlib.sha256(json.dumps(v,sort_keys=True,separators=(',',':'),ensure_ascii=False).encode()).hexdigest()
def finite(v): return isinstance(v,(int,float)) and not isinstance(v,bool) and math.isfinite(float(v))
def header_text(v): return str(v or '').strip()
def signal(sid,source_channel,semantic,unit,rows,key,formula_count):
    points=[r for r in rows if finite(r[key])]; x=[float(r['_sourceRow']) for r in points]; y=[float(r[key]) for r in points]
    if len(points)<20: raise RuntimeError(f'{sid}: insufficient numeric evidence: {len(points)}')
    rep={'xSemantic':'observation-index','xUnit':'index','xDirection':'increasing','reductionMethod':'contiguous-valid-record-window-no-interpolation','originalPointCount':len(rows),'x':x,'y':y}
    return {'id':sid,'label':sid.replace('-',' '),'sourceChannel':source_channel,'semantic':semantic,'unit':unit,'sourceValueMode':'delivered-direct-or-cached-formula-result','sourceInjectionFormulaCellCount':formula_count,'representation':rep,'representationFingerprint':sha(rep)}
def main():
    source=next(s for s in SOURCES if s['datasetId']=='mendeley-gtnb4j7bfx-v1')
    file_id,name,expected=source['files'][0]; _,meta=public_files(source['shortId'],source['version']); _,_,urls=resolve_file(meta,file_id,name,source['shortId'],source['version'])
    td=tempfile.TemporaryDirectory(); path=Path(td.name)/name
    try:
        download_first(urls,path); digest=hashlib.sha256(path.read_bytes()).hexdigest()
        if digest!=expected: raise RuntimeError(f'GTNB SHA mismatch: {digest}')
        formula_wb=load_workbook(path,read_only=True,data_only=False); value_wb=load_workbook(path,read_only=True,data_only=True)
        if 'nicky' not in formula_wb.sheetnames or 'nicky' not in value_wb.sheetnames: raise RuntimeError('GTNB rejection worksheet missing')
        formula_iter=formula_wb['nicky'].iter_rows(); value_iter=value_wb['nicky'].iter_rows(values_only=True)
        try:
            formula_header=next(formula_iter); value_header=next(value_iter)
        except StopIteration: raise RuntimeError('GTNB rejection worksheet is empty')
        headers=[header_text(v) for v in value_header]; formula_headers=[header_text(c.value) for c in formula_header]
        if headers!=formula_headers: raise RuntimeError('GTNB rejection formula/value header disagreement')
        nonempty=[h for h in headers if h]
        if len(nonempty)!=len(set(nonempty)): raise RuntimeError('GTNB rejection duplicate normalized headers')
        col={h:i for i,h in enumerate(headers) if h}
        required=['Maquina','Producto_rechazados','Produccion_total','Presion_inyeccion_bares','Tiempo_ciclo']
        missing=[h for h in required if h not in col]
        if missing: raise RuntimeError(f'GTNB rejection header drift after whitespace normalization: {missing}')
        keys=[('Producto_rechazados','rejected'),('Produccion_total','production'),('Presion_inyeccion_bares','pressure'),('Tiempo_ciclo','cycle')]
        formula_counts={key:0 for _,key in keys}; rows=[]
        for source_row,pair in enumerate(zip_longest(formula_iter,value_iter,fillvalue=None),start=2):
            formula_cells,values=pair
            if formula_cells is None or values is None: raise RuntimeError('GTNB rejection formula/value row-count disagreement')
            machine=header_text(values[col['Maquina']] if col['Maquina']<len(values) else None).upper()
            if not (re.fullmatch(r'I(?:-|_)?\d+',machine) or machine.startswith('INY')): continue
            row={'_sourceRow':source_row}
            for h,key in keys:
                idx=col[h]; v=values[idx] if idx<len(values) else None; row[key]=float(v) if finite(v) else None
                if idx<len(formula_cells) and formula_cells[idx].data_type=='f': formula_counts[key]+=1
            rows.append(row)
        if len(rows)!=4502: raise RuntimeError(f'GTNB injection row drift: {len(rows)}')
        valid=[r for r in rows if finite(r['rejected']) and finite(r['production']) and r['production']>0]
        if len(valid)<400: raise RuntimeError(f'GTNB insufficient valid rejection records: {len(valid)}')
        start=max(0,(len(valid)-400)//2); selected=valid[start:start+400]
        signals=[
            signal('rejected-products','Rejected_Products','recorded-rejected-product-count','units',selected,'rejected',formula_counts['rejected']),
            signal('production-total','Total_Production','recorded-total-production-count','units',selected,'production',formula_counts['production']),
            signal('injection-pressure','Injection_Pressure','recorded-injection-pressure','bar',selected,'pressure',formula_counts['pressure']),
            signal('cycle-time','Cycle_Time','recorded-cycle-time','s',selected,'cycle',formula_counts['cycle']),
        ]
        recommended_features=[{
            'id':'rejection-rate-percent',
            'label':'Rejected products as a share of total production',
            'method':'ratio_of_sums_percent',
            'methodVersion':1,
            'inputs':['signal:rejected-products','signal:production-total'],
            'params':{},
            'calculationScope':'displayed-reviewed-representation',
            'unit':'%'
        }]
        candidate={'candidateId':'GTNB-REJECTION-NUMERATOR-DENOMINATOR-01','datasetId':source['datasetId'],'sourceArtifact':'modelo.xlsx','sourceFingerprint':'sha256:'+digest,'sourceScope':{'selection':'bounded 400-record interval from injection records with finite rejection/production values and non-zero production totals','validInjectionRecords':len(valid),'selectedOrdinalStart':start,'selectedOrdinalEndExclusive':start+len(selected),'sourceValueMode':'publisher-workbook-stored-values-consistent-with-existing-benchmark-reader','sourceInjectionFormulaCellCounts':formula_counts},'signals':signals,'candidateFingerprint':sha(signals),'suggestedCatalogueCases':['MLM-039'],'recommendedFeatures':recommended_features,'bindingBlockers':[],'evidenceBoundary':'The source workbook supplies recorded rejected-product and total-production counts as stored values, including publisher-cached formula results. The governed ratio-of-sums feature can quantify their aggregate association in the selected records; it does not establish a process mechanism or root cause.'}
        result={'schemaVersion':1,'status':'unreviewed-source-derived-candidates','promotionEligible':False,'candidateCount':1,'candidates':[candidate],'boundary':'Numeric source evidence, count-channel governance and a versioned feature recipe are available. This remains authoring evidence only: case-specific wording, novelty review and independent engineering review are still required before learner promotion.'}
        OUT.parent.mkdir(parents=True,exist_ok=True); OUT.write_text(json.dumps(result,indent=2,ensure_ascii=False)+'\n',encoding='utf-8')
        print(json.dumps({'status':result['status'],'candidateId':candidate['candidateId'],'validInjectionRecords':len(valid),'formulaCounts':formula_counts,'recommendedFeatureMethods':[f['method'] for f in recommended_features],'bindingBlockers':candidate['bindingBlockers']},separators=(',',':')))
    finally: td.cleanup()
    return 0
if __name__=='__main__': raise SystemExit(main())
