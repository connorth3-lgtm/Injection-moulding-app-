#!/usr/bin/env python3
"""Extract a compact direct 6k8 isobar comparison candidate for MLM-052.

The originally targeted Figure10abc columns contain only three direct paired cells in the
published workbook, so they are not inflated into a learning trace. This extractor uses
the already source-proven Figure2 200/400/800 bar cooling isobars and keeps only direct
source points inside a common temperature band. No interpolation or formula evaluation
is performed.
"""
from __future__ import annotations
import hashlib, json, math, tempfile
from pathlib import Path
from openpyxl import load_workbook
from prove_mendeley_open_sources import SOURCES, public_files, resolve_file, download_first

OUT=Path('measured-source-proof/6k8-pressure-unreviewed-learning-candidate.json')
TEMPERATURE_MIN_C=80.0
TEMPERATURE_MAX_C=200.0

def sha(v): return 'sha256:'+hashlib.sha256(json.dumps(v,sort_keys=True,separators=(',',':'),ensure_ascii=False).encode()).hexdigest()
def finite(v): return isinstance(v,(int,float)) and not isinstance(v,bool) and math.isfinite(float(v))

def maximal_decreasing_branch(x,y):
    if len(x)!=len(y) or len(x)<3: raise RuntimeError('invalid Figure2 source pairs')
    cut=len(x)
    for i,(a,b) in enumerate(zip(x,x[1:])):
        if b>a:
            cut=i+1
            break
    selected_x=x[:cut]; selected_y=y[:cut]; excluded=len(x)-cut
    if not all(a>=b for a,b in zip(selected_x,selected_x[1:])):
        raise RuntimeError('Figure2 branch is not decreasing before terminal reversal')
    if excluded!=1:
        raise RuntimeError(f'Figure2 source-shape drift: expected one terminal reversal pair, got {excluded}')
    return selected_x,selected_y,excluded

def make_signal(sid,source_channel,semantic,x,y,source_pair_count):
    if not all(a>=b for a,b in zip(x,x[1:])): raise RuntimeError(f'{sid}: temperature axis is not decreasing')
    rep={'xSemantic':'temperature','xUnit':'degC','xDirection':'decreasing','reductionMethod':'direct-source-points-common-temperature-band-no-interpolation','originalPointCount':source_pair_count,'x':x,'y':y}
    return {'id':sid,'label':sid.replace('-',' '),'sourceChannel':source_channel,'semantic':semantic,'unit':'mm3/g','representation':rep,'representationFingerprint':sha(rep)}

def main():
    source=next(s for s in SOURCES if s['datasetId']=='mendeley-6k8fpbrd9s-v1')
    file_id,name,expected=source['files'][0]; _,meta=public_files(source['shortId'],source['version']); _,_,urls=resolve_file(meta,file_id,name,source['shortId'],source['version'])
    td=tempfile.TemporaryDirectory(); path=Path(td.name)/name
    try:
        download_first(urls,path); digest=hashlib.sha256(path.read_bytes()).hexdigest()
        if digest!=expected: raise RuntimeError(f'6k8 SHA mismatch: {digest}')
        wb=load_workbook(path,read_only=False,data_only=False); ws=wb['Figure2']
        specs=[
            ('200bar','A','B','Figure2!B','specific-volume-200bar-isobaric-cooling'),
            ('400bar','E','F','Figure2!F','specific-volume-400bar-isobaric-cooling'),
            ('800bar','I','J','Figure2!J','specific-volume-800bar-isobaric-cooling')]
        signals=[]; selection={}
        for sid,xc,yc,source_channel,semantic in specs:
            raw_x=[]; raw_y=[]
            for r in range(1,ws.max_row+1):
                a,b=ws[f'{xc}{r}'].value,ws[f'{yc}{r}'].value
                if finite(a) and finite(b): raw_x.append(float(a)); raw_y.append(float(b))
            branch_x,branch_y,excluded=maximal_decreasing_branch(raw_x,raw_y)
            chosen=[(x,y) for x,y in zip(branch_x,branch_y) if TEMPERATURE_MIN_C <= x <= TEMPERATURE_MAX_C]
            if len(chosen)<15: raise RuntimeError(f'6k8 {sid}: insufficient direct points in common temperature band: {len(chosen)}')
            x=[p[0] for p in chosen]; y=[p[1] for p in chosen]
            signals.append(make_signal(sid,source_channel,semantic,x,y,len(raw_x)))
            selection[sid]={'sourceNumericPairCount':len(raw_x),'terminalReversalPairsExcluded':excluded,'commonBandDirectPairCount':len(chosen),'temperatureBandDegC':[TEMPERATURE_MIN_C,TEMPERATURE_MAX_C]}
        candidate={'candidateId':'MEND-6K8-FIGURE2-PRESSURE-CONTRAST-01','datasetId':source['datasetId'],'sourceArtifact':'Data.xlsx','sourceFingerprint':'sha256:'+digest,'sourceScope':{'sheet':'Figure2','pressureSeriesBar':[200,400,800],'selection':selection,'interpolationPerformed':False},'signals':signals,'candidateFingerprint':sha(signals),'suggestedCatalogueCases':['MLM-052'],'evidenceBoundary':'Direct source points from the 200, 400 and 800 bar isobaric cooling series are compared only within a common 80–200 °C band. This can support a learner comparison of measured pressure-level differences in pvT behaviour; it does not establish a production process window or causal production mechanism.'}
        result={'schemaVersion':1,'status':'unreviewed-source-derived-candidates','promotionEligible':False,'candidateCount':1,'candidates':[candidate],'boundary':'Authoring evidence only; independent engineering review and a case-specific governed binding are required before promotion.'}
        OUT.parent.mkdir(exist_ok=True); OUT.write_text(json.dumps(result,indent=2,ensure_ascii=False)+'\n',encoding='utf-8')
        print(json.dumps({'status':result['status'],'candidateId':candidate['candidateId'],'selection':selection},separators=(',',':')))
    finally: td.cleanup()
    return 0
if __name__=='__main__': raise SystemExit(main())
