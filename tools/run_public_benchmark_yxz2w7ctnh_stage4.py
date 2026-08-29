#!/usr/bin/env python3
from __future__ import annotations
import argparse, hashlib, io, json, urllib.request
from pathlib import Path
from openpyxl import load_workbook

ROOT=Path(__file__).resolve().parents[1]
CONTRACT=ROOT/'data/public-benchmark-contracts/yxz2w7ctnh-v1.json'
DATASET_ID='yxz2w7ctnh'; PUBLIC_FILE_ROOT=f'https://data.mendeley.com/public-files/datasets/{DATASET_ID}/files'; UA='MouldMaster-Educational-Evidence-Profiler/1.0'
EXPECTED_TOTAL=450

def get(url):
    req=urllib.request.Request(url,headers={'User-Agent':UA,'Accept':'*/*'})
    with urllib.request.urlopen(req,timeout=120) as r: return r.read(),r.geturl()

def count_numeric_constants(ws, columns):
    counts={c:0 for c in columns}; formulas={c:0 for c in columns}
    for c in columns:
        for cell in ws[c]:
            v=cell.value
            if v is None: continue
            if cell.data_type=='f' or (isinstance(v,str) and v.startswith('=')):
                formulas[c]+=1; continue
            if isinstance(v,(int,float)) and not isinstance(v,bool): counts[c]+=1
    return counts,formulas

def main():
    ap=argparse.ArgumentParser(); ap.add_argument('--output',required=True); ap.add_argument('--retrieved-date',required=True); a=ap.parse_args()
    c=json.loads(CONTRACT.read_text()); by_name={f['filename']:f for f in c['stage1Evidence']['mechanicalFiles']}; canonical=c['stage2Evidence']['canonicalTestSheets']; regions=c['stage3Evidence']['injectionRegions']
    needed_files={canonical['bending']['file'],canonical['tensile']['file']}; books={}; hashes={}; hashes_ok=True
    for name in sorted(needed_files):
        f=by_name[name]; data,final=get(f"{PUBLIC_FILE_ROOT}/{f['id']}/file_downloaded"); digest=hashlib.sha256(data).hexdigest(); matched=digest.lower()==f['sha256'].lower(); hashes_ok &= matched
        books[name]=load_workbook(io.BytesIO(data),read_only=False,data_only=False); hashes[name]={'fileId':f['id'],'expectedSha256':f['sha256'],'sha256':digest,'publisherSha256Matched':matched,'sizeBytes':len(data),'resolvedUrl':final}
    sheet_profiles=[]; total=0; formula_excluded=0
    specs=[('bending','bending_PLA'),('bending','bending_ABS'),('tensile','tensile_PLA'),('tensile','tensile_ABS')]
    for family,sheet in specs:
        book_name=canonical[family]['file']; ws=books[book_name][sheet]; r=regions[sheet]
        anchor_value=str(ws[r['anchor']].value or '')
        cols=r['directNumericColumns']; counts,forms=count_numeric_constants(ws,cols); subtotal=sum(counts.values()); total+=subtotal; formula_excluded+=sum(forms.values())
        sheet_profiles.append({'testFamily':family,'sheet':sheet,'sourceFile':book_name,'routeAnchorCell':r['anchor'],'routeAnchorExpected':r['anchorText'],'routeAnchorObserved':anchor_value,'routeAnchorMatched':anchor_value==r['anchorText'],'directNumericColumns':cols,'numericConstantCellsByColumn':counts,'formulaCellsExcludedByColumn':forms,'acceptedDirectPhysicalValues':subtotal,'rawNumericValuesEmitted':False})
    anchors_ok=all(x['routeAnchorMatched'] for x in sheet_profiles)
    accepted=bool(hashes_ok and anchors_ok and total==EXPECTED_TOTAL and c['stage3Evidence']['impactExcludedFromAcceptance'] is True)
    result={'schema':1,'status':'completed-profiled-injection-material-test-scalars' if accepted else 'profile-needs-review','retrievedDate':a.retrieved_date,'source':{'datasetId':c['datasetId'],'datasetDoi':c['source']['datasetDoi'],'license':c['source']['license'],'version':c['source']['version']},'retrieval':{'sourceFiles':hashes,'allPublisherHashesMatched':hashes_ok,'impactWorkbookRetrieved':False,'energyWorkbookRetrieved':False,'rawPublisherFilesCommitted':False,'rawRowsOrNumericValuesUploadedAsArtifact':False},'profile':{'injectionRouteAnchorsMatched':anchors_ok,'acceptedSheets':sheet_profiles,'acceptedInjectionMaterialTestDirectPhysicalValues':total if accepted else 0,'formulaCellsInPinnedDirectColumnsExcluded':formula_excluded,'fdmRegionsExcluded':True,'impactSheetsExcludedRouteUnresolved':True,'duplicateEmbeddedSheetsExcluded':True,'energySheetsExcluded':True,'sampleIdentifiersExcluded':True,'tensileS0GeometryExcluded':True,'rawNumericValuesEmitted':False},'acceptance':{'countsAsFullyProfiledMeasuredDataset':accepted,'acceptedInjectionSubsetFullyProfiled':accepted,'acceptedMaterialTestDirectPhysicalValues':total if accepted else 0,'acceptedMaterialTestTraceValues':0,'acceptedInjectionProcessTimeSeriesSamples':0},'evidenceBoundary':c['evidenceBoundary']}
    Path(a.output).parent.mkdir(parents=True,exist_ok=True); Path(a.output).write_text(json.dumps(result,indent=2)+'\n')
    print(json.dumps({'status':result['status'],'hashesMatched':hashes_ok,'anchorsMatched':anchors_ok,'directPhysicalValues':result['acceptance']['acceptedMaterialTestDirectPhysicalValues'],'formulaCellsExcluded':formula_excluded,'sheets':[{'sheet':x['sheet'],'count':x['acceptedDirectPhysicalValues'],'columns':x['numericConstantCellsByColumn']} for x in sheet_profiles]},indent=2))
if __name__=='__main__': main()
