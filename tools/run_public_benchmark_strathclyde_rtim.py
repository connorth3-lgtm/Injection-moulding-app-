#!/usr/bin/env python3
from __future__ import annotations
import argparse, hashlib, html, http.cookiejar, io, json, re, tempfile, urllib.error, urllib.parse, urllib.request, zipfile
from pathlib import Path
from openpyxl import load_workbook
ROOT=Path(__file__).resolve().parents[1]
CONTRACT=ROOT/'data/public-benchmark-contracts/strathclyde-rtim-tablets-v1.json'
UA='Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 Chrome/124 Safari/537.36 MouldMaster-Evidence-Profiler/1.0'
JAR=http.cookiejar.CookieJar(); OPENER=urllib.request.build_opener(urllib.request.HTTPCookieProcessor(JAR))
def request(url,accept='*/*',referer=None):
    h={'User-Agent':UA,'Accept':accept,'Accept-Language':'en-GB,en;q=0.9','Cache-Control':'no-cache'}
    if referer:h['Referer']=referer
    req=urllib.request.Request(url,headers=h)
    try:
        with OPENER.open(req,timeout=120) as r:return {'ok':True,'status':r.status,'body':r.read(),'finalUrl':r.geturl(),'headers':dict(r.headers)}
    except urllib.error.HTTPError as e:
        return {'ok':False,'status':e.code,'body':b'','finalUrl':e.geturl(),'headers':dict(e.headers or {})}
def sanit(v):
    s=' '.join(str(v or '').replace('\x00',' ').split())
    if not s:return None
    s=re.sub(r'(?<![A-Za-z])[-+]?\d+(?:[.,]\d+)?(?:[Ee][-+]?\d+)?(?![A-Za-z])','<n>',s)
    return s[:220]
def discover_xlsx(page,expected):
    r=request(page,'text/html,application/xhtml+xml')
    if not r['ok']:raise RuntimeError(f'publisher page HTTP {r["status"]}')
    text=html.unescape(r['body'].decode('utf-8','replace')).replace('\\/','/')
    hrefs=re.findall(r'href=["\']([^"\']+)["\']',text,re.I); out=[]
    for href in hrefs:
        u=urllib.parse.urljoin(r['finalUrl'],href); label=urllib.parse.unquote(u)
        if expected.lower() in label.lower() or label.lower().endswith('.xlsx'):
            if u not in out:out.append(u)
    if not out:raise RuntimeError('publisher page exposed no XLSX download link')
    preferred=[u for u in out if expected.lower() in urllib.parse.unquote(u).lower()]
    return (preferred or out)[0],out,r['finalUrl']
def variants(url):
    out=[url]
    for q in ('download=1','download=true'):
        sep='&' if '?' in url else '?'; out.append(url+sep+q)
    return out
def profile_workbook(data):
    with tempfile.TemporaryDirectory() as td:
        p=Path(td)/'source.xlsx';p.write_bytes(data); wb=load_workbook(p,data_only=False,read_only=False); sheets=[]
        for ws in wb.worksheets:
            rows=[];labels=[];numeric=formulas=nonempty=0
            for row in ws.iter_rows():
                rc=fc=nc=0;rlabels=[]
                for cell in row:
                    v=cell.value
                    if v is None:continue
                    nonempty+=1;nc+=1
                    if cell.data_type=='f' or (isinstance(v,str) and v.startswith('=')):formulas+=1;fc+=1
                    elif isinstance(v,(int,float)) and not isinstance(v,bool):numeric+=1;rc+=1
                    elif isinstance(v,str):
                        s=sanit(v)
                        if s:rlabels.append(s);labels.append(s)
                if nc:rows.append({'row':row[0].row,'nonEmptyCells':nc,'numericLiteralCells':rc,'formulaCells':fc,'safeTextLabels':rlabels[:30]})
            uniq=[]
            for s in labels:
                if s not in uniq:uniq.append(s)
            sheets.append({'sheet':ws.title,'maxRow':ws.max_row,'maxColumn':ws.max_column,'nonEmptyCells':nonempty,'numericLiteralCells':numeric,'formulaCells':formulas,'safeTextLabels':uniq[:120],'rows':rows,'rawNumericValuesEmitted':False})
        wb.close();return sheets
def main():
    ap=argparse.ArgumentParser();ap.add_argument('--output',type=Path,required=True);ap.add_argument('--retrieved-date',required=True);a=ap.parse_args();c=json.loads(CONTRACT.read_text());page=c['source']['datasetPage'];expected=c['source']['expectedPublisherFile']
    url,candidates,page_final=discover_xlsx(page,expected); attempts=[]; chosen=None
    for u in variants(url):
        r=request(u,referer=page_final); attempts.append({'url':u,'httpStatus':r['status'],'finalUrl':r['finalUrl'],'contentType':r['headers'].get('Content-Type'),'contentDisposition':r['headers'].get('Content-Disposition'),'zipStructureValid':bool(r['ok'] and zipfile.is_zipfile(io.BytesIO(r['body']))),'responseBodyEmitted':False})
        if r['ok'] and zipfile.is_zipfile(io.BytesIO(r['body'])):chosen=r;break
    base={'schema':1,'retrievedDate':a.retrieved_date,'source':{'datasetId':c['datasetId'],'datasetDoi':c['source']['datasetDoi'],'license':c['source']['license'],'publisherFileName':expected,'discoveredDownloadUrl':url,'candidateDownloadLinksFound':len(candidates)},'retrieval':{'attempts':attempts,'rawPublisherFileCommitted':False,'rawNumericValuesUploadedAsArtifact':False,'rawResponseBodiesEmitted':False},'evidenceBoundary':c['evidenceBoundary']}
    if chosen is None:
        result={**base,'status':'retrieval-blocked-http','profile':None,'acceptance':{'countsAsFullyProfiledMeasuredDataset':False,'acceptedMeasuredTimeSeriesSamples':0,'semanticReviewRequired':True}}
    else:
        data=chosen['body'];digest=hashlib.sha256(data).hexdigest();sheets=profile_workbook(data)
        result={**base,'status':'retrieved-profile-needs-semantic-review','source':{**base['source'],'resolvedUrl':chosen['finalUrl'],'retrievedSizeBytes':len(data),'sha256':digest,'contentType':chosen['headers'].get('Content-Type')},'profile':{'sheetCount':len(sheets),'sheets':sheets,'totalNumericLiteralCells':sum(x['numericLiteralCells'] for x in sheets),'totalFormulaCells':sum(x['formulaCells'] for x in sheets),'rawRowsOrNumericValuesEmitted':False},'acceptance':{'countsAsFullyProfiledMeasuredDataset':False,'acceptedMeasuredTimeSeriesSamples':0,'semanticReviewRequired':True}}
    a.output.parent.mkdir(parents=True,exist_ok=True);a.output.write_text(json.dumps(result,indent=2,ensure_ascii=False)+'\n');print(json.dumps({'status':result['status'],'attempts':attempts,'profile':None if result['profile'] is None else {'sheetCount':result['profile']['sheetCount'],'numericLiteralCells':result['profile']['totalNumericLiteralCells'],'formulaCells':result['profile']['totalFormulaCells']}},indent=2))
if __name__=='__main__':main()
