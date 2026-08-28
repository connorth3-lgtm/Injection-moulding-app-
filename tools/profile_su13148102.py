#!/usr/bin/env python3
from __future__ import annotations
import argparse, csv, hashlib, io, json, re, time, urllib.request, zipfile
from pathlib import Path

DOI = "10.3390/su13148102"
LANDING = "https://www.mdpi.com/article/10.3390/su13148102/s1"
DIRECT_CANDIDATES = [
    "https://mdpi-res.com/d_attachment/sustainability/sustainability-13-08102/article_deploy/sustainability-13-08102-s001.zip",
    LANDING,
]
UA = "Mozilla/5.0 (compatible; MouldMasterEvidenceProfiler/1.0; +https://github.com/connorth3-lgtm/Injection-moulding-app-)"


def sha256(b: bytes) -> str:
    return hashlib.sha256(b).hexdigest()


def fetch(url: str, attempts: int = 4) -> tuple[bytes, str, str]:
    last = None
    for n in range(attempts):
        try:
            req = urllib.request.Request(url, headers={"User-Agent": UA, "Accept": "*/*"})
            with urllib.request.urlopen(req, timeout=60) as r:
                return r.read(), r.headers.get_content_type(), r.geturl()
        except Exception as e:
            last = e
            time.sleep(1.5 * (n + 1))
    raise RuntimeError(f"download failed for {url}: {last}")


def discover() -> tuple[bytes, str, str]:
    errors = []
    for url in DIRECT_CANDIDATES:
        try:
            body, ctype, final = fetch(url)
            if body[:4] == b"PK\x03\x04":
                return body, "application/zip", final
            text = body.decode("utf-8", "replace")
            links = re.findall(r'https?://[^"\'<> ]+\.zip(?:\?[^"\'<> ]*)?', text, flags=re.I)
            links += [
                urllib.request.urljoin(final, x)
                for x in re.findall(r'href=["\']([^"\']+\.zip(?:\?[^"\']*)?)["\']', text, flags=re.I)
            ]
            for link in dict.fromkeys(links):
                try:
                    z, ztype, zfinal = fetch(link)
                    if z[:4] == b"PK\x03\x04":
                        return z, ztype, zfinal
                except Exception as e:
                    errors.append(f"{link}: {e}")
        except Exception as e:
            errors.append(f"{url}: {e}")
    raise RuntimeError("; ".join(errors))


def profile_csv(raw: bytes, name: str) -> dict:
    text = raw.decode("utf-8-sig", "replace")
    sample = text[:65536]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except Exception:
        dialect = csv.excel
    rows = list(csv.reader(io.StringIO(text), dialect))
    if not rows:
        return {"name": name, "format": "csv", "dataRows": 0, "columns": 0}
    header = [str(x).strip() for x in rows[0]]
    data = [r for r in rows[1:] if any(str(x).strip() for x in r)]
    return {"name": name, "format": "csv", "dataRows": len(data), "columns": len(header), "headers": header, "sizeBytes": len(raw), "sha256": sha256(raw)}


def profile_xlsx(raw: bytes, name: str) -> dict:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    sheets = []
    for ws in wb.worksheets:
        rows = ws.iter_rows(values_only=True)
        try:
            first = next(rows)
        except StopIteration:
            sheets.append({"name": ws.title, "dataRows": 0, "columns": 0, "headers": []})
            continue
        headers = ["" if x is None else str(x).strip() for x in first]
        n = 0
        missing = [0] * len(headers)
        numeric = [0] * len(headers)
        for row in rows:
            vals = list(row[:len(headers)]) + [None] * max(0, len(headers) - len(row))
            if not any(v not in (None, "") for v in vals):
                continue
            n += 1
            for i, v in enumerate(vals):
                if v in (None, ""):
                    missing[i] += 1
                elif isinstance(v, (int, float)) and not isinstance(v, bool):
                    numeric[i] += 1
        sheets.append({"name": ws.title, "dataRows": n, "columns": len(headers), "headers": headers, "missingCounts": missing, "numericCounts": numeric})
    return {"name": name, "format": "xlsx", "sizeBytes": len(raw), "sha256": sha256(raw), "sheets": sheets}


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--output", default="su13148102-v1.json")
    args = ap.parse_args()
    archive, ctype, final_url = discover()
    members = []
    with zipfile.ZipFile(io.BytesIO(archive)) as z:
        for info in z.infolist():
            if info.is_dir():
                continue
            raw = z.read(info)
            ext = Path(info.filename).suffix.lower()
            if ext == ".xlsx":
                rec = profile_xlsx(raw, info.filename)
            elif ext in {".csv", ".tsv", ".txt"}:
                rec = profile_csv(raw, info.filename)
            else:
                rec = {"name": info.filename, "format": ext.lstrip(".") or "unknown", "sizeBytes": len(raw), "sha256": sha256(raw)}
            members.append(rec)
    candidate_tables = []
    for m in members:
        if m.get("format") == "xlsx":
            for s in m.get("sheets", []):
                candidate_tables.append({"member": m["name"], **s})
        elif m.get("format") == "csv":
            candidate_tables.append({"member": m["name"], "name": None, "dataRows": m.get("dataRows"), "columns": m.get("columns"), "headers": m.get("headers")})
    best = max(candidate_tables, key=lambda x: (int(x.get("dataRows") or 0) * int(x.get("columns") or 0)), default={})
    payload = {
        "schema_version": 1,
        "status": "profile-generated-review-required",
        "completed_date": "2026-08-28",
        "source": {
            "title": "Multivariate Modeling of Mechanical Properties for Hot Runner Molded Bioplastics and a Recycled Polypropylene Blend — supplementary experimental data",
            "doi": DOI,
            "supplementLanding": LANDING,
            "downloadUrlResolved": final_url,
            "articleLicense": "CC BY 4.0",
            "process": "two-cavity hot-runner injection moulding",
            "materials": 5,
            "publishedObservationRows": 955,
            "publishedColumns": 42
        },
        "archive": {"sizeBytes": len(archive), "sha256": sha256(archive), "memberCount": len(members), "contentType": ctype},
        "members": members,
        "largestTable": best,
        "rawSourceRowsCommitted": False,
        "acceptanceChecks": {
            "observed955Rows": best.get("dataRows") == 955,
            "observed42Columns": best.get("columns") == 42,
            "hasFiveMaterialContext": True,
            "licenseContextRecorded": True
        },
        "boundary": "Profile contains hashes, schema and aggregate missingness/type counts only. It does not emit supplementary raw observations or convert study correlations into universal processing limits."
    }
    Path(args.output).write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(json.dumps({"resolved": final_url, "archiveBytes": len(archive), "members": len(members), "largestTable": best, "acceptanceChecks": payload["acceptanceChecks"]}, indent=2))


if __name__ == "__main__":
    main()
