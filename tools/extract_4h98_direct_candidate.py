#!/usr/bin/env python3
"""Extract direct 4h98 replicate vectors without using publisher average columns."""
from __future__ import annotations
import hashlib, json, math, tempfile
from pathlib import Path
from openpyxl import load_workbook
from prove_mendeley_open_sources import SOURCES, public_files, resolve_file, download_first

OUT=Path('measured-source-proof/4h98-direct-unreviewed-learning-candidate.json')

def sha(v): return 'sha256:'+hashlib.sha256(json.dumps(v,sort_keys=True,separators=(',',':'),ensure_ascii=False).encode()).hexdigest()
def finite(v): return isinstance(v,(int,float)) and not isinstance(v,bool) and math.isfinite(float(v))
def make_signal(sid,source_channel,semantic,unit,values):
    x=[float(i+1) for i in range(len(values))]
    rep={'xSemantic':'observation-index','xUnit':'index','xDirection':'increasing','reductionMethod':'direct-replicate-flattening-by-experiment-then-replicate-column','originalPointCount':len(values),'x':x,'y':values}
    return {'id':sid,'label':sid.replace('-',' '),'sourceChannel':source_channel,'semantic':semantic,'unit':unit,'representation':rep,'representationFingerprint':sha(rep)}
def main():
    source=next(s for s in SOURCES if s['datasetId']=='mendeley-4h98rz9f92-v3')
    file_id,name,expected=source['files'][0]; _,meta=public_files(source['shortId'],source['version']); _,_,urls=resolve_file(meta,file_id,name,source['shortId'],source['version'])
    td=tempfile.TemporaryDirectory(); path=Path(td.name)/name
    try:
        download_first(urls,path); digest=hashlib.sha256(path.read_bytes()).hexdigest()
        if digest!=expected: raise RuntimeError(f'4h98 SHA mismatch: {digest}')
        wb=load_workbook(path,read_only=False,data_only=False); ws=wb['Sheet1']
        specs=[('tensile-replicates',list('EFGHI'),'Sheet1!E:I','tensile-modulus-replicates','GPa'),('hardness-replicates',list('KLMNO'),'Sheet1!K:O','hardness-replicates','HV'),('toughness-replicates',list('QRSTU'),'Sheet1!Q:U','toughness-replicates','J')]
        signals=[]
        for sid,cols,source_channel,semantic,unit in specs:
            values=[]
            for row in range(4,39):
                for col in cols:
                    cell=ws[f'{col}{row}']
                    if cell.data_type=='f': raise RuntimeError(f'unexpected formula in direct replicate block {col}{row}')
                    if not finite(cell.value): raise RuntimeError(f'non-numeric direct replicate {col}{row}: {cell.value!r}')
                    values.append(float(cell.value))
            if len(values)!=175: raise RuntimeError(f'{sid}: expected 175 values, got {len(values)}')
            signals.append(make_signal(sid,source_channel,semantic,unit,values))
        candidate={'candidateId':'MEND-4H98-DIRECT-REPLICATES-01','datasetId':source['datasetId'],'sourceArtifact':'Raw Data.xlsx','sourceFingerprint':'sha256:'+digest,'sourceScope':{'sheet':'Sheet1','experimentRows':'4:38','directReplicateColumns':['E:I','K:O','Q:U'],'excludedDerivedAverageColumns':['J','P','V'],'flatteningOrder':'experiment-row then left-to-right direct replicate column'},'signals':signals,'candidateFingerprint':sha(signals),'suggestedCatalogueCases':['MLM-030','MLM-049','MLM-056'],'evidenceBoundary':'Exactly 525 direct replicate outcome values are represented. Publisher average columns J/P/V are excluded. Replicate variability and measured property differences can be taught; production root cause is not established.'}
        result={'schemaVersion':1,'status':'unreviewed-source-derived-candidates','promotionEligible':False,'candidateCount':1,'candidates':[candidate],'boundary':'Authoring evidence only; independent engineering review and a case-specific governed binding are required before learner promotion.'}
        OUT.parent.mkdir(exist_ok=True); OUT.write_text(json.dumps(result,indent=2,ensure_ascii=False)+'\n',encoding='utf-8')
        print(json.dumps({'status':result['status'],'candidateId':candidate['candidateId']},separators=(',',':')))
    finally: td.cleanup()
    return 0
if __name__=='__main__': raise SystemExit(main())
