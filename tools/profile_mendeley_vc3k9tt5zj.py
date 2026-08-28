#!/usr/bin/env python3
from __future__ import annotations
import argparse, csv, hashlib, html, io, json, re, urllib.parse, urllib.request
from pathlib import Path

DATASET_ID = "vc3k9tt5zj"
VERSION = "2"
DOI = "10.17632/vc3k9tt5zj.2"
TITLE = "Preform injection molding analysis -Database"
LICENSE = "CC BY 4.0"
PAGE = f"https://data.mendeley.com/datasets/{DATASET_ID}/{VERSION}"
PUBLIC_FILES_API = f"https://data.mendeley.com/public-api/datasets/{DATASET_ID}/files?folder_id=root&version={VERSION}"
UA = "MouldMaster-Academy-PET-profiler/1.0"
SUPPORTED = {".csv", ".tsv", ".txt", ".xlsx", ".xlsm"}


def sha256(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def get(url: str, accept: str = "*/*") -> tuple[bytes, str, str]:
    req = urllib.request.Request(url, headers={"User-Agent": UA, "Accept": accept})
    with urllib.request.urlopen(req, timeout=90) as r:
        return r.read(), r.headers.get_content_type(), r.geturl()


def flatten(payload):
    if isinstance(payload, list):
        return payload
    if isinstance(payload, dict):
        for key in ("results", "files", "items", "data"):
            if isinstance(payload.get(key), list):
                return payload[key]
    return []


def discover_files() -> list[dict]:
    out = []
    try:
        raw, _, _ = get(PUBLIC_FILES_API, "application/json")
        for item in flatten(json.loads(raw.decode("utf-8"))):
            url = item.get("download_url") or item.get("downloadUrl") or item.get("content_details", {}).get("download_url")
            name = item.get("name") or item.get("filename") or item.get("file_name")
            if url:
                out.append({"name": name, "url": url, "publisherId": item.get("id")})
    except Exception:
        pass
    if out:
        return out
    raw, _, final = get(PAGE, "text/html,application/xhtml+xml")
    text = html.unescape(raw.decode("utf-8", "replace")).replace("\\u002F", "/").replace("\\/", "/")
    pattern = re.compile(
        rf"https://data\.mendeley\.com/public-files/datasets/{DATASET_ID}/files/([0-9a-fA-F-]{{36}})/file_downloaded"
    )
    seen = set()
    for m in pattern.finditer(text):
        url = m.group(0)
        if url in seen:
            continue
        seen.add(url)
        out.append({"name": None, "url": url, "publisherId": m.group(1)})
    if not out:
        raise RuntimeError("No version-pinned public Mendeley file links were discovered")
    return out


def sniff_filename(url: str, content_disposition: str | None, fallback: str) -> str:
    if content_disposition:
        m = re.search(r"filename\*?=(?:UTF-8''|\")?([^\";]+)", content_disposition, flags=re.I)
        if m:
            return urllib.parse.unquote(m.group(1).strip().strip('"'))
    path = urllib.parse.urlparse(url).path
    name = Path(path).name
    if Path(name).suffix:
        return name
    return fallback


def download_file(rec: dict, index: int) -> tuple[bytes, str, str]:
    req = urllib.request.Request(rec["url"], headers={"User-Agent": UA, "Accept": "*/*"})
    with urllib.request.urlopen(req, timeout=90) as r:
        data = r.read()
        final = r.geturl()
        name = rec.get("name") or sniff_filename(final, r.headers.get("Content-Disposition"), f"publisher-file-{index}")
        return data, name, final


def profile_csv(raw: bytes, name: str) -> dict:
    text = raw.decode("utf-8-sig", "replace")
    try:
        dialect = csv.Sniffer().sniff(text[:65536], delimiters=",;\t|")
    except Exception:
        dialect = csv.excel
    rows = list(csv.reader(io.StringIO(text), dialect))
    if not rows:
        return {"name": name, "format": "csv", "dataRows": 0, "columns": 0, "headers": []}
    headers = [str(v).strip() for v in rows[0]]
    data = [r for r in rows[1:] if any(str(v).strip() for v in r)]
    missing = [0] * len(headers)
    numeric = [0] * len(headers)
    for row in data:
        vals = list(row[:len(headers)]) + [""] * max(0, len(headers) - len(row))
        for i, value in enumerate(vals):
            value = str(value).strip()
            if not value:
                missing[i] += 1
            else:
                try:
                    float(value)
                    numeric[i] += 1
                except ValueError:
                    pass
    return {"name": name, "format": "csv", "dataRows": len(data), "columns": len(headers), "headers": headers, "missingCounts": missing, "numericCounts": numeric}


def profile_xlsx(raw: bytes, name: str) -> dict:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    sheets = []
    for ws in wb.worksheets:
        it = ws.iter_rows(values_only=True)
        try:
            first = next(it)
        except StopIteration:
            sheets.append({"name": ws.title, "dataRows": 0, "columns": 0, "headers": []})
            continue
        headers = ["" if v is None else str(v).strip() for v in first]
        n = 0
        missing = [0] * len(headers)
        numeric = [0] * len(headers)
        for row in it:
            vals = list(row[:len(headers)]) + [None] * max(0, len(headers) - len(row))
            if not any(v not in (None, "") for v in vals):
                continue
            n += 1
            for i, v in enumerate(vals):
                if v in (None, ""):
                    missing[i] += 1
                elif isinstance(v, (int, float)) and not isinstance(v, bool):
                    numeric[i] += 1
        sheets.append({"name": ws.title, "dataRows": n, "columns": len(headers), "headers": headers, "missingCounts": missing, "numericCounts": numeric})
    return {"name": name, "format": "xlsx", "sheets": sheets}


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--output", default="vc3k9tt5zj-v2.json")
    args = ap.parse_args()
    discovered = discover_files()
    files = []
    tables = []
    for i, rec in enumerate(discovered, 1):
        raw, name, final = download_file(rec, i)
        ext = Path(name).suffix.lower()
        base = {"name": name, "urlResolved": final, "publisherId": rec.get("publisherId"), "sizeBytes": len(raw), "sha256": sha256(raw), "suffix": ext}
        if ext in {".csv", ".tsv", ".txt"}:
            p = profile_csv(raw, name)
            base.update(p)
            tables.append({"file": name, **p})
        elif ext in {".xlsx", ".xlsm"}:
            p = profile_xlsx(raw, name)
            base.update(p)
            for sheet in p.get("sheets", []):
                tables.append({"file": name, **sheet})
        else:
            base["format"] = ext.lstrip(".") or "unknown"
        files.append(base)
    largest = max(tables, key=lambda t: int(t.get("dataRows") or 0) * int(t.get("columns") or 0), default={})
    payload = {
        "schema_version": 1,
        "status": "profile-generated-review-required",
        "completed_date": "2026-08-28",
        "source": {"datasetId": DATASET_ID, "version": VERSION, "doi": DOI, "title": TITLE, "license": LICENSE, "publisher": "Mendeley Data", "page": PAGE, "materialContext": "polyethylene terephthalate (PET) water-bottle preforms"},
        "files": files,
        "tableCount": len(tables),
        "largestTable": largest,
        "rawSourceRowsCommitted": False,
        "boundary": "This discovery/profile pass records exact version-pinned file hashes, schemas and aggregate type/missingness counts only. It does not emit raw PET-preform observations and does not promote the dataset until process, material, machine/mould, measurement and quality semantics are reviewed."
    }
    Path(args.output).write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(json.dumps({"files": [{"name": f["name"], "bytes": f["sizeBytes"], "sha256": f["sha256"], "format": f.get("format")} for f in files], "tableCount": len(tables), "largestTable": largest}, indent=2))


if __name__ == "__main__":
    main()
