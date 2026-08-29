#!/usr/bin/env python3
from __future__ import annotations
import argparse, hashlib, io, json, subprocess, tempfile, urllib.parse, urllib.request, zipfile
from pathlib import Path, PurePosixPath
import openpyxl

ROOT=Path(__file__).resolve().parents[1]
CONTRACT=ROOT/'data/public-benchmark-contracts/pmc4753395-hdpe-cenosphere-v1.json'
UA='MouldMaster-Educational-Evidence-Profiler/1.0'


def get(url,accept='*/*'):
    req=urllib.request.Request(url,headers={'Accept':accept,'User-Agent':UA})
    with urllib.request.urlopen(req,timeout=120) as r: return r.read(),r.geturl(),r.headers.get('Content-Type')

def safe(name):
    p=PurePosixPath(name); return not p.is_absolute() and '..' not in p.parts

def s3_https(url):
    if not url.startswith('s3://pmc-oa-opendata/'): raise RuntimeError('unexpected S3 URL')
    return 'https://pmc-oa-opendata.s3.amazonaws.com/'+url[len('s3://pmc-oa-opendata/'):]

def retrieve(c):
    raw,_,_=get(c['source']['cloudMetadataUrl'],'application/json,*/*'); meta=json.loads(raw.decode('utf-8'))
    if meta.get('pmcid')!=c['source']['pmcid'] or str(meta.get('license_code')).upper().replace('-',' ')!='CC BY': raise RuntimeError('PMC metadata identity/licence mismatch')
    target=c['source']['supplementLabel'].lower(); matches=[]
    for u in meta.get('media_urls') or []:
        if Path(urllib.parse.urlparse(str(u)).path).name.lower()==target: matches.append(str(u))
    if len(matches)!=1: raise RuntimeError('exact supplement media object not found')
    outer,_,_=get(s3_https(matches[0]),'application/zip,*/*')
    if hashlib.sha256(outer).hexdigest()!=c['stage1Evidence']['supplementSha256']: raise RuntimeError('supplement hash drift')
    with zipfile.ZipFile(io.BytesIO(outer)) as z:
        rars=[i for i in z.infolist() if not i.is_dir() and Path(i.filename).suffix.lower()=='.rar']
        if len(rars)!=1 or not safe(rars[0].filename): raise RuntimeError('nested RAR structure drift')
        rar=z.read(rars[0])
    if hashlib.sha256(rar).hexdigest()!=c['stage1Evidence']['nestedRarSha256']: raise RuntimeError('RAR hash drift')
    with tempfile.TemporaryDirectory(prefix='mouldmaster-pmc-stage2-') as td:
        p=Path(td)/'source.rar'; p.write_bytes(rar)
        listing=subprocess.run(['7z','l','-slt','-ba',str(p)],check=True,capture_output=True,text=True).stdout
        names=[line.split(' = ',1)[1].strip() for line in listing.splitlines() if line.startswith('Path = ')]
        measured=[n for n in names if Path(n).name.lower()=='tensile-data.xlsx']
        theoretical=[n for n in names if 'porfiri' in Path(n).name.lower() and n.lower().endswith('.xlsx')]
        if len(measured)!=1 or len(theoretical)!=1: raise RuntimeError('workbook identity drift')
        m=subprocess.run(['7z','x','-so',str(p),measured[0]],check=True,capture_output=True).stdout
        t=subprocess.run(['7z','x','-so',str(p),theoretical[0]],check=True,capture_output=True).stdout
    if hashlib.sha256(m).hexdigest()!=c['stage1Evidence']['measuredWorkbookSha256']: raise RuntimeError('measured workbook hash drift')
    if hashlib.sha256(t).hexdigest()!=c['stage1Evidence']['theoreticalWorkbookSha256']: raise RuntimeError('theoretical workbook hash drift')
    return m, measured[0], hashlib.sha256(raw).hexdigest()

def sheet_profile(ws):
    labels=[]; label_counts={'Strain (%)':0,'Stress (MPa)':0}; numeric_by_col=[]
    for col in range(1,ws.max_column+1):
        numeric=0; first=None; last=None; col_labels=[]
        for row in range(1,ws.max_row+1):
            v=ws.cell(row=row,column=col).value
            if isinstance(v,bool) or v is None: continue
            if isinstance(v,(int,float)):
                numeric+=1; first=row if first is None else first; last=row
            elif isinstance(v,str) and v.strip():
                s=' '.join(v.split())[:120]; col_labels.append(s)
                if s in label_counts: label_counts[s]+=1
                labels.append({'row':row,'column':col,'label':s})
        numeric_by_col.append({'column':col,'numericCells':numeric,'firstNumericRow':first,'lastNumericRow':last,'textLabels':col_labels[:12]})
    return {'rows':ws.max_row,'columns':ws.max_column,'textLabelCells':labels[:80],'textLabelCounts':label_counts,'numericCellsByColumn':numeric_by_col,'numericCellsTotal':sum(x['numericCells'] for x in numeric_by_col),'rawNumericValuesEmitted':False}

def main():
    a=argparse.ArgumentParser(); a.add_argument('--output',type=Path,required=True); a.add_argument('--retrieved-date',required=True); args=a.parse_args()
    c=json.loads(CONTRACT.read_text()); measured,name,metadata_sha=retrieve(c)
    with tempfile.NamedTemporaryFile(suffix='.xlsx') as f:
        f.write(measured); f.flush(); wb=openpyxl.load_workbook(f.name,data_only=False,read_only=False)
        composition=c['stage2Rules']['compositionSheets']; profiles={}
        for s in composition:
            if s not in wb.sheetnames: raise RuntimeError(f'missing composition sheet {s}')
            profiles[s]=sheet_profile(wb[s])
        if 'PLots' not in wb.sheetnames: raise RuntimeError('PLots exclusion sheet missing')
        plot=sheet_profile(wb['PLots'])
    strain=sum(p['numericCellsByColumn'][i]['numericCells'] for p in profiles.values() for i in range(p['columns']) if 'Strain (%)' in p['numericCellsByColumn'][i]['textLabels'])
    stress=sum(p['numericCellsByColumn'][i]['numericCells'] for p in profiles.values() for i in range(p['columns']) if 'Stress (MPa)' in p['numericCellsByColumn'][i]['textLabels'])
    candidate=sum(p['numericCellsTotal'] for p in profiles.values())
    header_ok=all(p['columns']==10 and p['textLabelCounts']=={'Strain (%)':5,'Stress (MPa)':5} and all(len(x['textLabels'])==1 for x in p['numericCellsByColumn']) for p in profiles.values())
    complete=header_ok and candidate==strain+stress and strain>0 and stress>0
    result={'schema':1,'status':'semantic-profile-complete-needs-acceptance' if complete else 'semantic-profile-needs-review','retrievedDate':args.retrieved_date,'source':{'datasetId':c['datasetId'],'datasetDoi':c['source']['datasetDoi'],'license':c['source']['license'],'metadataSha256':metadata_sha,'measuredWorkbookMember':name,'measuredWorkbookSha256':hashlib.sha256(measured).hexdigest()},'semanticProfile':{'compositionSheets':profiles,'excludedPlotSheet':{'sheet':'PLots','numericCells':plot['numericCellsTotal']},'pairedHeaderPatternValidated':header_ok,'sourceDeliveredStressTraceValues':stress,'sourceDeliveredStrainTraceValues':strain,'sourceDeliveredStressStrainTraceValues':candidate,'theoreticalWorkbookValuesAccepted':0,'injectionMachineTimeSeriesSamplesAdded':0,'rawNumericValuesEmitted':False},'acceptance':{'countsAsFullyProfiledMeasuredDataset':False,'stage3AcceptanceRequired':complete,'acceptedInjectionProcessTimeSeriesSamplesAdded':0},'evidenceBoundary':c['evidenceBoundary']}
    args.output.parent.mkdir(parents=True,exist_ok=True); args.output.write_text(json.dumps(result,indent=2)+'\n')
    print(json.dumps({'status':result['status'],'pairedHeaderPatternValidated':header_ok,'stressValues':stress,'strainValues':strain,'traceValues':candidate,'plotNumericExcluded':plot['numericCellsTotal']},indent=2))
if __name__=='__main__': main()
