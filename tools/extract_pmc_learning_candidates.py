#!/usr/bin/env python3
"""Extract compact, unreviewed PMC4753395 stress-strain authoring candidates.

The exact benchmarked Tensile-Data.xlsx is re-retrieved through the current PMC
Article Datasets route by ``prove_pmc_hdpe_source``. This extractor reads only
six governed stress/strain pairs, retains source values, removes only backward
strain points to obtain an explicitly monotonic display coordinate, and performs
no interpolation or synthetic curve fitting.
"""
from __future__ import annotations

import hashlib
import io
import json
import math
from pathlib import Path

from openpyxl import load_workbook
from prove_pmc_hdpe_source import EXPECTED_WORKBOOK_SHA, retrieve_workbook, sha256

ROOT=Path(__file__).resolve().parents[1]
OUT=ROOT/"measured-source-proof"/"pmc-unreviewed-learning-candidates.json"
ARTIFACT="Tensile-Data.xlsx"
FINGERPRINT="sha256:"+EXPECTED_WORKBOOK_SHA

CHANNELS={
    "HDPE!A:B":("HDPE","A","B","hdpe-tensile-stress-trace-replicate-1","HDPE tensile replicate 1"),
    "HDPE!C:D":("HDPE","C","D","hdpe-tensile-stress-trace-replicate-2","HDPE tensile replicate 2"),
    "HDPE!E:F":("HDPE","E","F","hdpe-tensile-stress-trace-replicate-3","HDPE tensile replicate 3"),
    "HDPE20!A:B":("HDPE20","A","B","hdpe20-tensile-stress-trace-replicate-1","HDPE20 tensile replicate 1"),
    "HDPE40!A:B":("HDPE40","A","B","hdpe40-tensile-stress-trace-replicate-1","HDPE40 tensile replicate 1"),
    "HDPE60!A:B":("HDPE60","A","B","hdpe60-tensile-stress-trace-replicate-1","HDPE60 tensile replicate 1"),
}


def sha(value) -> str:
    raw=json.dumps(value,sort_keys=True,separators=(",",":"),ensure_ascii=False).encode()
    return "sha256:"+hashlib.sha256(raw).hexdigest()


def numeric(value):
    if isinstance(value,bool) or not isinstance(value,(int,float)): return None
    v=float(value)
    return v if math.isfinite(v) else None


def compact_pair(ws, xcol: str, ycol: str, target: int=400):
    raw=[]
    for row in range(3, ws.max_row+1):
        x=numeric(ws[f"{xcol}{row}"].value); y=numeric(ws[f"{ycol}{row}"].value)
        if x is not None and y is not None: raw.append((row,x,y))
    if len(raw)<3: raise AssertionError(f"{ws.title}!{xcol}:{ycol}: fewer than three numeric pairs")
    monotonic=[]; backward=0; last=None
    for item in raw:
        if last is None or item[1] >= last:
            monotonic.append(item); last=item[1]
        else:
            backward+=1
    if len(monotonic)<3: raise AssertionError(f"{ws.title}!{xcol}:{ycol}: insufficient monotonic source-order pairs")
    if len(monotonic)<=target:
        selected=monotonic; method="source-order-nondecreasing-strain-subsequence-no-interpolation"
    else:
        # Reserve one slot for the measured peak stress, then sample source order uniformly.
        slots=target-1
        idx={round(i*(len(monotonic)-1)/(slots-1)) for i in range(slots)}
        peak=max(range(len(monotonic)),key=lambda i:monotonic[i][2]); idx.add(peak)
        if len(idx)>target:
            removable=sorted(i for i in idx if i!=peak and i not in {0,len(monotonic)-1})
            idx.remove(removable[len(removable)//2])
        selected=[monotonic[i] for i in sorted(idx)]
        method="source-order-nondecreasing-strain-subsequence-uniform-plus-peak-no-interpolation"
    return {
        "rawNumericPairCount":len(raw),
        "monotonicPairCount":len(monotonic),
        "backwardStrainPairsExcluded":backward,
        "sourceRowStart":raw[0][0],
        "sourceRowEnd":raw[-1][0],
        "reductionMethod":method,
        "x":[v[1] for v in selected],
        "y":[v[2] for v in selected],
    }


def signal(workbook, channel: str) -> dict:
    sheet,xcol,ycol,semantic,label=CHANNELS[channel]
    ws=workbook[sheet]
    compact=compact_pair(ws,xcol,ycol)
    rep={
        "xSemantic":"strain","xUnit":"%","xDirection":"increasing",
        "reductionMethod":compact["reductionMethod"],
        "originalPointCount":compact["rawNumericPairCount"],
        "x":compact["x"],"y":compact["y"],
    }
    return {
        "id":channel.replace("!","-").replace(":","-").lower(),
        "label":label,"sourceChannel":channel,"semantic":semantic,"unit":"MPa",
        "sourceDiagnostics":{
            "rawNumericPairCount":compact["rawNumericPairCount"],
            "monotonicPairCount":compact["monotonicPairCount"],
            "backwardStrainPairsExcluded":compact["backwardStrainPairsExcluded"],
            "sourceRowRangeInclusive":[compact["sourceRowStart"],compact["sourceRowEnd"]],
        },
        "representation":rep,"representationFingerprint":sha(rep),
    }


def candidate(workbook, cid: str, cases: list[str], channels: list[str], boundary: str) -> dict:
    signals=[signal(workbook,ch) for ch in channels]
    return {
        "candidateId":cid,"datasetId":"pmc4753395-hdpe-cenosphere-v1",
        "sourceArtifact":ARTIFACT,"sourceFingerprint":FINGERPRINT,
        "sourceScope":{"workbook":ARTIFACT,"governedSourcePairs":channels,"selection":"direct stress/strain worksheet pairs; source order preserved; no interpolation"},
        "signals":signals,"candidateFingerprint":sha(signals),"suggestedCatalogueCases":cases,
        "evidenceBoundary":boundary,
    }


def main() -> int:
    _,_,blob,_,_=retrieve_workbook()
    if sha256(blob)!=EXPECTED_WORKBOOK_SHA: raise AssertionError("PMC workbook identity drift")
    workbook=load_workbook(io.BytesIO(blob),read_only=True,data_only=True)
    candidates=[
        candidate(workbook,"PMC-HDPE-TRACE-EXCURSION-01",["MLM-031"],["HDPE!A:B"],"One direct HDPE tensile stress-versus-strain trace. A trace excursion is observable; it does not identify a production-process cause."),
        candidate(workbook,"PMC-HDPE-REPLICATE-TRACES-01",["MLM-053"],["HDPE!A:B","HDPE!C:D","HDPE!E:F"],"Three direct HDPE specimen traces support repeatability comparison. Specimen-to-specimen differences remain material-test evidence, not production root-cause proof."),
        candidate(workbook,"PMC-COMPOSITION-TRACE-COMPARISON-01",["MLM-054"],["HDPE!A:B","HDPE20!A:B","HDPE40!A:B","HDPE60!A:B"],"One direct trace from each source-defined composition supports bounded composition comparison; it does not establish universal material behavior or a moulding-process mechanism."),
    ]
    doc={"schemaVersion":1,"status":"unreviewed-source-derived-candidates","promotionEligible":False,"candidateCount":len(candidates),"candidates":candidates,"boundary":"Hash-verified CC BY 4.0 material-test authoring data only. These candidates are not independently reviewed learner cases."}
    OUT.parent.mkdir(exist_ok=True)
    OUT.write_text(json.dumps(doc,indent=2)+"\n",encoding="utf-8")
    print(json.dumps({"status":doc["status"],"candidateCount":len(candidates),"catalogueCoverage":["MLM-031","MLM-053","MLM-054"],"pointCounts":{c["candidateId"]:[len(s["representation"]["x"]) for s in c["signals"]] for c in candidates}},separators=(",",":")))
    return 0


if __name__=="__main__": raise SystemExit(main())
