#!/usr/bin/env python3
from __future__ import annotations
import argparse,csv,hashlib,html,io,json,re,urllib.parse,urllib.request
from pathlib import Path

DATASET_ID='4h98rz9f92'; VERSION='3'; DOI='10.17632/4h98rz9f92.3'
TITLE='Dataset on Graphite nanoplatelet enhanced HDPE composites: tensile modulus, hardness and toughness'
LICENSE='CC BY 4.0'; COMPANION='10.1016/j.dib.2024.110987'
PAGE=f'https://data.mendeley.com/datasets/{DATASET_ID}/{VERSION}'
API=f'https://data.mendeley.com/public-api/datasets/{DATASET_ID}/files?folder_id=root&version={VERSION}'
UA='MouldMaster-HDPE-GNP-profiler/1.1'

def sha256(b):return hashlib.sha256(b).hexdigest()
def get(url,accept='*/*'):
    req=urllib.request.Request(url,headers={'User-Agent':UA,'Accept':accept})
    with urllib.request.urlopen(req,timeout=180) as r:return r.read(),r.headers,r.geturl()
def flatten(x):
    if isinstance(x,list):return x
    if isinstance(x,dict):
        for k in ('results','files','items','data'):
            if isinstance(x.get(k),list):return x[k]
    return []
def discover():
    out=[]
    try:
        raw,_,_=get(API,'application/json')
        for item in flatten(json.loads(raw.decode('utf-8'))):
            u=item.get('download_url') or item.get('downloadUrl') or (item.get('content_details') or {}).get('download_url')
            n=item.get('name') or item.get('filename') or item.get('file_name')
            if u:out.append({'name':n,'url':u,'publisherId':item.get('id')})
    except Exception:pass
    if out:return out
    raw,_,_=get(PAGE,'text/html')
    text=html.unescape(raw.decode('utf-8','replace')).replace('\\u002F','/').replace('\\/','/')
    pat=re.compile(rf'https://data\.mendeley\.com/public-files/datasets/{DATASET_ID}/files/([0-9a-fA-F-]{{36}})/file_downloaded')
    seen=set()
    for m in pat.finditer(text):
        u=m.group(0)
        if u not in seen:seen.add(u);out.append({'name':None,'url':u,'publisherId':m.group(1)})
    if not out:raise RuntimeError('No version-pinned Mendeley file links discovered')
    return out
def fname(final,h,fallback):
    cd=h.get('Content-Disposition')
    if cd:
        m=re.search(r'filename\*?=(?:UTF-8\'\'|")?([^";]+)',cd,re.I)
        if m:return urllib.parse.unquote(m.group(1).strip().strip('"'))
    n=Path(urllib.parse.urlparse(final).path).name
    return n if Path(n).suffix else fallback
def finite(v):
    try:float(str(v).strip());return True
    except Exception:return False
def profile_rows(headers,rows):
    missing=[0]*len(headers); numeric=[0]*len(headers); mins=[None]*len(headers); maxs=[None]*len(headers)
    for row in rows:
        vals=list(row[:len(headers)])+['']*max(0,len(headers)-len(row))
        for i,v in enumerate(vals):
            if v in (None,''):missing[i]+=1;continue
            try:x=float(str(v).strip());numeric[i]+=1;mins[i]=x if mins[i] is None else min(mins[i],x);maxs[i]=x if maxs[i] is None else max(maxs[i],x)
            except Exception:pass
    return [{'name':headers[i],'numericCount':numeric[i],'missing':missing[i],'min':mins[i],'max':maxs[i]} for i in range(len(headers))]
def profile_csv(raw):
    text=raw.decode('utf-8-sig','replace')
    try:d=csv.Sniffer().sniff(text[:65536],delimiters=',;\t|')
    except Exception:d=csv.excel
    allrows=list(csv.reader(io.StringIO(text),d));h=[str(x).strip() for x in allrows[0]] if allrows else [];rows=[r for r in allrows[1:] if any(str(x).strip() for x in r)] if allrows else []
    return {'format':'csv','rows':len(rows),'columns':len(h),'header':h,'columnStats':profile_rows(h,rows)}
def profile_xlsx(raw):
    from openpyxl import load_workbook
    wb=load_workbook(io.BytesIO(raw),read_only=False,data_only=True);sheets=[]
    for ws in wb.worksheets:
        matrix=[[cell.value for cell in row] for row in ws.iter_rows()]
        width=max((len(r) for r in matrix),default=0)
        matrix=[r+[None]*(width-len(r)) for r in matrix]
        nonempty=[r for r in matrix if any(v not in (None,'') for v in r)]
        first=['' if v is None else str(v).strip() for v in (nonempty[0] if nonempty else [])]
        body=nonempty[1:] if nonempty else []
        preview=[['' if v is None else str(v) for v in r] for r in nonempty[:5]]
        merged=[str(rng) for rng in ws.merged_cells.ranges]
        # Raw Data.xlsx uses two heading rows followed by 35 experimental rows.
        experimental=None
        if ws.title=='Sheet1' and len(nonempty)>=37:
            candidates=nonempty[2:]
            candidates=[r for r in candidates if finite(r[0]) and finite(r[1]) and finite(r[2]) and finite(r[3])]
            if len(candidates)==35:
                experimental={
                    'rows':35,
                    'inputColumns':{
                        'experimentNumber':{'index':0,'min':min(float(r[0]) for r in candidates),'max':max(float(r[0]) for r in candidates)},
                        'gnpFraction':{'index':1,'min':min(float(r[1]) for r in candidates),'max':max(float(r[1]) for r in candidates)},
                        'temperatureC':{'index':2,'min':min(float(r[2]) for r in candidates),'max':max(float(r[2]) for r in candidates)},
                        'pressureMPa':{'index':3,'min':min(float(r[3]) for r in candidates),'max':max(float(r[3]) for r in candidates)},
                    },
                    'outputNumericColumnRanges':[
                        {'index':i,'min':min(float(r[i]) for r in candidates if finite(r[i])),'max':max(float(r[i]) for r in candidates if finite(r[i])),'numericRows':sum(1 for r in candidates if finite(r[i]))}
                        for i in range(4,width) if sum(1 for r in candidates if finite(r[i]))>=30
                    ],
                }
        sheets.append({'name':ws.title,'rows':len(body),'columns':len(first),'header':first,'columnStats':profile_rows(first,body),'preview':preview,'mergedRanges':merged,'experimentalStructure':experimental})
    return {'format':'xlsx','sheets':sheets}
def main():
    ap=argparse.ArgumentParser();ap.add_argument('--output',default='hdpe-gnp-v3.json');a=ap.parse_args()
    files=[];tables=[]
    for i,rec in enumerate(discover(),1):
        raw,h,final=get(rec['url']);name=rec.get('name') or fname(final,h,f'publisher-file-{i}');ext=Path(name).suffix.lower()
        p={'name':name,'publisherId':rec.get('publisherId'),'sizeBytes':len(raw),'sha256':sha256(raw),'suffix':ext}
        if ext in {'.csv','.tsv','.txt'}:
            q=profile_csv(raw);p.update(q);tables.append({'file':name,**q})
        elif ext in {'.xlsx','.xlsm'}:
            q=profile_xlsx(raw);p.update(q)
            for s in q['sheets']:tables.append({'file':name,**s})
        else:p['format']=ext.lstrip('.') or 'unknown'
        files.append(p)
    biggest=max(tables,key=lambda x:x.get('rows',0)*x.get('columns',0),default={})
    headers=' '.join(str(x.get('header') or '')+' '+str(x.get('preview') or '') for x in tables).lower()
    markers={k:(k in headers) for k in ['injection','temperature','pressure','gnp','graphite','tensile','toughness','hardness']}
    raw_data=next((s for f in files if f['name']=='Raw Data.xlsx' for s in f.get('sheets',[]) if s['name']=='Sheet1'),None)
    payload={'schema':2,'status':'profile-generated-review-required','completedDate':'2026-08-28','source':{'datasetId':DATASET_ID,'version':VERSION,'doi':DOI,'title':TITLE,'license':LICENSE,'publisher':'Mendeley Data','page':PAGE,'peerReviewedCompanion':COMPANION,'materialContext':'HDPE with graphite nanoplatelet reinforcement'},'files':files,'tableCount':len(tables),'largestTable':biggest,'semanticHeaderMarkers':markers,'rawDataExperimentalStructure':(raw_data or {}).get('experimentalStructure'),'acceptedMeasuredRecords':0,'acceptedMeasuredTimeSeriesSamples':0,'rawSourceRowsCommitted':False,'boundary':'Exact version-3 publisher files are fingerprinted. Raw Data.xlsx is inspected using its two-row/merged heading structure and must reconcile to 35 experimental injection-moulding configurations before promotion. Peer-reviewed semantics define GNP fraction, injection temperature (deg C), injection pressure (MPa), tensile modulus (GPa), toughness (MPa), and Vickers hardness (HV). Record-level measurements do not count as waveform samples.'}
    Path(a.output).write_text(json.dumps(payload,indent=2,ensure_ascii=False)+'\n',encoding='utf-8')
    print(json.dumps({'files':[(f['name'],f['format'],f['sizeBytes'],f['sha256']) for f in files],'tableCount':len(tables),'rawData':raw_data,'semanticHeaderMarkers':markers},indent=2,ensure_ascii=False))
if __name__=='__main__':main()
