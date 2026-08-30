#!/usr/bin/env python3
from __future__ import annotations

import argparse
import datetime as dt
import hashlib
import io
import json
import urllib.request
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

ROOT = Path(__file__).resolve().parents[1]
CONTRACT = ROOT / "data/public-benchmark-contracts/yxz2w7ctnh-v1.json"
UA = "MouldMaster-Educational-Evidence-Profiler/1.0"
DATASET_ID = "yxz2w7ctnh"
VERSION = 1
API_ROOT = "https://api.data.mendeley.com"


def get(url: str) -> tuple[bytes, str]:
    req = urllib.request.Request(url, headers={"User-Agent": UA, "Accept": "*/*"})
    with urllib.request.urlopen(req, timeout=120) as r:
        return r.read(), r.geturl()


def listing():
    url = f"https://data.mendeley.com/public-api/datasets/{DATASET_ID}/files?folder_id=root&version={VERSION}"
    raw, _ = get(url)
    x = json.loads(raw.decode("utf-8"))
    if isinstance(x, list):
        return x
    if isinstance(x, dict):
        for key in ("results", "files", "items", "data"):
            if isinstance(x.get(key), list):
                return x[key]
    return []


def fid(item):
    return str(item.get("id") or item.get("file_id") or item.get("uuid") or "")


def furl(item):
    d = item.get("content_details") or item.get("contentDetails") or {}
    for key in ("download_url", "downloadUrl"):
        if d.get(key):
            return str(d[key])
        if item.get(key):
            return str(item[key])
    return f"{API_ROOT}/datasets/{DATASET_ID}/files/{fid(item)}/file_downloaded?version={VERSION}"


def token(v):
    if isinstance(v, dt.datetime): return v.isoformat()
    if isinstance(v, dt.date): return v.isoformat()
    return repr(v)


def sheet_hash(ws):
    h = hashlib.sha256()
    for row in ws.iter_rows():
        for c in row:
            if c.value is None: continue
            h.update(c.coordinate.encode()); h.update(b"\0")
            h.update(str(c.data_type).encode()); h.update(b"\0")
            h.update(token(c.value).encode("utf-8", "replace")); h.update(b"\n")
    return h.hexdigest()


def count_numeric_region(ws, a1_range: str):
    min_col, min_row, max_col, max_row = range_boundaries(a1_range)
    numeric = formulas = nonempty = 0
    rows_with_numeric = 0
    for r in range(min_row, max_row + 1):
        row_numeric = 0
        for c in range(min_col, max_col + 1):
            cell = ws.cell(r, c)
            if cell.value is None: continue
            nonempty += 1
            if cell.data_type == "f": formulas += 1
            elif isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                numeric += 1; row_numeric += 1
        if row_numeric: rows_with_numeric += 1
    return {"numericConstantCells": numeric, "formulaCellsExcluded": formulas, "nonEmptyCells": nonempty, "rowsWithNumericConstants": rows_with_numeric}


def count_direct_columns(ws, columns, start_row, end_row):
    numeric = formulas = 0
    by_column = {}
    rows = set()
    for col in columns:
        cn = cf = 0
        for r in range(start_row, end_row + 1):
            cell = ws[f"{col}{r}"]
            if cell.value is None: continue
            if cell.data_type == "f":
                formulas += 1; cf += 1
            elif isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                numeric += 1; cn += 1; rows.add(r)
        by_column[col] = {"numericConstantCells": cn, "formulaCellsExcluded": cf}
    return {"numericConstantCells": numeric, "formulaCellsExcluded": formulas, "rowsWithDirectMeasurements": len(rows), "byColumn": by_column}


def main():
    ap = argparse.ArgumentParser(); ap.add_argument("--output", type=Path, required=True); ap.add_argument("--retrieved-date", required=True); args = ap.parse_args()
    contract = json.loads(CONTRACT.read_text(encoding="utf-8"))
    pub = {fid(x): x for x in listing()}
    unique_sheets = {}
    duplicate_occurrences = {}
    files_out = []
    for expected in contract["source"]["files"]:
        item = pub.get(expected["id"])
        if item is None: raise RuntimeError(f"publisher file disappeared: {expected['id']}")
        data, final_url = get(furl(item))
        digest = hashlib.sha256(data).hexdigest()
        if digest.lower() != expected["sha256"].lower(): raise RuntimeError(f"publisher SHA mismatch for {expected['name']}")
        wb = load_workbook(io.BytesIO(data), read_only=False, data_only=False)
        sheet_rows = []
        for ws in wb.worksheets:
            sh = sheet_hash(ws)
            duplicate_occurrences.setdefault(sh, []).append({"file": expected["name"], "sheet": ws.title})
            if sh not in unique_sheets:
                unique_sheets[sh] = ws
            sheet_rows.append({"sheet": ws.title, "sheetContentSha256": sh})
        files_out.append({"fileName": expected["name"], "sha256": digest, "publisherSha256Matched": True, "resolvedUrl": final_url, "sheets": sheet_rows, "rawPublisherFileCommitted": False})

    rules = contract["semanticRules"]
    selected = []
    total_direct = 0
    for spec in rules["tensileInjectionBlocks"]:
        prefix = contract["verifiedUniqueSheetHashes"][spec["sheet"]]
        matches = [(h, ws) for h, ws in unique_sheets.items() if h.startswith(prefix) and ws.title == spec["sheet"]]
        if len(matches) != 1: raise RuntimeError(f"unique tensile sheet mismatch: {spec['sheet']}")
        h, ws = matches[0]
        marker = str(ws[spec["marker"]].value or "")
        if spec["expectedMarkerContains"].lower() not in marker.lower(): raise RuntimeError(f"injection marker drifted: {spec['sheet']}")
        prof = count_numeric_region(ws, spec["numericRegion"])
        total_direct += prof["numericConstantCells"]
        selected.append({"sheet": spec["sheet"], "sheetContentSha256": h, "kind": "tensile-injection-block", "markerCoordinate": spec["marker"], "markerText": marker, "numericRegion": spec["numericRegion"], **prof})

    for spec in rules["bendingDirectMeasurementBlocks"]:
        prefix = contract["verifiedUniqueSheetHashes"][spec["sheet"]]
        matches = [(h, ws) for h, ws in unique_sheets.items() if h.startswith(prefix) and ws.title == spec["sheet"]]
        if len(matches) != 1: raise RuntimeError(f"unique bending sheet mismatch: {spec['sheet']}")
        h, ws = matches[0]
        marker = str(ws[spec["marker"]].value or "")
        if spec["expectedMarkerContains"].lower() not in marker.lower(): raise RuntimeError(f"injection marker drifted: {spec['sheet']}")
        start, end = spec["dataRows"]
        prof = count_direct_columns(ws, spec["directColumns"], start, end)
        total_direct += prof["numericConstantCells"]
        selected.append({"sheet": spec["sheet"], "sheetContentSha256": h, "kind": "bending-injection-direct-measurements", "markerCoordinate": spec["marker"], "markerText": marker, "directColumns": spec["directColumns"], "headers": spec["headers"], "dataRows": spec["dataRows"], **prof})

    duplicate_groups = [v for v in duplicate_occurrences.values() if len(v) > 1]
    result = {
        "schema": 1,
        "status": "completed-profiled-record-level-injection-mechanical-testing",
        "retrievedDate": args.retrieved_date,
        "source": {"datasetId": contract["datasetId"], "datasetDoi": contract["source"]["datasetDoi"], "license": contract["source"]["license"], "version": VERSION},
        "files": files_out,
        "profile": {
            "publisherFilesVerified": len(files_out),
            "uniqueWorksheetContentHashes": len(unique_sheets),
            "duplicateWorksheetGroupsExcluded": duplicate_groups,
            "selectedInjectionBlocks": selected,
            "directRecordLevelInjectionMeasuredValues": total_direct,
            "acceptedMeasuredTimeSeriesSamples": 0,
            "rawRowsOrNumericValuesEmitted": False
        },
        "acceptance": {
            "countsAsFullyProfiledMeasuredDataset": total_direct > 0,
            "recordLevelOnly": True,
            "acceptedMeasuredTimeSeriesSamples": 0,
            "energySheetsExcluded": True,
            "impactSheetsExcludedForRouteAmbiguity": True,
            "duplicateWorksheetsExcluded": True,
            "formulaCellsExcludedAsDerived": True
        },
        "retrieval": {"rawPublisherFilesCommitted": False, "rawRowsOrNumericValuesUploadedAsArtifact": False},
        "evidenceBoundary": contract["evidenceBoundary"]
    }
    args.output.parent.mkdir(parents=True, exist_ok=True); args.output.write_text(json.dumps(result, indent=2, ensure_ascii=False)+"\n", encoding="utf-8")
    print(json.dumps({"status": result["status"], "directRecordLevelInjectionMeasuredValues": total_direct, "duplicateWorksheetGroups": len(duplicate_groups)}, indent=2))


if __name__ == "__main__": main()
