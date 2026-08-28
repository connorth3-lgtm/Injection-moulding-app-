#!/usr/bin/env python3
from __future__ import annotations
import argparse, ast, csv, hashlib, io, json, math, re, tempfile, urllib.request, zipfile
from pathlib import Path

BASE='https://raw.githubusercontent.com/sc4t1m/scatimdata/main'
ARCHIVES=['dataset1.zip','dataset2.zip','dataset3.zip']
EXPECTED_POINTS=2049
EXPECTED_SAMPLING_MS=6

def sha256_bytes(b: bytes) -> str:
    return hashlib.sha256(b).hexdigest()

def download(url: str) -> bytes:
    req=urllib.request.Request(url, headers={'User-Agent':'MouldMaster/1.0 evidence profiler'})
    with urllib.request.urlopen(req, timeout=90) as r:
        return r.read()

def numeric_vector_len(value: str) -> int|None:
    s=(value or '').strip()
    if len(s)<8: return None
    # Common serialized-vector formats: JSON/Python lists and whitespace/semicolon separated vectors.
    if s[0] in '[(' and s[-1] in '])':
        try:
            x=ast.literal_eval(s)
            if isinstance(x,(list,tuple)) and len(x)>=16 and all(isinstance(v,(int,float)) for v in x[:min(len(x),64)]):
                return len(x)
        except Exception:
            pass
    if s.count(';')>=15:
        parts=[p for p in s.split(';') if p.strip()]
        try:
            [float(p.replace(',','.')) for p in parts[:32]]
            return len(parts)
        except Exception: pass
    if s.count(' ')>=15 and ',' not in s[:100]:
        parts=s.split()
        try:
            [float(p) for p in parts[:32]]
            return len(parts)
        except Exception: pass
    return None

def profile_csv(raw: bytes, name: str) -> dict:
    sample=raw[:65536].decode('utf-8-sig', errors='replace')
    try:
        dialect=csv.Sniffer().sniff(sample, delimiters=',;\t|')
    except Exception:
        dialect=csv.excel
    text=io.TextIOWrapper(io.BytesIO(raw), encoding='utf-8-sig', errors='replace', newline='')
    reader=csv.reader(text, dialect)
    try: header=next(reader)
    except StopIteration: return {'path':name,'format':'csv','rows':0,'columns':0}
    rows=0
    vector_cols={}
    nonempty=[0]*len(header)
    for row in reader:
        if not row: continue
        rows+=1
        if len(row)<len(header): row += ['']*(len(header)-len(row))
        for i,val in enumerate(row[:len(header)]):
            if val.strip(): nonempty[i]+=1
            if i not in vector_cols or vector_cols[i].get('length') is None:
                n=numeric_vector_len(val)
                if n:
                    vector_cols[i]={'name':header[i] if i<len(header) else f'col{i}','length':n}
    return {
        'path':name,'format':'csv','rows':rows,'columns':len(header),'headers':header[:200],
        'vectorColumns':list(vector_cols.values()),'nonEmptyCounts':nonempty[:200]
    }

def profile_json(raw: bytes, name: str) -> dict:
    try: obj=json.loads(raw)
    except Exception as e: return {'path':name,'format':'json','parseError':str(e)}
    out={'path':name,'format':'json','topType':type(obj).__name__}
    if isinstance(obj,list): out['records']=len(obj)
    elif isinstance(obj,dict): out['keys']=list(obj)[:200]
    return out

def profile_xlsx(raw: bytes, name: str) -> dict:
    try:
        from openpyxl import load_workbook
    except Exception as e:
        return {'path':name,'format':'xlsx','parseError':f'openpyxl unavailable: {e}'}
    with tempfile.NamedTemporaryFile(suffix='.xlsx') as f:
        f.write(raw); f.flush()
        wb=load_workbook(f.name, read_only=True, data_only=True)
        sheets=[]
        for ws in wb.worksheets:
            rows=ws.max_row or 0; cols=ws.max_column or 0
            header=[]
            if rows:
                header=[c.value for c in next(ws.iter_rows(min_row=1,max_row=1))]
            sheets.append({'name':ws.title,'rowsIncludingHeader':rows,'columns':cols,'headers':header[:200]})
        return {'path':name,'format':'xlsx','sheets':sheets}

def profile_numpy(raw: bytes, name: str) -> dict:
    try:
        import numpy as np
    except Exception as e: return {'path':name,'format':'numpy','parseError':f'numpy unavailable: {e}'}
    bio=io.BytesIO(raw)
    try:
        obj=np.load(bio, allow_pickle=False)
        if hasattr(obj,'files'):
            arrays={k:list(obj[k].shape) for k in obj.files}
            return {'path':name,'format':'npz','arrays':arrays}
        return {'path':name,'format':'npy','shape':list(obj.shape),'dtype':str(obj.dtype)}
    except Exception as e: return {'path':name,'format':'numpy','parseError':str(e)}


def profile_hdf5(raw: bytes, name: str) -> dict:
    try:
        import h5py
    except Exception as e:
        return {'path':name,'format':'h5','parseError':f'h5py unavailable: {e}'}
    with tempfile.NamedTemporaryFile(suffix='.h5') as f:
        f.write(raw); f.flush()
        datasets=[]
        with h5py.File(f.name,'r') as h:
            def visit(path,obj):
                if isinstance(obj,h5py.Dataset):
                    attrs={}
                    for k,v in obj.attrs.items():
                        try:
                            if hasattr(v,'tolist'): v=v.tolist()
                            if isinstance(v,bytes): v=v.decode('utf-8','replace')
                            attrs[str(k)]=v if isinstance(v,(str,int,float,bool,list,tuple)) else str(v)
                        except Exception:
                            attrs[str(k)]=str(v)
                    datasets.append({'path':path,'shape':list(obj.shape),'dtype':str(obj.dtype),'size':int(obj.size),'attrs':attrs})
            h.visititems(visit)
        return {'path':name,'format':'h5','datasets':datasets}

def profile_member(z: zipfile.ZipFile, info: zipfile.ZipInfo) -> dict:
    raw=z.read(info)
    ext=Path(info.filename).suffix.lower()
    base={'path':info.filename,'sizeBytes':len(raw),'sha256':sha256_bytes(raw)}
    try:
        if ext in {'.csv','.txt','.tsv'}: base.update(profile_csv(raw,info.filename))
        elif ext=='.json': base.update(profile_json(raw,info.filename))
        elif ext in {'.xlsx','.xlsm'}: base.update(profile_xlsx(raw,info.filename))
        elif ext in {'.npy','.npz'}: base.update(profile_numpy(raw,info.filename))
        elif ext in {'.h5','.hdf5'}: base.update(profile_hdf5(raw,info.filename))
        else: base['format']=ext.lstrip('.') or 'unknown'
    except Exception as e:
        base['profileError']=f'{type(e).__name__}: {e}'
    return base

def infer_samples(files: list[dict]) -> dict:
    vectors=[]
    matrix_candidates=[]
    row_counts=[]
    for f in files:
        if f.get('format')=='csv':
            row_counts.append(f.get('rows') or 0)
            for v in f.get('vectorColumns') or []:
                nm=(v.get('name') or '').lower()
                if any(k in nm for k in ['pressure','flow','druck','volumenstrom','injection']):
                    vectors.append({'file':f['path'],**v,'rows':f.get('rows',0)})
            # Published CSVs use one time column plus one column per injection cycle.
            nm=f['path'].lower()
            if any(k in nm for k in ['pressure','flow','druck','volumenstrom']):
                r=f.get('rows') or 0; c=f.get('columns') or 0
                if r >= 1000 and c > 1:
                    samples=r*(c-1)
                    matrix_candidates.append({'file':f['path'],'rows':r,'columns':c,'cycleColumns':c-1,'samples':samples,'timeAxisColumn':True})
        elif f.get('format')=='h5':
            for d in f.get('datasets') or []:
                nm=(f['path']+' '+d.get('path','')).lower()
                if any(k in nm for k in ['pressure','flow','druck','volumenstrom']) and (d.get('size') or 0)>=1000:
                    matrix_candidates.append({'file':f['path'],'dataset':d.get('path'),'shape':d.get('shape'),'samples':d.get('size'),'hdf5':True})
        elif f.get('format') in {'npy','npz'}:
            pass
    vector_samples=sum(v['rows']*v['length'] for v in vectors if v.get('length') and v.get('rows'))
    matrix_samples=sum(m['samples'] for m in matrix_candidates)
    return {'vectorColumns':vectors,'matrixCandidates':matrix_candidates,'detectedTimeSeriesScalarSamples':max(vector_samples,matrix_samples),'tabularRowCounts':row_counts}

def main():
    ap=argparse.ArgumentParser(); ap.add_argument('--output',default='data/public-benchmark-results/scatimdata-v1.json'); args=ap.parse_args()
    archives=[]; all_files=[]
    for name in ARCHIVES:
        url=f'{BASE}/{name}'; b=download(url)
        rec={'name':name,'url':url,'sizeBytes':len(b),'sha256':sha256_bytes(b),'members':[]}
        with zipfile.ZipFile(io.BytesIO(b)) as z:
            infos=[i for i in z.infolist() if not i.is_dir()]
            rec['memberCount']=len(infos)
            for info in infos:
                p=profile_member(z,info); rec['members'].append(p); all_files.append(p)
        archives.append(rec)
    detected=infer_samples(all_files)
    payload={
        'schema_version':1,'status':'completed-public-measured-timeseries-benchmark','completed_date':'2026-08-28',
        'source':{'title':'High-resolution injection-moulding time-series datasets (scatimdata)','repository':'https://github.com/sc4t1m/scatimdata','license':'CC BY 4.0','peerReviewedCompanion':'10.3390/polym15040978'},
        'publishedStructure':{'samplingIntervalMs':6,'pointsPerTimeSeries':2049,'signalsPerCycle':['injection pressure curve','injection flow curve'],'cycleLinkage':True,'qualityPerCycle':['part weight','geometric dimension']},
        'archives':archives,'detected':detected,
        'acceptedMeasuredTimeSeriesSamples':detected['detectedTimeSeriesScalarSamples'],
        'boundary':'Counts are derived from lawfully downloaded CC BY 4.0 source archives. No raw rows are emitted by this profile. A scalar sample is one measured value in a time-series channel, not a moulding cycle. If automatic structure detection cannot establish the sample count, acceptedMeasuredTimeSeriesSamples remains zero until an adapter is added.'
    }
    out=Path(args.output); out.parent.mkdir(parents=True,exist_ok=True); out.write_text(json.dumps(payload,indent=2,ensure_ascii=False)+'\n')
    print(json.dumps({'archives':[(a['name'],a['sizeBytes'],a['memberCount']) for a in archives],'detectedSamples':payload['acceptedMeasuredTimeSeriesSamples']},indent=2))

if __name__=='__main__': main()
