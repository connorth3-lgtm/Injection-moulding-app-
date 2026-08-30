#!/usr/bin/env python3
from __future__ import annotations

import argparse
import hashlib
import json
import tempfile
import urllib.request
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

DATASET_ID = "4h98rz9f92"
VERSION = 3
DOI = "10.17632/4h98rz9f92.3"
EXPECTED_FILE_ID = "368356fe-618c-4eab-82e6-53dc86762943"
EXPECTED_FILE = "Raw Data.xlsx"
EXPECTED_SHA256 = "39210169aac62a1455603d37cdffaca93cf0c46189ea4258c5f3c0a4a37255c9"
PUBLIC_FILES_ENDPOINT = f"https://data.mendeley.com/public-api/datasets/{DATASET_ID}/files?folder_id=root&version={VERSION}"
API_ROOT = "https://api.data.mendeley.com"
UA = "MouldMaster-Educational-Evidence-Profiler/1.0"

EXPECTED_LABELS = {
    "B1":"Input","E1":"Output","A2":"Expt No.","B2":"GNP %","C2":"Temperature","D2":"Pressure (MPa)",
    "E2":"EXP 1","F2":"EXP 2","G2":"EXP 3","H2":"EXP 4","I2":"EXP 5","J2":"Average Tensile modulus",
    "K2":"EXP 1","L2":"EXP 2","M2":"EXP 3","N2":"EXP 4","O2":"EXP 5","P2":"Average Hardness",
    "Q2":"EXP 1","R2":"EXP 2","S2":"EXP 3","T2":"EXP 4","U2":"EXP 5","V2":"Average Toughness (J)",
    "E3":"Tensile modulus (GPa)","F3":"Tensile modulus (GPa)","G3":"Tensile modulus (GPa)","H3":"Tensile modulus (GPa)","I3":"Tensile modulus (GPa)",
    "K3":"(HV)","L3":"(HV)","M3":"(HV)","N3":"(HV)","O3":"(HV)",
    "Q3":"Toughness (J)","R3":"Toughness (J)","S3":"Toughness (J)","T3":"Toughness (J)","U3":"Toughness (J)"
}
ROLE_COLUMNS = {
    "experimentIdentifier": ["A"],
    "processInputs": ["B","C","D"],
    "tensileModulusDirectReplicates": ["E","F","G","H","I"],
    "tensileModulusDerivedAverage": ["J"],
    "hardnessDirectReplicates": ["K","L","M","N","O"],
    "hardnessDerivedAverage": ["P"],
    "toughnessDirectReplicates": ["Q","R","S","T","U"],
    "toughnessDerivedAverage": ["V"]
}
EXPECTED_ROWS_PER_COLUMN = 35
EXPECTED_TOTAL_NUMERIC = 770
EXPECTED_DIRECT_MEASUREMENTS = 525
EXPECTED_DERIVED_AVERAGES = 105
EXPECTED_PROCESS_INPUT_VALUES = 105
EXPECTED_EXPERIMENT_IDENTIFIERS = 35


def get(url: str, accept: str = "*/*"):
    req=urllib.request.Request(url,headers={"Accept":accept,"User-Agent":UA})
    with urllib.request.urlopen(req,timeout=120) as r: return r.read(),r.geturl()


def flatten_files(payload):
    if isinstance(payload,list): return payload
    if isinstance(payload,dict):
        for key in ("results","files","items","data"):
            if isinstance(payload.get(key),list): return payload[key]
    return []


def fid(x): return str(x.get("id") or x.get("file_id") or x.get("uuid") or "").strip()
def fname(x): return str(x.get("filename") or x.get("name") or "").strip()


def furl(x):
    d=x.get("content_details") or x.get("contentDetails") or {}
    for u in (d.get("download_url"),d.get("downloadUrl"),x.get("download_url"),x.get("downloadUrl")):
        if u: return str(u)
    return f"{API_ROOT}/datasets/{DATASET_ID}/files/{fid(x)}/file_downloaded?version={VERSION}"


def psha(x):
    d=x.get("content_details") or x.get("contentDetails") or {}
    return str(d.get("sha256_hash") or d.get("sha256Hash") or x.get("sha256") or x.get("sha256_hash") or "").lower()


def main():
    ap=argparse.ArgumentParser(); ap.add_argument("--output",required=True); ap.add_argument("--retrieved-date",required=True); args=ap.parse_args()
    raw,_=get(PUBLIC_FILES_ENDPOINT,"application/json"); files=flatten_files(json.loads(raw.decode("utf-8")))
    matches=[x for x in files if fid(x)==EXPECTED_FILE_ID and fname(x)==EXPECTED_FILE]
    if len(matches)!=1: raise RuntimeError("exact HDPE/GNP raw workbook identity drifted")
    item=matches[0]
    if psha(item)!=EXPECTED_SHA256: raise RuntimeError("publisher HDPE/GNP SHA drifted")
    data,final_url=get(furl(item)); digest=hashlib.sha256(data).hexdigest()
    if digest!=EXPECTED_SHA256: raise RuntimeError(f"retrieved HDPE/GNP SHA drifted: {digest}")
    with tempfile.TemporaryDirectory() as td:
        p=Path(td)/EXPECTED_FILE; p.write_bytes(data); wb=load_workbook(p,read_only=True,data_only=False)
        if wb.sheetnames != ["Sheet1"]: raise RuntimeError(f"HDPE/GNP sheet set drifted: {wb.sheetnames}")
        ws=wb["Sheet1"]
        for coord,label in EXPECTED_LABELS.items():
            if str(ws[coord].value or "").strip()!=label: raise RuntimeError(f"HDPE/GNP label drift at {coord}: {ws[coord].value!r}")
        counts={}; formulas=0
        for row in ws.iter_rows():
            for cell in row:
                v=cell.value
                if isinstance(v,str) and v.startswith("="): formulas+=1
                elif isinstance(v,(int,float)) and not isinstance(v,bool):
                    c=get_column_letter(cell.column); counts[c]=counts.get(c,0)+1
    if set(counts)!=set("ABCDEFGHIJKLMNOPQRSTUV"): raise RuntimeError(f"HDPE/GNP numeric column set drifted: {sorted(counts)}")
    if any(counts[c]!=EXPECTED_ROWS_PER_COLUMN for c in counts): raise RuntimeError(f"HDPE/GNP row counts drifted: {counts}")
    role_counts={role:sum(counts[c] for c in cols) for role,cols in ROLE_COLUMNS.items()}
    direct=role_counts["tensileModulusDirectReplicates"]+role_counts["hardnessDirectReplicates"]+role_counts["toughnessDirectReplicates"]
    derived=role_counts["tensileModulusDerivedAverage"]+role_counts["hardnessDerivedAverage"]+role_counts["toughnessDerivedAverage"]
    process=role_counts["processInputs"]; identifiers=role_counts["experimentIdentifier"]; total=sum(counts.values())
    if total!=EXPECTED_TOTAL_NUMERIC or formulas!=0: raise RuntimeError(f"HDPE/GNP total/formula drift: {total}/{formulas}")
    if direct!=EXPECTED_DIRECT_MEASUREMENTS or derived!=EXPECTED_DERIVED_AVERAGES or process!=EXPECTED_PROCESS_INPUT_VALUES or identifiers!=EXPECTED_EXPERIMENT_IDENTIFIERS: raise RuntimeError("HDPE/GNP semantic role totals drifted")
    result={
      "schema":1,"status":"completed-public-measured-record-level-benchmark","retrievedDate":args.retrieved_date,
      "source":{"datasetId":"mendeley-4h98rz9f92-v3","datasetDoi":DOI,"version":VERSION,"license":"CC BY 4.0","companionArticleDoi":"10.1016/j.dib.2024.110987","publisherFileId":EXPECTED_FILE_ID,"publisherFileName":EXPECTED_FILE,"sha256":digest,"publisherSha256Matched":True,"retrievedSizeBytes":len(data),"resolvedUrl":final_url},
      "profile":{"process":"injection moulding","material":"HDPE with graphite nanoplatelets","experimentalRows":35,"numericColumns":22,"deliveredNumericCells":total,"roleCounts":role_counts,"directMeasuredPropertyValues":direct,"directMeasuredPropertyBreakdown":{"tensileModulus":175,"hardness":175,"toughness":175},"derivedAverageValuesExcluded":derived,"processInputValuesExcludedFromMeasuredOutcomeCount":process,"experimentIdentifiersExcluded":identifiers,"formulaCells":formulas,"acceptedMeasuredTimeSeriesSamples":0,"rawRowsOrCellValuesEmitted":False,"numericMeasurementValuesEmitted":False},
      "acceptance":{"countsAsFullyProfiledMeasuredDataset":True,"recordLevelMeasuredDataset":True,"acceptedMeasuredTimeSeriesSamples":0,"derivedAverageValuesExcluded":True,"modelAndClassificationFilesExcluded":True},
      "retrieval":{"rawPublisherFileCommitted":False,"rawRowsOrCellValuesUploadedAsArtifact":False},
      "evidenceBoundary":"The exact CC BY 4.0 Raw Data.xlsx workbook is SHA-256 matched. Across 35 injection-moulding experiments, 15 direct replicate outcome columns provide 525 measured mechanical-property values: 175 tensile-modulus, 175 hardness and 175 toughness measurements. Three 35-row average columns are derived summaries and excluded. GNP percentage, injection temperature and injection pressure are process inputs, experiment numbers are identifiers, and Random-Forest/model/classification/supporting files are excluded. This record-level dataset contributes zero to the high-frequency time-series sample metric."
    }
    out=Path(args.output); out.parent.mkdir(parents=True,exist_ok=True); out.write_text(json.dumps(result,indent=2,ensure_ascii=False)+"\n",encoding="utf-8")
    print(json.dumps({"status":result["status"],"experimentalRows":35,"deliveredNumericCells":total,"directMeasuredPropertyValues":direct,"derivedAverageValuesExcluded":derived,"acceptedMeasuredTimeSeriesSamples":0},indent=2))


if __name__=="__main__": main()
