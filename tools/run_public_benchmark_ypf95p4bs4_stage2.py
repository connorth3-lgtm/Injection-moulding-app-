#!/usr/bin/env python3
from __future__ import annotations
import argparse,hashlib,json,re,tempfile,urllib.request
from pathlib import Path
from openpyxl import load_workbook
ROOT=Path(__file__).resolve().parents[1];M=ROOT/'data/public-benchmark-results/ypf95p4bs4-stage1.json';DATASET='ypf95p4bs4';VERSION=1;UA='MouldMaster-Educational-Evidence-Profiler/1.0';API='https://api.data.mendeley.com'
def get(url):
    req=urllib.request.Request(url,headers={'User-Agent':UA,'Accept':'*/*'})
    with urllib.request.urlopen(req,timeout=240) as r:return r.read(),r.geturl()
def sanit(v):
    s=' '.join(str(v or '').replace('\x00',' ').split())
    if not s:return None
    s=re.sub(r'(?<![A-Za-z])[-+]?\d+(?:[.,]\d+)?(?:[Ee][-+]?\d+)?(?![A-Za-z])','<n>',s)
    return s[:220]
def file_url(fid):return f'{API}/datasets/{DATASET}/files/{fid}/file_downloaded?version={VERSION}'
def profile_book(data):
    with tempfile.TemporaryDirectory() as td:
        p=Path(td)/'source.xlsx';p.write_bytes(data);wb=load_workbook(p,data_only=False,read_only=False);sheets=[]
        for ws in wb.worksheets:
            labels=[];rows=[];numeric=formulas=texts=nonempty=0
            for row in ws.iter_rows():
                rnum=rform=rtext=rnon=0;rlabs=[]
                for c in row:
                    v=c.value
                    if v is None:continue
                    nonempty+=1;rnon+=1
                    if c.data_type=='f' or (isinstance(v,str) and v.startswith('=')):formulas+=1;rform+=1
                    elif isinstance(v,(int,float)) and not isinstance(v,bool):numeric+=1;rnum+=1
                    elif isinstance(v,str):
                        texts+=1;rtext+=1;s=sanit(v)
                        if s:labels.append(s);rlabs.append(s)
                if rnon:rows.append({'row':row[0].row,'nonEmptyCells':rnon,'numericLiteralCells':rnum,'formulaCells':rform,'textCells':rtext,'safeTextLabels':rlabs[:40]})
            uniq=[]
            for s in labels:
                if s not in uniq:uniq.append(s)
            low=' '.join(x.lower() for x in uniq)
            markers={k:[x for x in terms if x in low] for k,terms in {
              'operational':['production','producción','downtime','parada','availability','disponibilidad','machine','máquina','molde','mold','cambio','setup','tiempo','time','turno','shift'],
              'simulation':['arena','simulation','simulación','replication','replicación','entity','queue','cola','resource','recurso'],
              'maintenance':['tpm','mantenimiento','maintenance','falla','failure','avería'],
              'smed':['smed','cambio de molde','changeover','setup'],
              'quality':['rechazo','reject','defect','defecto','scrap','merma']}.items()}
            sheets.append({'sheet':ws.title,'maxRow':ws.max_row,'maxColumn':ws.max_column,'nonEmptyCells':nonempty,'numericLiteralCells':numeric,'formulaCells':formulas,'textCells':texts,'safeTextLabels':uniq[:160],'semanticMarkers':markers,'rows':rows,'rawNumericValuesEmitted':False})
        wb.close();return sheets
def main():
    ap=argparse.ArgumentParser();ap.add_argument('--output',type=Path,required=True);ap.add_argument('--retrieved-date',required=True);a=ap.parse_args();m=json.loads(M.read_text());cands=[x for x in m['manifest']['files'] if x['classification'].startswith('stage2-')];files=[]
    for x in cands:
        data,final=get(file_url(x['id']));digest=hashlib.sha256(data).hexdigest()
        if digest!=x['sha256']:raise RuntimeError(f'publisher SHA mismatch for {x["filename"]}')
        sheets=profile_book(data);files.append({'fileId':x['id'],'fileName':x['filename'],'classification':x['classification'],'sizeBytes':len(data),'sha256':digest,'publisherSha256Matched':True,'resolvedUrl':final,'sheetCount':len(sheets),'sheets':sheets,'rawPublisherFileCommitted':False})
    result={'schema':1,'status':'retrieved-operational-workbooks-needs-semantic-decision','retrievedDate':a.retrieved_date,'source':{'datasetId':'mendeley-ypf95p4bs4-v1','datasetDoi':'10.17632/ypf95p4bs4.1','license':'CC BY 4.0'},'files':files,'profile':{'xlsxFilesProfiled':len(files),'totalSheets':sum(f['sheetCount'] for f in files),'totalNumericLiteralCells':sum(s['numericLiteralCells'] for f in files for s in f['sheets']),'totalFormulaCells':sum(s['formulaCells'] for f in files for s in f['sheets']),'rawNumericValuesEmitted':False},'acceptance':{'countsAsFullyProfiledMeasuredDataset':False,'acceptedMeasuredTimeSeriesSamples':0,'acceptedRecordLevelMeasuredValues':0,'semanticDecisionRequired':True},'retrieval':{'doeFilesDownloaded':False,'rawPublisherFilesCommitted':False,'rawNumericValuesUploadedAsArtifact':False},'evidenceBoundary':'Only the two XLSX publisher files are retrieved. All .doe Arena model files remain untouched and non-counting. Aggregate workbook structure, sanitized text labels and numeric/formula counts are emitted; no raw numeric cell values are retained. Operational measurements must be distinguished from validation/simulation and formulas before acceptance.'}
    a.output.parent.mkdir(parents=True,exist_ok=True);a.output.write_text(json.dumps(result,indent=2,ensure_ascii=False)+'\n');print(json.dumps({'status':result['status'],'files':[(f['fileName'],f['sheetCount']) for f in files],'numericLiteralCells':result['profile']['totalNumericLiteralCells'],'formulaCells':result['profile']['totalFormulaCells']},indent=2))
if __name__=='__main__':main()
